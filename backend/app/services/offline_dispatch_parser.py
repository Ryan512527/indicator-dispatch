"""
线下派单处理情况 Parser
从"sheet：通报"中提取横山区的汇总指标
"""
import openpyxl
import re
from datetime import datetime
from sqlalchemy.orm import Session
from .models import OfflineDispatchSummary


def parse_offline_dispatch_file(file_path: str, district_filter: str = "横山") -> dict:
    """
    解析线下派单处理情况报表的"通报"sheet
    返回横山区的汇总指标
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)

    if "通报" not in wb.sheetnames:
        raise ValueError("报表中未找到'通报'sheet")

    ws = wb["通报"]

    # 提取通报日期（从Row 1）
    report_date = _extract_report_date(ws)

    # 找到横山区数据行
    hengshan_row = None
    for row in range(5, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=1).value
        if cell_value and district_filter in str(cell_value):
            hengshan_row = row
            break

    if not hengshan_row:
        raise ValueError(f"未找到{district_filter}的数据")

    # 提取指标（根据之前分析的结构）
    monthly_dispatch = ws.cell(row=hengshan_row, column=5).value  # [5] 月派单量
    overdue_backlog = ws.cell(row=hengshan_row, column=6).value  # [6] 超时积压
    not_overdue_backlog = ws.cell(row=hengshan_row, column=7).value  # [7] 未超时积压
    total_in_transit = ws.cell(row=hengshan_row, column=8).value  # [8] 累计在途
    warn_4h_overdue = ws.cell(row=hengshan_row, column=10).value  # [10] 预警4小时超时

    return {
        "report_date": report_date,
        "district": district_filter,
        "monthly_dispatch": str(monthly_dispatch) if monthly_dispatch is not None else "",
        "overdue_backlog": str(overdue_backlog) if overdue_backlog is not None else "",
        "not_overdue_backlog": str(not_overdue_backlog) if not_overdue_backlog is not None else "",
        "total_in_transit": str(total_in_transit) if total_in_transit is not None else "",
        "warn_4h_overdue": str(warn_4h_overdue) if warn_4h_overdue is not None else "",
    }


def _extract_report_date(ws) -> str:
    """从Row 1提取通报日期"""
    cell_value = ws.cell(row=1, column=1).value
    if not cell_value:
        return ""

    # 尝试提取日期，格式如："线下派单处理情况（装维调度） 截止6月5日   17:10"
    match = re.search(r'截止(\d+)月(\d+)日', str(cell_value))
    if match:
        month, day = match.groups()
        return f"2026-{month.zfill(2)}-{day.zfill(2)}"

    return str(cell_value)


def save_offline_dispatch_data(db: Session, data: dict):
    """保存线下派单数据到数据库（删除旧数据，插入新数据）"""
    # 删除旧数据
    db.query(OfflineDispatchSummary).delete()

    # 创建新记录
    summary = OfflineDispatchSummary(
        report_date=data.get("report_date"),
        district=data.get("district", "横山"),
        monthly_dispatch=data.get("monthly_dispatch"),
        overdue_backlog=data.get("overdue_backlog"),
        not_overdue_backlog=data.get("not_overdue_backlog"),
        total_in_transit=data.get("total_in_transit"),
        warn_4h_overdue=data.get("warn_4h_overdue"),
    )
    db.add(summary)
    db.commit()
