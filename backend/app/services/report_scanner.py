"""
报表扫描与通用解析服务
- 按关键词识别报表类型
- 出现 ≥3 次的报表类型视为"目标报表"
- 通用解析：读取 Excel 全部行列，存入 report_records 表
"""
import os
import re
import glob
import logging
import asyncio
from datetime import datetime, timezone, timedelta
from collections import defaultdict
from typing import Optional, List, Dict

import openpyxl
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, delete, func

from app.core.models import ReportType, ReportFile, ReportRecord
from app.core.config import settings

logger = logging.getLogger(__name__)


# ── 通知记录辅助函数 ──

# 通知去重时间窗口（秒）：同一 report_type+filename 在此窗口内只产生一条通知
_NOTIFICATION_DEDUP_SECONDS = 60

async def _add_notification(db: AsyncSession, report_type: str, filename: Optional[str]) -> None:
    """写入一条指标更新通知记录（含去重：同一文件同一类型在窗口期内不重复创建）"""
    if not filename:
        return
    try:
        from app.core.models import Notification
        # ── 去重检查：查找同类型同文件的最近未读通知 ──
        cutoff = datetime.now(timezone.utc) - timedelta(seconds=_NOTIFICATION_DEDUP_SECONDS)
        dup_query = select(Notification).where(
            Notification.report_type == report_type,
            Notification.filename == filename,
            Notification.is_read == False,
            Notification.event_time >= cutoff,
        ).order_by(Notification.event_time.desc()).limit(1)
        dup_result = await db.execute(dup_query)
        existing = dup_result.scalar_one_or_none()
        if existing:
            logger.debug(f"通知去重跳过 [{report_type}] {filename}（{existing.event_time} 已有未读通知）")
            return
        n = Notification(report_type=report_type, filename=filename)
        db.add(n)
        await db.commit()
    except Exception as e:
        logger.warning(f"通知记录失败 [{report_type}]: {e}")


def record_notification(report_type_name: str):
    """装饰器：在 reparse 成功后自动写入通知记录"""
    def decorator(func):
        async def wrapper(db: AsyncSession, directory: Optional[str] = None):
            result = await func(db, directory)
            if result and result.get("filename"):
                await _add_notification(db, report_type_name, result["filename"])
            return result
        return wrapper
    return decorator


# 报表类型关键词（按优先级匹配，长的优先）
REPORT_KEYWORDS: List[str] = [
    "家宽+FTTR遗留工单安装进度通报",
    "质差小区弱光工单处理完成率",
    "宽带在途投诉清单",
    "成功率攻坚通报",
    "质差客户整治完成率通报",
    "全市装维工作量统计",
    "投诉积压大于3单人员通报",
    "10086投诉积压(督办)",
    "企宽开通及时率通报",
    "企宽故障率",
    "企宽装机通报",
    "线下派单处理情况",
    "重投预警工单梳理",
    "家宽重投2次清单明细",
    "投诉三类工单在途情况",
    "H5当日闭环测评清单",
    "2200000及时率通报",
    "一二级分支真实处理通报",
    "长历时通报",
    "触点用后即评",
    "五类工单退撤单情况",
    "装机履约及时率",
    "一户一案",
    "企宽弱光通报",
    "企宽分支在途工单",
    "榆林未恢复故障统计",
    "接入层通报",
    "皮站故障清单",
    "无线退服清单",
    "日报",
]


def identify_report_type(filename: str) -> Optional[str]:
    """根据文件名识别报表类型，匹配最长关键词"""
    for keyword in REPORT_KEYWORDS:
        if keyword in filename:
            return keyword
    return None


def scan_target_reports(directory: str, min_occurrences: int = 3) -> List[Dict]:
    """
    递归扫描目录，按报表类型分组，返回出现次数 ≥ min_occurrences 的报表类型列表。
    每条记录：{"report_type": str, "file_count": int, "files": [filename, ...]}
    """
    patterns = (".xlsx", ".xls", ".csv")
    all_files: List[str] = []
    for root, _dirs, files in os.walk(directory):
        for f in files:
            if f.startswith("~$") or f.startswith("."):
                continue
            if f.lower().endswith(patterns):
                all_files.append(os.path.join(root, f))

    groups: Dict[str, List[str]] = defaultdict(list)
    for filepath in all_files:
        filename = os.path.basename(filepath)
        report_type = identify_report_type(filename)
        if report_type:
            groups[report_type].append(filepath)

    result: List[Dict] = []
    for report_type, files in groups.items():
        if len(files) >= min_occurrences:
            result.append({
                "report_type": report_type,
                "file_count": len(files),
                "files": [os.path.basename(f) for f in files],
            })

    return sorted(result, key=lambda x: -x["file_count"])


async def scan_and_register(directory: str, db: AsyncSession) -> List[Dict]:
    """
    扫描目录，识别目标报表类型，注册到 report_types 表。
    返回 [{"name": ..., "status": "registered"/"exists", "file_count": ...}, ...]
    """
    target_reports = scan_target_reports(directory)
    result = []

    for tr in target_reports:
        name = tr["report_type"]
        stmt = select(ReportType).where(ReportType.name == name)
        r = await db.execute(stmt)
        existing = r.scalar_one_or_none()

        if not existing:
            rt = ReportType(name=name, category="")
            db.add(rt)
            await db.flush()
            result.append({"name": name, "status": "registered", "file_count": tr["file_count"]})
        else:
            result.append({"name": name, "status": "exists", "file_count": tr["file_count"]})

    await db.commit()
    return result


def _parse_excel_sync(file_path: str) -> List[Dict]:
    """
    同步解析 Excel 文件，返回 [row_dict, ...]。
    在线程池中调用，避免阻塞事件循环。
    """
    ext = os.path.splitext(file_path)[1].lower()
    rows_data: List[tuple] = []

    if ext == ".xlsx":
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        rows_data = list(ws.iter_rows(values_only=True))
        wb.close()
    else:
        # 尝试用 openpyxl 读取 .xls（部分兼容）
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active
            rows_data = list(ws.iter_rows(values_only=True))
            wb.close()
        except Exception:
            raise ValueError(f"不支持的文件格式: {ext}，仅支持 .xlsx")

    if not rows_data or len(rows_data) < 2:
        return []

    # 第一行作为表头
    raw_headers = rows_data[0]
    headers: List[str] = []
    for i, h in enumerate(raw_headers):
        if h is not None:
            headers.append(str(h).strip())
        else:
            # 查找最近的非空表头
            headers.append(f"col_{i}")

    data: List[Dict] = []
    for row in rows_data[1:]:
        row_dict: Dict[str, str] = {}
        for i, cell in enumerate(row):
            header = headers[i] if i < len(headers) else f"col_{i}"
            if cell is None:
                row_dict[header] = ""
            elif isinstance(cell, datetime):
                row_dict[header] = cell.isoformat()
            else:
                row_dict[header] = str(cell)
        data.append(row_dict)

    return data


async def parse_report_type(
    report_type_name: str,
    db: AsyncSession,
    directory: Optional[str] = None,
    update_column_hint: bool = True,
) -> Dict:
    """
    解析指定报表类型的所有文件，入库。
    返回 {"total_records": int, "files_parsed": int, "files_skipped": int}
    """
    if directory is None:
        directory = settings.watch_dir

    # 收集匹配文件（递归）
    all_files: List[str] = []
    for root, _dirs, files in os.walk(directory):
        for f in files:
            if f.startswith("~$") or f.startswith("."):
                continue
            if f.lower().endswith((".xlsx", ".xls", ".csv")):
                all_files.append(os.path.join(root, f))
    matching_files = [f for f in all_files if report_type_name in os.path.basename(f)]

    # 获取或创建 ReportType
    stmt = select(ReportType).where(ReportType.name == report_type_name)
    r = await db.execute(stmt)
    report_type = r.scalar_one_or_none()
    if not report_type:
        report_type = ReportType(name=report_type_name, category="")
        db.add(report_type)
        await db.flush()

    total_records = 0
    files_parsed = 0
    files_skipped = 0
    all_columns: set = set()

    for filepath in matching_files:
        filename = os.path.basename(filepath)
        # 检查是否已解析
        stmt2 = select(ReportFile).where(ReportFile.filename == filename)
        r2 = await db.execute(stmt2)
        existing = r2.scalar_one_or_none()

        if existing and existing.parse_status == "parsed":
            logger.info(f"跳过已解析文件: {filename}")
            files_skipped += 1
            continue

        try:
            rows = await asyncio.to_thread(_parse_excel_sync, filepath)
            if not rows:
                files_skipped += 1
                continue

            # 创建/更新 ReportFile
            if existing:
                report_file = existing
            else:
                report_file = ReportFile(
                    report_type_id=report_type.id,
                    filename=filename,
                    file_path=filepath,
                )
                db.add(report_file)
            await db.flush()

            # 幂等：删除旧记录
            await db.execute(delete(ReportRecord).where(ReportRecord.report_file_id == report_file.id))

            # 写入新记录
            for row in rows:
                db.add(ReportRecord(report_file_id=report_file.id, data=row))
                all_columns.update(row.keys())

            report_file.record_count = len(rows)
            report_file.parse_status = "parsed"
            report_file.parse_error = None
            total_records += len(rows)
            files_parsed += 1
            logger.info(f"已解析 {filename}: {len(rows)} 条记录")

        except Exception as e:
            logger.error(f"解析失败 {filename}: {e}")
            if existing is None:
                report_file = ReportFile(
                    report_type_id=report_type.id,
                    filename=filename,
                    file_path=filepath,
                    parse_status="failed",
                    parse_error=str(e),
                )
                db.add(report_file)
            else:
                existing.parse_status = "failed"
                existing.parse_error = str(e)
            files_skipped += 1

    # 更新 column_hint
    if update_column_hint and all_columns:
        report_type.column_hint = list(all_columns)
        report_type.updated_at = datetime.now(timezone.utc)

    await db.commit()
    return {
        "total_records": total_records,
        "files_parsed": files_parsed,
        "files_skipped": files_skipped,
    }


async def get_report_types(db: AsyncSession) -> List[Dict]:
    """获取所有报表类型及统计信息"""
    stmt = select(ReportType).order_by(ReportType.id)
    r = await db.execute(stmt)
    types = r.scalars().all()

    result = []
    for rt in types:
        # 统计文件数和记录数
        f_stmt = select(func.count(ReportFile.id)).where(
            ReportFile.report_type_id == rt.id,
            ReportFile.parse_status == "parsed",
        )
        r2 = await db.execute(f_stmt)
        file_count = r2.scalar() or 0

        r_stmt = (
            select(func.count(ReportRecord.id))
            .join(ReportFile)
            .where(ReportFile.report_type_id == rt.id)
        )
        r3 = await db.execute(r_stmt)
        record_count = r3.scalar() or 0

        # 最新文件时间 & 预览数据
        t_stmt = select(ReportFile).where(
            ReportFile.report_type_id == rt.id,
            ReportFile.parse_status == "parsed",
        ).order_by(ReportFile.created_at.desc()).limit(1)
        r4 = await db.execute(t_stmt)
        latest_file = r4.scalar_one_or_none()
        latest_time = latest_file.created_at if latest_file else None
        latest_filename = latest_file.filename if latest_file else None

        # 最新文件的前5条记录作为预览
        latest_preview: List[Dict] = []
        if latest_file:
            p_stmt = select(ReportRecord.data).where(
                ReportRecord.report_file_id == latest_file.id
            ).limit(5)
            r5 = await db.execute(p_stmt)
            latest_preview = [dict(row[0]) for row in r5.all()]

        result.append({
            "id": rt.id,
            "name": rt.name,
            "category": rt.category,
            "column_hint": rt.column_hint or [],
            "file_count": file_count,
            "record_count": record_count,
            "latest_time": latest_time.isoformat() if latest_time else None,
            "latest_filename": latest_filename,
            "latest_preview": latest_preview,
            "created_at": rt.created_at.isoformat() if rt.created_at else None,
        })

    return result


# ── 无线退服横山数据专用处理 ──

# 无线退服清单保留字段（仅提取这9个字段）
WIRELESS_OUTAGE_FIELDS = [
    "基站类型",      # station_type
    "站址名称",      # site_name
    "告警名称",      # alarm_name
    "告警时间",      # alarm_time
    "退服时长(h)",   # outage_duration_hours
    "保障场景",      # guarantee_scenario
    "是否超时",      # is_timeout
    "是否塔维",      # is_tower_maintenance
    "机房名称",      # room_name
]

# 英文到中文映射（用于 API 返回，保留中文原名）
WIRELESS_OUTAGE_FIELD_MAP = {
    "基站类型": "基站类型",
    "站址名称": "站址名称",
    "告警名称": "告警名称",
    "告警时间": "告警时间",
    "退服时长(h)": "退服时长(h)",
    "保障场景": "保障场景",
    "是否超时": "是否超时",
    "是否塔维": "是否塔维",
    "机房名称": "机房名称",
}


# 无线退服: parser 英文字段 → 展示中文名称的映射
WIRELESS_OUTAGE_EN_TO_CN = {
    "station_type": "基站类型",
    "site_name": "站址名称",
    "alarm_name": "告警名称",
    "alarm_time": "告警时间",
    "outage_duration_hours": "退服时长(h)",
    "guarantee_scenario": "保障场景",
    "is_timeout": "是否超时",
    "is_tower_maintenance": "是否塔维",
    "room_name": "机房名称",
}


def _parse_wireless_outage_files(directory: str) -> List[Dict]:
    """
    解析所有"无线退服清单"文件，使用 parser 模块正确检测表头，
    仅保留 县区=="横山" 且 9 个指定字段的记录。
    返回 [{"filename": str, "records": [dict]}, ...]
    """
    from app.parser.service import parse_file as parser_parse_file
    all_results: List[Dict] = []

    # 递归查找匹配文件
    matching: List[str] = []
    for root, _dirs, files in os.walk(directory):
        for f in files:
            if f.startswith("~$") or f.startswith("."):
                continue
            if "无线退服清单" in f and f.lower().endswith((".xlsx", ".xls")):
                matching.append(os.path.join(root, f))

    # 按文件修改时间排序（旧的先处理，新的后处理 → 新文件 id 更大）
    matching.sort(key=lambda f: os.path.getmtime(f))

    for filepath in matching:
        filename = os.path.basename(filepath)
        try:
            # 使用 parser 模块正确解析（会检测表头行并做中英映射）
            rows = parser_parse_file(filepath)
            if not rows:
                continue

            filtered: List[Dict] = []
            for row in rows:
                # parser 返回的是英文字段名，district 对应 "县区"
                district = row.get("district", "")
                if district != "横山":
                    continue

                # 只提取 9 个指定字段，映射为中文名称
                clean: Dict[str, str] = {}
                for en_key, cn_key in WIRELESS_OUTAGE_EN_TO_CN.items():
                    val = row.get(en_key, "")
                    clean[cn_key] = str(val) if val is not None else ""
                filtered.append(clean)

            all_results.append({
                "filename": filename,
                "records": filtered,
            })
            if filtered:
                logger.info(f"无线退服横山: {filename} -> {len(filtered)} 条横山记录")
            else:
                logger.info(f"无线退服横山: {filename} -> 0 条横山记录（最新）")

        except Exception as e:
            logger.error(f"解析无线退服文件失败 {filename}: {e}")

    return all_results


@record_notification("无线退服清单")
async def reparse_wireless_outage(db: AsyncSession, directory: Optional[str] = None) -> Dict:
    """
    重新解析无线退服清单文件，仅保留横山区 9 个字段，更新数据库。
    删除旧的无线退服 report_records 并写入新数据。
    """
    if directory is None:
        directory = settings.watch_dir

    # 获取或创建"无线退服清单"报表类型
    stmt = select(ReportType).where(ReportType.name == "无线退服清单")
    r = await db.execute(stmt)
    report_type = r.scalar_one_or_none()
    if not report_type:
        report_type = ReportType(name="无线退服清单", category="无线")
        db.add(report_type)
        await db.flush()

    # 删除旧的无线退服记录（通过 report_files 关联删除 report_records）
    old_files_stmt = select(ReportFile.id).where(ReportFile.report_type_id == report_type.id)
    r2 = await db.execute(old_files_stmt)
    old_file_ids = [row[0] for row in r2.all()]
    if old_file_ids:
        from sqlalchemy import delete as _delete
        await db.execute(_delete(ReportRecord).where(ReportRecord.report_file_id.in_(old_file_ids)))
        await db.execute(_delete(ReportFile).where(ReportFile.report_type_id == report_type.id))

    # 在线程池中解析
    file_results = await asyncio.to_thread(_parse_wireless_outage_files, directory)

    total_records = 0
    files_parsed = 0
    files_skipped = 0
    all_columns = set(WIRELESS_OUTAGE_FIELDS)

    for fr in file_results:
        filename = fr["filename"]
        records = fr["records"]

        report_file = ReportFile(
            report_type_id=report_type.id,
            filename=filename,
            file_path=os.path.join(directory, filename),
            parse_status="parsed",
            record_count=len(records),
        )
        db.add(report_file)
        await db.flush()

        for row in records:
            db.add(ReportRecord(report_file_id=report_file.id, data=row))

        total_records += len(records)
        if records:
            files_parsed += 1
        else:
            files_skipped += 1

    # 更新 column_hint
    report_type.column_hint = list(all_columns)
    report_type.updated_at = datetime.now(timezone.utc)

    await db.commit()

    return {
        "total_records": total_records,
        "files_parsed": files_parsed,
        "files_skipped": files_skipped,
    }


async def get_wireless_outage_summary(db: AsyncSession) -> Dict:
    """获取无线退服横山数据概要：仅最新一份文件的退服数 + 告警名称列表"""
    stmt = select(ReportType).where(ReportType.name == "无线退服清单")
    r = await db.execute(stmt)
    report_type = r.scalar_one_or_none()
    if not report_type:
        return {"total": 0, "alarm_names": [], "latest_time": None, "latest_filename": None}

    # 找到最新的 ReportFile（按 id 降序，id 大的 = 最新入库）
    latest_file_stmt = (
        select(ReportFile.id, ReportFile.filename, ReportFile.record_count)
        .where(ReportFile.report_type_id == report_type.id)
        .order_by(ReportFile.id.desc())
        .limit(1)
    )
    r2 = await db.execute(latest_file_stmt)
    latest_row = r2.first()
    if not latest_row:
        return {"total": 0, "alarm_names": [], "latest_time": None, "latest_filename": None}

    latest_file_id, latest_filename, record_count = latest_row

    # 如果最新文件记录数为 0，直接返回
    if record_count == 0:
        return {"total": 0, "alarm_names": [], "latest_time": None, "latest_filename": latest_filename}

    # 仅查询最新文件的记录
    data_stmt = (
        select(ReportRecord.data)
        .where(ReportRecord.report_file_id == latest_file_id)
        .order_by(ReportRecord.id.desc())
    )
    r3 = await db.execute(data_stmt)
    rows = r3.all()

    records = [dict(row[0]) for row in rows]

    alarm_names: List[str] = []
    seen_names = set()
    for rec in records:
        name = rec.get("告警名称", "")
        if name and name not in seen_names:
            alarm_names.append(name)
            seen_names.add(name)

    latest_time = records[0].get("告警时间", "") if records else None

    return {
        "total": len(records),
        "alarm_names": alarm_names,
        "latest_time": latest_time,
        "latest_filename": latest_filename,
    }


async def get_wireless_outage_detail(
    db: AsyncSession,
    page: int = 1,
    page_size: int = 50,
) -> Dict:
    """分页获取无线退服横山详细数据：仅最新一份文件的数据（9个字段）"""
    stmt = select(ReportType).where(ReportType.name == "无线退服清单")
    r = await db.execute(stmt)
    report_type = r.scalar_one_or_none()
    if not report_type:
        return {"records": [], "total": 0, "page": page, "page_size": page_size, "latest_filename": None}

    # 找到最新的 ReportFile
    latest_file_stmt = (
        select(ReportFile.id, ReportFile.filename, ReportFile.record_count)
        .where(ReportFile.report_type_id == report_type.id)
        .order_by(ReportFile.id.desc())
        .limit(1)
    )
    r2 = await db.execute(latest_file_stmt)
    latest_row = r2.first()
    if not latest_row:
        return {"records": [], "total": 0, "page": page, "page_size": page_size, "latest_filename": None}

    latest_file_id, latest_filename, record_count = latest_row

    if record_count == 0:
        return {"records": [], "total": 0, "page": page, "page_size": page_size, "latest_filename": latest_filename}

    # 仅查询最新文件的记录
    data_stmt = (
        select(ReportRecord.data, ReportFile.filename, ReportRecord.created_at)
        .join(ReportFile, ReportRecord.report_file_id == ReportFile.id)
        .where(ReportRecord.report_file_id == latest_file_id)
        .order_by(ReportRecord.id.desc())
    )
    r3 = await db.execute(data_stmt)
    rows = r3.all()

    all_records = []
    for row in rows:
        data, filename, created_at = row
        record = dict(data)
        record["_source_file"] = filename
        record["_created_at"] = created_at.isoformat() if created_at else ""
        all_records.append(record)

    total = len(all_records)

    # 分页
    offset = (page - 1) * page_size
    records = all_records[offset:offset + page_size]

    return {
        "records": records,
        "total": total,
        "page": page,
        "page_size": page_size,
        "latest_filename": latest_filename,
    }


async def get_wireless_outage_trend(db: AsyncSession, hours: int = 48) -> List[Dict]:
    """
    获取最近 N 小时无线退服横山数量趋势（按告警发生时间聚合）。
    去重：同一 (站址名称, 告警名称, 告警时间) 只计一次。
    返回 [{"hour": "2026-06-05T10:00", "count": 3}, ...]
    """
    from collections import Counter
    from datetime import timedelta, timezone as _tz

    stmt = select(ReportType).where(ReportType.name == "无线退服清单")
    r = await db.execute(stmt)
    report_type = r.scalar_one_or_none()
    if not report_type:
        return []

    j = ReportRecord.__table__.join(ReportFile.__table__,
        ReportRecord.report_file_id == ReportFile.id)

    data_stmt = (
        select(ReportRecord.data)
        .select_from(j)
        .where(ReportFile.report_type_id == report_type.id)
    )
    r2 = await db.execute(data_stmt)
    rows = r2.all()

    # 北京时间 UTC+8
    beijing_tz = _tz(timedelta(hours=8))
    now_bj = datetime.now(beijing_tz)
    cutoff = now_bj.replace(minute=0, second=0, microsecond=0)

    hour_counter: Counter = Counter()
    seen_alarms: set = set()

    for row in rows:
        data = dict(row[0])
        alarm_time_str = data.get("告警时间", "")
        site_name = data.get("站址名称", "")
        alarm_name = data.get("告警名称", "")

        if not alarm_time_str:
            continue

        # 去重：同一站点+同一告警+同一时间只计一次
        dedup_key = (site_name, alarm_name, alarm_time_str)
        if dedup_key in seen_alarms:
            continue
        seen_alarms.add(dedup_key)

        try:
            # 解析北京时间字符串
            alarm_dt = None
            for fmt_str in [
                "%Y-%m-%d %H:%M:%S",
                "%Y-%m-%d %H:%M",
                "%Y/%m/%d %H:%M:%S",
                "%Y/%m/%d %H:%M",
            ]:
                try:
                    alarm_dt = datetime.strptime(alarm_time_str[:19], fmt_str[:len(alarm_time_str[:19])])
                    break
                except ValueError:
                    continue

            if alarm_dt is None:
                continue

            # 添加北京时区使其与 cutoff 对齐
            alarm_dt = alarm_dt.replace(tzinfo=beijing_tz)

            # 按小时聚合（北京时间）
            hour_key = alarm_dt.replace(minute=0, second=0, microsecond=0)
            hour_counter[hour_key] += 1

        except Exception:
            continue

    # 生成最近 hours 小时的时间序列（北京时间）
    trend = []
    for i in range(hours - 1, -1, -1):
        actual_slot = cutoff - timedelta(hours=i)
        count = hour_counter.get(actual_slot, 0)
        trend.append({
            "hour": actual_slot.isoformat(),
            "count": count,
        })

    return trend


# ── 皮站故障横山数据专用处理 ──

# 皮站故障清单保留字段（仅提取这5个字段）
PISITE_FAULT_FIELDS = [
    "网络类型",      # network_type
    "基站名称",      # station_name
    "网管状态",      # nms_status
    "设备厂商",      # vendor
    "设备类型",      # device_type
]

# 英文到中文映射
PISITE_FAULT_EN_TO_CN = {
    "network_type": "网络类型",
    "station_name": "基站名称",
    "nms_status": "网管状态",
    "vendor": "设备厂商",
    "device_type": "设备类型",
}


def _parse_pisite_fault_files(directory: str) -> list[dict]:
    """
    解析所有"皮站故障清单"文件，使用 parser 模块正确检测表头，
    仅保留 县区=="横山" 且 5 个指定字段的记录。
    返回 [{"filename": str, "records": [dict]}, ...]
    """
    from app.parser.service import parse_file as parser_parse_file
    all_results: list[dict] = []

    # 递归查找匹配文件
    matching: list[str] = []
    for root, _dirs, files in os.walk(directory):
        for f in files:
            if f.startswith("~$") or f.startswith("."):
                continue
            if "皮站故障清单" in f and f.lower().endswith((".xlsx", ".xls")):
                matching.append(os.path.join(root, f))

    # 按文件修改时间排序（旧的先处理，新的后处理 → 新文件 id 更大）
    matching.sort(key=lambda f: os.path.getmtime(f))

    for filepath in matching:
        filename = os.path.basename(filepath)
        try:
            # 使用 parser 模块正确解析（会检测表头行并做中英映射）
            rows = parser_parse_file(filepath)
            if not rows:
                continue

            filtered: list[dict] = []
            for row in rows:
                # parser 返回的是英文字段名，district 对应 "县区"
                district = row.get("district", "")
                if district != "横山":
                    continue

                # 只提取 5 个指定字段，映射为中文名称
                clean: dict[str, str] = {}
                for en_key, cn_key in PISITE_FAULT_EN_TO_CN.items():
                    val = row.get(en_key, "")
                    clean[cn_key] = str(val) if val is not None else ""
                filtered.append(clean)

            all_results.append({
                "filename": filename,
                "records": filtered,
            })
            if filtered:
                logger.info(f"皮站故障横山: {filename} -> {len(filtered)} 条横山记录")
            else:
                logger.info(f"皮站故障横山: {filename} -> 0 条横山记录")

        except Exception as e:
            logger.error(f"解析皮站故障文件失败 {filename}: {e}")

    return all_results


@record_notification("皮站故障清单")
async def reparse_pisite_fault(db: AsyncSession, directory: Optional[str] = None) -> dict:
    """
    重新解析皮站故障清单文件，仅保留横山区 5 个字段，更新数据库。
    删除旧的皮站故障 report_records 并写入新数据。
    """
    if directory is None:
        directory = settings.watch_dir

    # 获取或创建"皮站故障清单"报表类型
    stmt = select(ReportType).where(ReportType.name == "皮站故障清单")
    r = await db.execute(stmt)
    report_type = r.scalar_one_or_none()
    if not report_type:
        report_type = ReportType(name="皮站故障清单", category="故障")
        db.add(report_type)
        await db.flush()

    # 删除旧的皮站故障记录（通过 report_files 关联删除 report_records）
    old_files_stmt = select(ReportFile.id).where(ReportFile.report_type_id == report_type.id)
    r2 = await db.execute(old_files_stmt)
    old_file_ids = [row[0] for row in r2.all()]
    if old_file_ids:
        from sqlalchemy import delete as _delete
        await db.execute(_delete(ReportRecord).where(ReportRecord.report_file_id.in_(old_file_ids)))
        await db.execute(_delete(ReportFile).where(ReportFile.report_type_id == report_type.id))

    # 在线程池中解析
    file_results = await asyncio.to_thread(_parse_pisite_fault_files, directory)

    total_records = 0
    files_parsed = 0
    files_skipped = 0
    all_columns = set(PISITE_FAULT_FIELDS)

    for fr in file_results:
        filename = fr["filename"]
        records = fr["records"]

        report_file = ReportFile(
            report_type_id=report_type.id,
            filename=filename,
            file_path=os.path.join(directory, filename),
            parse_status="parsed",
            record_count=len(records),
        )
        db.add(report_file)
        await db.flush()

        for row in records:
            db.add(ReportRecord(report_file_id=report_file.id, data=row))

        total_records += len(records)
        if records:
            files_parsed += 1
        else:
            files_skipped += 1

    # 更新 column_hint
    report_type.column_hint = list(all_columns)
    report_type.updated_at = datetime.now(timezone.utc)

    await db.commit()

    return {
        "total_records": total_records,
        "files_parsed": files_parsed,
        "files_skipped": files_skipped,
    }


async def get_pisite_fault_summary(db: AsyncSession) -> dict:
    """获取皮站故障横山数据概要：仅最新一份文件的故障总数 + 设备厂商列表"""
    stmt = select(ReportType).where(ReportType.name == "皮站故障清单")
    r = await db.execute(stmt)
    report_type = r.scalar_one_or_none()
    if not report_type:
        return {"total": 0, "vendors": [], "latest_time": None, "latest_filename": None}

    # 找到最新的 ReportFile（按 id 降序，id 大的 = 最新入库）
    latest_file_stmt = (
        select(ReportFile.id, ReportFile.filename, ReportFile.record_count)
        .where(ReportFile.report_type_id == report_type.id)
        .order_by(ReportFile.id.desc())
        .limit(1)
    )
    r2 = await db.execute(latest_file_stmt)
    latest_row = r2.first()
    if not latest_row:
        return {"total": 0, "vendors": [], "latest_time": None, "latest_filename": None}

    latest_file_id, latest_filename, record_count = latest_row

    if record_count == 0:
        return {"total": 0, "vendors": [], "latest_time": None, "latest_filename": latest_filename}

    # 仅查询最新文件的记录
    data_stmt = (
        select(ReportRecord.data)
        .where(ReportRecord.report_file_id == latest_file_id)
        .order_by(ReportRecord.id.desc())
    )
    r3 = await db.execute(data_stmt)
    rows = r3.all()

    records = [dict(row[0]) for row in rows]

    # 统计设备厂商（去重）
    vendors: list[str] = []
    seen_vendors = set()
    for rec in records:
        vendor = rec.get("设备厂商", "")
        if vendor and vendor not in seen_vendors:
            vendors.append(vendor)
            seen_vendors.add(vendor)

    return {
        "total": len(records),
        "vendors": vendors,
        "latest_time": None,  # 皮站故障清单没有时间字段
        "latest_filename": latest_filename,
    }


async def get_pisite_fault_detail(
    db: AsyncSession,
    page: int = 1,
    page_size: int = 50,
) -> dict:
    """分页获取皮站故障横山详细数据：仅最新一份文件的5个字段数据"""
    stmt = select(ReportType).where(ReportType.name == "皮站故障清单")
    r = await db.execute(stmt)
    report_type = r.scalar_one_or_none()
    if not report_type:
        return {"records": [], "total": 0, "page": page, "page_size": page_size, "latest_filename": None}

    # 找到最新的 ReportFile
    latest_file_stmt = (
        select(ReportFile.id, ReportFile.filename, ReportFile.record_count)
        .where(ReportFile.report_type_id == report_type.id)
        .order_by(ReportFile.id.desc())
        .limit(1)
    )
    r2 = await db.execute(latest_file_stmt)
    latest_row = r2.first()
    if not latest_row:
        return {"records": [], "total": 0, "page": page, "page_size": page_size, "latest_filename": None}

    latest_file_id, latest_filename, record_count = latest_row

    if record_count == 0:
        return {"records": [], "total": 0, "page": page, "page_size": page_size, "latest_filename": latest_filename}

    # 仅查询最新文件的记录
    data_stmt = (
        select(ReportRecord.data, ReportFile.filename, ReportRecord.created_at)
        .join(ReportFile, ReportRecord.report_file_id == ReportFile.id)
        .where(ReportRecord.report_file_id == latest_file_id)
        .order_by(ReportRecord.id.desc())
    )
    r3 = await db.execute(data_stmt)
    rows = r3.all()

    all_records = []
    for row in rows:
        data, filename, created_at = row
        record = dict(data)
        record["_source_file"] = filename
        record["_created_at"] = created_at.isoformat() if created_at else ""
        all_records.append(record)

    total = len(all_records)

    # 分页
    offset = (page - 1) * page_size
    records = all_records[offset:offset + page_size]

    return {
        "records": records,
        "total": total,
        "page": page,
        "page_size": page_size,
        "latest_filename": latest_filename,
    }


# ── 接入层通报横山数据专用处理 ──

# 接入层通报保留字段（仅提取这6个字段）
ACCESS_LAYER_FAULT_FIELDS = [
    "接入层断纤链路",   # fiber_break_link
    "告警码名称",       # alarm_code_name
    "发生时间",         # occurrence_time
    "具体原因",         # specific_reason
    "是否影响业务",     # business_affected
    "故障历时",         # fault_duration
]

# 英文到中文映射
ACCESS_LAYER_FAULT_EN_TO_CN = {
    "fiber_break_link": "接入层断纤链路",
    "alarm_code_name": "告警码名称",
    "occurrence_time": "发生时间",
    "specific_reason": "具体原因",
    "business_affected": "是否影响业务",
    "fault_duration": "故障历时",
}


def _extract_date_from_filename(filename: str) -> Optional[str]:
    """从文件名中提取日期，用于排序确保最新文件在后。

    支持格式: 2026-06-04, 20260604, 2026年06月04日 等。
    返回 ISO 日期字符串 "YYYY-MM-DD"，失败返回 None。
    """
    patterns = [
        r"(\d{4})[-/年](\d{1,2})[-/月](\d{1,2})",
        r"(\d{4})(\d{2})(\d{2})",
    ]
    for pat in patterns:
        m = re.search(pat, filename)
        if m:
            y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
            if 2020 <= y <= 2030 and 1 <= mo <= 12 and 1 <= d <= 31:
                return f"{y:04d}-{mo:02d}-{d:02d}"
    return None


# 接入层通报列关键字 → 英文字段名映射（用于包含匹配）
_ACCESS_LAYER_KEYWORD_MAP: list[tuple[str, str]] = [
    ("接入层断纤链路", "fiber_break_link"),
    ("告警码名称", "alarm_code_name"),
    ("发生时间", "occurrence_time"),
    ("具体原因", "specific_reason"),
    ("是否影响业务", "business_affected"),
    ("故障历时", "fault_duration"),
    ("县区", "district"),
    ("责任人", "responsible_person"),
]


def _parse_access_layer_fault_files(directory: str) -> list[dict]:
    """
    直接使用 openpyxl 解析所有"接入层通报"文件。
    关键字包含匹配表头 → 提取字段 → 过滤县区=="横山"。
    返回 [{"filename": str, "records": [dict]}, ...]
    """
    all_results: list[dict] = []

    # 递归查找匹配文件
    matching: list[str] = []
    for root, _dirs, files in os.walk(directory):
        for f in files:
            if f.startswith("~$") or f.startswith("."):
                continue
            if "接入层通报" in f and f.lower().endswith((".xlsx", ".xls")):
                matching.append(os.path.join(root, f))

    # 排序：优先按文件名中的日期，回退到 mtime（旧的先处理，新的后处理 → 新文件 id 更大）
    def _sort_key(filepath: str) -> str:
        fname = os.path.basename(filepath)
        date_str = _extract_date_from_filename(fname)
        if date_str:
            return date_str
        # 回退：用 mtime 生成日期
        mtime = os.path.getmtime(filepath)
        dt = datetime.fromtimestamp(mtime, tz=timezone.utc)
        return dt.strftime("%Y-%m-%d")

    matching.sort(key=_sort_key)

    for filepath in matching:
        filename = os.path.basename(filepath)
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
            ws = wb.active

            # ── 步骤1：检测表头行 ──
            header_row_idx: Optional[int] = None
            col_map: dict[int, str] = {}  # 列索引 → 英文字段名

            for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                matched = 0
                temp_map: dict[int, str] = {}
                for col_idx, cell in enumerate(row):
                    if cell is None:
                        continue
                    cell_str = str(cell).strip()
                    if not cell_str:
                        continue
                    # 关键字包含匹配：表头中"接入层断纤链路清单（2026-06-04）"包含"接入层断纤链路"
                    for kw, en_key in _ACCESS_LAYER_KEYWORD_MAP:
                        if len(kw) >= 2 and kw in cell_str:
                            temp_map[col_idx] = en_key
                            matched += 1
                            break

                if matched >= 4:  # 至少匹配 4 个关键字才认定是表头行（含 县区 共需 8 个）
                    header_row_idx = row_idx
                    col_map = temp_map
                    break

            if header_row_idx is None:
                logger.warning(f"接入层通报: {filename} 未检测到表头行")
                wb.close()
                continue

            logger.info(
                f"接入层通报: {filename} 表头行={header_row_idx}, 匹配列={list(col_map.values())}"
            )

            # ── 步骤2：提取数据行 ──
            filtered: list[dict] = []
            for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                if row_idx <= header_row_idx:
                    continue

                # 构建英文字段记录
                record: dict[str, str] = {}
                for col_idx, en_key in col_map.items():
                    val = ""
                    if col_idx < len(row) and row[col_idx] is not None:
                        val = str(row[col_idx]).strip()
                    record[en_key] = val

                # 过滤：县区 == 横山
                district = record.get("district", "")
                if district != "横山":
                    continue

                # 只保留 6 个指定字段，映射为中文名称
                clean: dict[str, str] = {}
                for en_key, cn_key in ACCESS_LAYER_FAULT_EN_TO_CN.items():
                    val = record.get(en_key, "")
                    clean[cn_key] = str(val) if val else ""
                filtered.append(clean)

            wb.close()

            all_results.append({
                "filename": filename,
                "records": filtered,
            })
            if filtered:
                logger.info(f"接入层通报横山: {filename} -> {len(filtered)} 条横山记录")
            else:
                logger.info(f"接入层通报横山: {filename} -> 0 条横山记录")

        except Exception as e:
            logger.error(f"解析接入层通报文件失败 {filename}: {e}", exc_info=True)

    return all_results


@record_notification("接入层通报")
async def reparse_access_layer_fault(db: AsyncSession, directory: Optional[str] = None) -> dict:
    """
    重新解析接入层通报文件，仅保留横山区 6 个字段，更新数据库。
    删除旧的接入层通报 report_records 并写入新数据。
    """
    if directory is None:
        directory = settings.watch_dir

    # 获取或创建"接入层通报"报表类型
    stmt = select(ReportType).where(ReportType.name == "接入层通报")
    r = await db.execute(stmt)
    report_type = r.scalar_one_or_none()
    if not report_type:
        report_type = ReportType(name="接入层通报", category="故障")
        db.add(report_type)
        await db.flush()

    # 删除旧的接入层通报记录（通过 report_files 关联删除 report_records）
    old_files_stmt = select(ReportFile.id).where(ReportFile.report_type_id == report_type.id)
    r2 = await db.execute(old_files_stmt)
    old_file_ids = [row[0] for row in r2.all()]
    if old_file_ids:
        from sqlalchemy import delete as _delete
        await db.execute(_delete(ReportRecord).where(ReportRecord.report_file_id.in_(old_file_ids)))
        await db.execute(_delete(ReportFile).where(ReportFile.report_type_id == report_type.id))

    # 在线程池中解析
    file_results = await asyncio.to_thread(_parse_access_layer_fault_files, directory)

    total_records = 0
    files_parsed = 0
    files_skipped = 0
    all_columns = set(ACCESS_LAYER_FAULT_FIELDS)

    for fr in file_results:
        filename = fr["filename"]
        records = fr["records"]

        report_file = ReportFile(
            report_type_id=report_type.id,
            filename=filename,
            file_path=os.path.join(directory, filename),
            parse_status="parsed",
            record_count=len(records),
        )
        db.add(report_file)
        await db.flush()

        for row in records:
            db.add(ReportRecord(report_file_id=report_file.id, data=row))

        total_records += len(records)
        if records:
            files_parsed += 1
        else:
            files_skipped += 1

    # 更新 column_hint
    report_type.column_hint = list(all_columns)
    report_type.updated_at = datetime.now(timezone.utc)

    await db.commit()

    return {
        "total_records": total_records,
        "files_parsed": files_parsed,
        "files_skipped": files_skipped,
    }


async def get_access_layer_fault_summary(db: AsyncSession) -> dict:
    """获取接入层通报横山数据概要：故障总数 + 影响业务数 + 不影响业务数 + 告警码名称列表"""
    stmt = select(ReportType).where(ReportType.name == "接入层通报")
    r = await db.execute(stmt)
    report_type = r.scalar_one_or_none()
    if not report_type:
        return {"total": 0, "business_affected": 0, "business_unaffected": 0, "alarm_code_names": [], "latest_time": None, "latest_filename": None}

    # 找到最新的 ReportFile（按 id 降序，id 大的 = 最新入库）
    latest_file_stmt = (
        select(ReportFile.id, ReportFile.filename, ReportFile.record_count)
        .where(ReportFile.report_type_id == report_type.id)
        .order_by(ReportFile.id.desc())
        .limit(1)
    )
    r2 = await db.execute(latest_file_stmt)
    latest_row = r2.first()
    if not latest_row:
        return {"total": 0, "business_affected": 0, "business_unaffected": 0, "alarm_code_names": [], "latest_time": None, "latest_filename": None}

    latest_file_id, latest_filename, record_count = latest_row

    if record_count == 0:
        return {"total": 0, "business_affected": 0, "business_unaffected": 0, "alarm_code_names": [], "latest_time": None, "latest_filename": latest_filename}

    # 仅查询最新文件的记录
    data_stmt = (
        select(ReportRecord.data)
        .where(ReportRecord.report_file_id == latest_file_id)
        .order_by(ReportRecord.id.desc())
    )
    r3 = await db.execute(data_stmt)
    rows = r3.all()

    records = [dict(row[0]) for row in rows]

    business_affected = 0
    business_unaffected = 0
    alarm_code_names: list[str] = []
    seen_names = set()

    for rec in records:
        # 统计影响业务 / 不影响业务
        affected = str(rec.get("是否影响业务", "")).strip()
        if affected == "是":
            business_affected += 1
        elif affected == "否":
            business_unaffected += 1

        # 收集告警码名称（去重）
        name = rec.get("告警码名称", "")
        if name and name not in seen_names:
            alarm_code_names.append(name)
            seen_names.add(name)

    latest_time = records[0].get("发生时间", "") if records else None

    return {
        "total": len(records),
        "business_affected": business_affected,
        "business_unaffected": business_unaffected,
        "alarm_code_names": alarm_code_names,
        "latest_time": latest_time,
        "latest_filename": latest_filename,
    }


async def get_access_layer_fault_detail(
    db: AsyncSession,
    page: int = 1,
    page_size: int = 50,
) -> dict:
    """分页获取接入层通报横山详细数据：仅最新一份文件的6个字段数据"""
    stmt = select(ReportType).where(ReportType.name == "接入层通报")
    r = await db.execute(stmt)
    report_type = r.scalar_one_or_none()
    if not report_type:
        return {"records": [], "total": 0, "page": page, "page_size": page_size, "latest_filename": None}

    # 找到最新的 ReportFile
    latest_file_stmt = (
        select(ReportFile.id, ReportFile.filename, ReportFile.record_count)
        .where(ReportFile.report_type_id == report_type.id)
        .order_by(ReportFile.id.desc())
        .limit(1)
    )
    r2 = await db.execute(latest_file_stmt)
    latest_row = r2.first()
    if not latest_row:
        return {"records": [], "total": 0, "page": page, "page_size": page_size, "latest_filename": None}

    latest_file_id, latest_filename, record_count = latest_row

    if record_count == 0:
        return {"records": [], "total": 0, "page": page, "page_size": page_size, "latest_filename": latest_filename}

    # 仅查询最新文件的记录
    data_stmt = (
        select(ReportRecord.data, ReportFile.filename, ReportRecord.created_at)
        .join(ReportFile, ReportRecord.report_file_id == ReportFile.id)
        .where(ReportRecord.report_file_id == latest_file_id)
        .order_by(ReportRecord.id.desc())
    )
    r3 = await db.execute(data_stmt)
    rows = r3.all()

    all_records = []
    for row in rows:
        data, filename, created_at = row
        record = dict(data)
        record["_source_file"] = filename
        record["_created_at"] = created_at.isoformat() if created_at else ""
        all_records.append(record)

    total = len(all_records)

    # 分页
    offset = (page - 1) * page_size
    records = all_records[offset:offset + page_size]

    return {
        "records": records,
        "total": total,
        "page": page,
        "page_size": page_size,
        "latest_filename": latest_filename,
    }


# ── 企宽装机通报横山数据专用处理 ──

def _parse_enterprise_broadband_files(directory: str) -> dict:
    """
    解析最新的"企宽装机通报"文件：
    1. 从"移动汇报"sheet 提取横山汇总指标
    2. 从"积压"sheet 提取横山积压清单（计算装机历时）
    返回 {"summary": dict, "backlog": [dict], "filename": str, "report_date": str}
    """
    from datetime import datetime as _dt

    # 找到最新的企宽装机通报文件
    matching: list[str] = []
    for root, _dirs, files in os.walk(directory):
        for f in files:
            if f.startswith("~$") or f.startswith("."):
                continue
            if "企宽装机通报" in f and f.lower().endswith((".xlsx", ".xls")):
                matching.append(os.path.join(root, f))

    if not matching:
        return {"summary": None, "backlog": [], "filename": None, "report_date": None}

    # 按修改时间排序，取最新
    matching.sort(key=lambda f: os.path.getmtime(f), reverse=True)
    latest_file = matching[0]
    filename = os.path.basename(latest_file)

    try:
        wb = openpyxl.load_workbook(latest_file, read_only=True, data_only=True)

        # ── 1. 解析"移动汇报"sheet ──
        summary = None
        report_date = None
        if "移动汇报" in wb.sheetnames:
            ws = wb["移动汇报"]
            rows_data = list(ws.iter_rows(values_only=True))

            # 提取标题中的日期
            if rows_data and rows_data[0][0]:
                title = str(rows_data[0][0])
                date_match = re.search(r'(\d+)月(\d+)日', title)
                if date_match:
                    now = _dt.now()
                    report_date = f"{now.year}-{date_match.group(1).zfill(2)}-{date_match.group(2).zfill(2)}"

            # 从 Row 2（实际表头行）动态构建列名→列索引映射
            # 模板可能增删列（如"当月退单率""当日退单率"），硬编码索引会错位
            header_row = rows_data[2] if len(rows_data) > 2 else ()
            col_map: dict[str, int] = {}
            target_fields = [
                "当月受理量", "当月归档量", "当月成功率", "当月退单量",
                "积压总量", "当日受理量", "当日归档量", "当日成功率",
                "当日退单量", "当日积压",
            ]
            for target in target_fields:
                for idx, cell in enumerate(header_row):
                    if cell and str(cell).strip() == target:
                        col_map[target] = idx
                        break

            # 查找横山行
            for row in rows_data[4:]:
                if row[0] and str(row[0]).strip() == "横山县":
                    def _safe_str(key: str) -> str:
                        idx = col_map.get(key)
                        if idx is not None and len(row) > idx and row[idx] is not None:
                            return str(row[idx])
                        return ""

                    summary = {
                        "district": "横山",
                        "month_accept": _safe_str("当月受理量"),
                        "month_archive": _safe_str("当月归档量"),
                        "month_success_rate": _safe_str("当月成功率"),
                        "month_reject": _safe_str("当月退单量"),
                        "total_backlog": _safe_str("积压总量"),
                        "day_accept": _safe_str("当日受理量"),
                        "day_archive": _safe_str("当日归档量"),
                        "day_success_rate": _safe_str("当日成功率"),
                        "day_reject": _safe_str("当日退单量"),
                        "day_backlog": _safe_str("当日积压"),
                    }
                    break

        # ── 2. 解析"积压"sheet ──
        backlog_records: list[dict] = []
        if "积压" in wb.sheetnames:
            ws2 = wb["积压"]
            # 流式读取表头（仅第一行），避免 list() 加载百万行导致 OOM
            row_iter = ws2.iter_rows(values_only=True)
            try:
                header_row = next(row_iter)
            except StopIteration:
                header_row = ()

            if not header_row:
                wb.close()
                return {"summary": summary, "backlog": [], "filename": filename, "report_date": report_date}

            col_map: dict[str, int] = {}
            target_fields = ["所属区县", "覆盖场景", "宽带账号", "施工地址", "施工人姓名",
                             "受理时间", "到装维时间", "完成时限", "用户品牌"]
            for target in target_fields:
                for idx, cell in enumerate(header_row):
                    if cell and str(cell).strip() == target:
                        col_map[target] = idx
                        break

            # 确保所有必要字段都找到了
            missing = [f for f in target_fields if f not in col_map]
            if missing:
                logger.warning(f"企宽装机通报积压sheet缺少字段: {missing}")
                wb.close()
                return {"summary": summary, "backlog": [], "filename": filename, "report_date": report_date}

            max_col_idx = max(col_map.values())
            # 流式遍历剩余数据行（不加载全量到内存）
            for row in row_iter:
                if not row or len(row) <= max_col_idx:
                    continue

                district_val = str(row[col_map.get("所属区县")]).strip() if col_map.get("所属区县") is not None and row[col_map["所属区县"]] else ""
                if district_val != "横山县":
                    continue
                
                # 新增：过滤覆盖场景=企宽场景
                scene_val = str(row[col_map.get("覆盖场景")]).strip() if col_map.get("覆盖场景") is not None and row[col_map["覆盖场景"]] else ""
                if scene_val != "企宽场景":
                    continue

                # 计算装机历时(h) = 完成时限 - 到装维时间
                install_duration = ""
                deadline_str = str(row[col_map["完成时限"]]) if "完成时限" in col_map and row[col_map["完成时限"]] else ""
                to_install_str = str(row[col_map["到装维时间"]]) if "到装维时间" in col_map and row[col_map["到装维时间"]] else ""

                if deadline_str and to_install_str:
                    try:
                        deadline_dt = _dt.strptime(deadline_str[:19], "%Y-%m-%d %H:%M:%S")
                        to_install_dt = _dt.strptime(to_install_str[:19], "%Y-%m-%d %H:%M:%S")
                        diff = deadline_dt - to_install_dt
                        hours = diff.total_seconds() / 3600.0
                        install_duration = f"{hours:.2f}"
                    except (ValueError, TypeError):
                        install_duration = ""

                record = {
                    "district": district_val,
                    "cover_scene": scene_val,
                    "account": str(row[col_map["宽带账号"]]) if "宽带账号" in col_map and row[col_map["宽带账号"]] else "",
                    "address": str(row[col_map["施工地址"]]) if "施工地址" in col_map and row[col_map["施工地址"]] else "",
                    "worker_name": str(row[col_map["施工人姓名"]]) if "施工人姓名" in col_map and row[col_map["施工人姓名"]] else "",
                    "accept_time": str(row[col_map["受理时间"]]) if "受理时间" in col_map and row[col_map["受理时间"]] else "",
                    "to_install_time": to_install_str,
                    "deadline": deadline_str,
                    "install_duration_hours": install_duration,
                    "user_brand": str(row[col_map["用户品牌"]]) if "用户品牌" in col_map and row[col_map["用户品牌"]] else "",
                }
                backlog_records.append(record)

            logger.info(f"企宽装机通报横山积压: {filename} -> {len(backlog_records)} 条")

        wb.close()
        return {
            "summary": summary,
            "backlog": backlog_records,
            "filename": filename,
            "report_date": report_date,
        }

    except Exception as e:
        logger.error(f"解析企宽装机通报失败 {filename}: {e}", exc_info=True)
        return {"summary": None, "backlog": [], "filename": filename, "report_date": None}


@record_notification("企宽装机通报")
async def reparse_enterprise_broadband(db: AsyncSession, directory: Optional[str] = None) -> dict:
    """
    重新解析企宽装机通报文件，提取横山汇总指标和积压清单，更新数据库。
    删除旧数据并写入新数据。
    """
    if directory is None:
        directory = settings.watch_dir

    # 在线程池中解析
    result = await asyncio.to_thread(_parse_enterprise_broadband_files, directory)

    # ── 写入汇总表 ──
    # 删除旧汇总数据
    from app.core.models import EnterpriseBroadbandSummary, EnterpriseBroadbandBacklog
    from sqlalchemy import delete as _delete
    await db.execute(_delete(EnterpriseBroadbandSummary))
    await db.execute(_delete(EnterpriseBroadbandBacklog))

    summary_count = 0
    if result["summary"]:
        s = result["summary"]
        ebs = EnterpriseBroadbandSummary(
            report_date=result["report_date"] or "",
            district=s["district"],
            latest_filename=result.get("filename") or "",
            month_accept=s["month_accept"],
            month_archive=s["month_archive"],
            month_success_rate=s["month_success_rate"],
            month_reject=s["month_reject"],
            total_backlog=s["total_backlog"],
            day_accept=s["day_accept"],
            day_archive=s["day_archive"],
            day_success_rate=s["day_success_rate"],
            day_reject=s["day_reject"],
            day_backlog=s["day_backlog"],
        )
        db.add(ebs)
        summary_count = 1

    # ── 写入积压清单 ──
    # 先获取或创建"企宽装机通报"报表类型用于关联 ReportFile
    stmt = select(ReportType).where(ReportType.name == "企宽装机通报")
    r = await db.execute(stmt)
    report_type = r.scalar_one_or_none()
    if not report_type:
        report_type = ReportType(name="企宽装机通报", category="装维生产")
        db.add(report_type)
        await db.flush()

    # 删除旧的企宽装机通报 report_files/report_records
    old_files_stmt = select(ReportFile.id).where(ReportFile.report_type_id == report_type.id)
    r2 = await db.execute(old_files_stmt)
    old_file_ids = [row[0] for row in r2.all()]
    if old_file_ids:
        await db.execute(_delete(ReportRecord).where(ReportRecord.report_file_id.in_(old_file_ids)))
        await db.execute(_delete(ReportFile).where(ReportFile.report_type_id == report_type.id))

    # 创建 ReportFile（用于积压清单关联）
    report_file = ReportFile(
        report_type_id=report_type.id,
        filename=result["filename"] or "",
        file_path=os.path.join(directory, result["filename"] or ""),
        parse_status="parsed",
        record_count=len(result["backlog"]),
    )
    db.add(report_file)
    await db.flush()

    backlog_count = 0
    for rec in result["backlog"]:
        ebb = EnterpriseBroadbandBacklog(
            report_file_id=report_file.id,
            district=rec["district"],
            cover_scene=rec["cover_scene"],
            account=rec["account"],
            address=rec["address"],
            worker_name=rec["worker_name"],
            accept_time=rec["accept_time"],
            to_install_time=rec["to_install_time"],
            deadline=rec["deadline"],
            install_duration_hours=rec["install_duration_hours"],
            user_brand=rec["user_brand"],
        )
        db.add(ebb)
        backlog_count += 1

    await db.commit()

    return {
        "summary_parsed": summary_count,
        "backlog_count": backlog_count,
        "filename": result["filename"],
        "report_date": result["report_date"],
    }


async def get_enterprise_broadband_summary(db: AsyncSession) -> dict:
    """获取企宽装机通报横山卡片指标"""
    from app.core.models import EnterpriseBroadbandSummary
    stmt = select(EnterpriseBroadbandSummary).order_by(EnterpriseBroadbandSummary.id.desc()).limit(1)
    r = await db.execute(stmt)
    row = r.scalar_one_or_none()
    if not row:
        return {
            "district": "横山",
            "month_accept": "", "month_archive": "", "month_success_rate": "",
            "month_reject": "", "total_backlog": "",
            "day_accept": "", "day_archive": "", "day_success_rate": "",
            "day_reject": "", "day_backlog": "",
            "report_date": "", "latest_filename": "",
        }
    return {
        "district": row.district,
        "month_accept": row.month_accept,
        "month_archive": row.month_archive,
        "month_success_rate": row.month_success_rate,
        "month_reject": row.month_reject,
        "total_backlog": row.total_backlog,
        "day_accept": row.day_accept,
        "day_archive": row.day_archive,
        "day_success_rate": row.day_success_rate,
        "day_reject": row.day_reject,
        "day_backlog": row.day_backlog,
        "report_date": row.report_date,
        "latest_filename": getattr(row, "latest_filename", "") or "",
    }


async def get_enterprise_broadband_backlog(
    db: AsyncSession,
    page: int = 1,
    page_size: int = 50,
) -> dict:
    """分页获取企宽装机通报横山积压清单"""
    from app.core.models import EnterpriseBroadbandBacklog
    from sqlalchemy import func as _func

    count_stmt = select(_func.count(EnterpriseBroadbandBacklog.id))
    r = await db.execute(count_stmt)
    total = r.scalar() or 0

    offset = (page - 1) * page_size
    data_stmt = (
        select(EnterpriseBroadbandBacklog)
        .order_by(EnterpriseBroadbandBacklog.id.desc())
        .offset(offset)
        .limit(page_size)
    )
    r2 = await db.execute(data_stmt)
    rows = r2.scalars().all()

    records = []
    for row in rows:
        records.append({
            "id": row.id,
            "district": row.district,
            "cover_scene": row.cover_scene,
            "account": row.account,
            "address": row.address,
            "worker_name": row.worker_name,
            "accept_time": row.accept_time,
            "to_install_time": row.to_install_time,
            "deadline": row.deadline,
            "install_duration_hours": row.install_duration_hours,
            "user_brand": row.user_brand,
        })

    return {
        "records": records,
        "total": total,
        "page": page,
        "page_size": page_size,
    }


# ── 日报横山数据专用处理 ──

# 两类指标保留字段
TWO_CAT_FIELDS = [
    "积压总量",      # backlog_total
    "家宽转化率",    # broadband_rate
    "FTTR转化率",    # fttr_rate
    "总装机转化率",  # total_rate
]

# 五类指标保留字段
FIVE_CAT_FIELDS = [
    "积压总量",      # backlog_total
    "家宽转化率",    # broadband_rate
    "智能组网",      # smart_network
    "平安乡村",      # safe_village
    "FTTR转化率",    # fttr_rate
    "总装机转化率",  # total_rate
]

# 宽带积压保留字段
BACKLOG_FIELDS = [
    "所属区县",
    "覆盖场景",
    "宽带账号",
    "服务",
    "施工地址",
    "施工人姓名",
    "工单状态",
    "受理时间",
    "到装维时间",
    "完成时限",
    "积压时长h",
    "用户品牌",
]


def _parse_daily_report_files(directory: str) -> dict:
    """
    解析所有"日报"文件，提取横山数据：
    1. "两类" sheet → 两类装机成功率概况
    2. "五类" sheet → 五类装机成功率概况
    3. "宽带积压" sheet → 装机积压清单（含计算装机历时）
    使用 openpyxl(read_only=True) + 迭代器流式读取。
    返回 {"summary": dict, "backlog": [dict], "filename": str, "report_date": str}
    """
    from datetime import datetime as _dt

    # 找到最新的日报文件
    matching: list[str] = []
    for root, _dirs, files in os.walk(directory):
        for f in files:
            if f.startswith("~$") or f.startswith("."):
                continue
            if "日报" in f and f.lower().endswith((".xlsx", ".xls")):
                matching.append(os.path.join(root, f))

    if not matching:
        return {"summary": None, "backlog": [], "filename": None, "report_date": None}

    # 按修改时间排序，取最新
    matching.sort(key=lambda f: os.path.getmtime(f), reverse=True)
    latest_file = matching[0]
    filename = os.path.basename(latest_file)

    # 尝试从文件名提取日期
    report_date = None
    date_match = re.search(r'(\d{4})[-年](\d{1,2})[-月](\d{1,2})', filename)
    if date_match:
        report_date = f"{date_match.group(1)}-{date_match.group(2).zfill(2)}-{date_match.group(3).zfill(2)}"
    else:
        report_date = _dt.now().strftime("%Y-%m-%d")

    try:
        wb = openpyxl.load_workbook(latest_file, read_only=True, data_only=True)
        summary: dict = {}
        backlog_records: list[dict] = []

        # ── 1. 解析"两类" sheet ──
        if "两类" in wb.sheetnames:
            ws = wb["两类"]
            summary["two_cat"] = _parse_category_sheet(ws, "两类", TWO_CAT_FIELDS)
        else:
            logger.warning(f"日报文件 {filename} 缺少'两类'sheet")

        # ── 2. 解析"五类" sheet ──
        if "五类" in wb.sheetnames:
            ws = wb["五类"]
            summary["five_cat"] = _parse_category_sheet(ws, "五类", FIVE_CAT_FIELDS)
        else:
            logger.warning(f"日报文件 {filename} 缺少'五类'sheet")

        # ── 3. 解析"宽带积压" sheet ──
        if "宽带积压" in wb.sheetnames:
            ws = wb["宽带积压"]
            backlog_records = _parse_backlog_sheet(ws, filename, source_label="宽带积压")
        else:
            logger.warning(f"日报文件 {filename} 缺少'宽带积压'sheet")

        # ── 4. 解析"FTTR积压" sheet ──
        if "FTTR积压" in wb.sheetnames:
            ws = wb["FTTR积压"]
            fttr_records = _parse_backlog_sheet(ws, filename, source_label="FTTR积压")
            backlog_records.extend(fttr_records)
        else:
            logger.warning(f"日报文件 {filename} 缺少'FTTR积压'sheet")

        wb.close()

        logger.info(
            f"日报横山: {filename} -> 两类={summary.get('two_cat')}, "
            f"五类={summary.get('five_cat')}, 积压={len(backlog_records)} 条"
        )

        return {
            "summary": summary,
            "backlog": backlog_records,
            "filename": filename,
            "report_date": report_date,
        }

    except Exception as e:
        logger.error(f"解析日报文件失败 {filename}: {e}", exc_info=True)
        return {"summary": None, "backlog": [], "filename": filename, "report_date": report_date}


def _parse_category_sheet(ws, sheet_label: str, target_fields: list[str]) -> dict | None:
    """
    解析"两类"/"五类"sheet，定位横山行并提取指标值。
    
    策略：流式遍历行，在第一列中查找"横山"或"横山县"。
    表头行通过标题行（包含"两类"/"五类"关键词）定位。
    找到横山行后，按目标字段顺序提取对应列的值。
    """
    try:
        rows_data = list(ws.iter_rows(values_only=True))
    except Exception:
        # 某些 read_only 模式下无法 list，回退逐行读取
        rows_data = []
        for row in ws.iter_rows(values_only=True):
            rows_data.append(row)

    if len(rows_data) < 2:
        return None

    # 查找横山行：扫描全表，找到第一列是"横山"或"横山县"的行，同时用该行上方最近的非空行作为可能的表头
    hengshan_row_idx = None
    hengshan_row = None
    header_row_idx = None

    for idx, row in enumerate(rows_data):
        first_cell = str(row[0]).strip() if row and row[0] else ""
        if first_cell in ("横山", "横山县"):
            hengshan_row_idx = idx
            hengshan_row = row
            break

    if hengshan_row_idx is None:
        logger.info(f"日报 {sheet_label} sheet 未找到横山行")
        return None

    # 查找表头行：从第1行开始（跳过标题行0），向横山行方向搜索包含指标关键字的行
    for idx in range(1, hengshan_row_idx):
        row = rows_data[idx]
        row_text = " ".join(str(c) for c in row if c)
        # 检查是否包含"积压总量"或"转化率"等指标关键字
        if any(kw in row_text for kw in ["积压总量", "转化率", "装机"]):
            header_row_idx = idx
            break

    result: dict[str, str] = {}
    for i, field in enumerate(target_fields):
        result[field] = ""

    if hengshan_row is None:
        return result

    # 如果有表头行，按表头列名匹配
    if header_row_idx is not None:
        header_row = rows_data[header_row_idx]
        for i, field in enumerate(target_fields):
            for col_idx, header_cell in enumerate(header_row):
                if header_cell and field in str(header_cell).strip():
                    val = hengshan_row[col_idx] if col_idx < len(hengshan_row) else ""
                    result[field] = str(val).strip() if val is not None else ""
                    break

    # 如果表头匹配失败，回退：按横山行数据顺序填充（跳过第一列区县名）
    if all(v == "" for v in result.values()):
        for i, field in enumerate(target_fields):
            col_idx = i + 1  # 跳过第一列（区县名）
            if col_idx < len(hengshan_row):
                val = hengshan_row[col_idx]
                result[field] = str(val).strip() if val is not None else ""

    return result


def _parse_backlog_sheet(ws, filename: str, source_label: str = "宽带积压") -> list[dict]:
    """
    解析"宽带积压"或"FTTR积压"sheet，流式读取，过滤横山区数据。
    计算装机历时(h) = 完成时限 - 到装维时间。
    返回 [record, ...]，每条记录带 data_source 标记。
    """
    from datetime import datetime as _dt

    row_iter = ws.iter_rows(values_only=True)

    # 读取前几行找表头（表头通常在前5行内）
    header_row = None
    header_rows_buf: list[tuple] = []
    for i in range(10):
        try:
            r = next(row_iter)
            header_rows_buf.append(r)
            # 检查是否包含关键字段
            row_text = " ".join(str(c) for c in r if c)
            if any(kw in row_text for kw in ["宽带账号", "所属区县", "施工地址"]):
                header_row = r
                break
        except StopIteration:
            break

    if header_row is None:
        logger.warning(f"日报{source_label} sheet 未检测到表头行: {filename}")
        return []

    # 建立列映射（字段名匹配，>=2字符防止单字符误匹配）
    col_map: dict[str, int] = {}
    
    # 先处理需要语义回退的特殊字段
    _SEMANTIC_ALIASES = {
        "用户品牌": "客户等级",
    }
    
    for idx, cell in enumerate(header_row):
        if cell is None:
            continue
        cell_str = str(cell).strip()
        for field in BACKLOG_FIELDS:
            if len(field) >= 2 and field in cell_str:
                col_map[field] = idx
                break
    
    # 语义回退：如果原始字段没匹配到，用别名再试
    for field, alias in _SEMANTIC_ALIASES.items():
        if field not in col_map:
            for idx, cell in enumerate(header_row):
                if cell is None:
                    continue
                cell_str = str(cell).strip()
                if alias in cell_str:
                    col_map[field] = idx
                    logger.info(f"日报{source_label}: '{field}' 语义回退到 '{alias}' col={idx}")
                    break

    # 验证关键字段
    missing = [f for f in ["所属区县", "覆盖场景", "宽带账号"] if f not in col_map]
    if missing:
        logger.warning(f"日报{source_label} sheet 缺少关键字段: {missing}")
        return []

    # "积压时长h" 特殊处理：表头单元格为合并残留（None或数字），
    # 查找 col1 位置（通常是积压时长数值列，值为数字型小时数）
    if "积压时长h" not in col_map:
        # 检查 col1：如果表头为 None/数字 且数据行为数值，则认定为积压时长列
        if len(header_row) > 1 and (header_row[1] is None or isinstance(header_row[1], (int, float))):
            col_map["积压时长h"] = 1
            logger.info(f"日报{source_label}: '积压时长h' 通过 col1 (合并单元格) 检测到")
        elif source_label == "FTTR积压" and len(header_row) > 2:
            # FTTR积压特殊处理：col1为积压天数（小数），col2为积压时长h，
            # col3为积压时长标签（如"48小时以上"）。优先使用col2。
            col_map["积压时长h"] = 2
            logger.info(f"日报{source_label}: '积压时长h' 通过 col2 (FTTR合并单元格回退) 检测到")

    max_col_idx = max(col_map.values())

    # 流式遍历数据行，过滤横山 + 家庭场景
    backlog_records: list[dict] = []
    for row in row_iter:
        if not row or len(row) <= max_col_idx:
            continue

        # 过滤横山
        district_val = ""
        if "所属区县" in col_map and row[col_map["所属区县"]]:
            district_val = str(row[col_map["所属区县"]]).strip()

        if district_val != "横山" and district_val != "横山县":
            continue

        # 过滤家庭场景
        scene_val = ""
        if "覆盖场景" in col_map and row[col_map["覆盖场景"]]:
            scene_val = str(row[col_map["覆盖场景"]]).strip()

        if scene_val != "家庭场景":
            continue

        # 提取字段
        def _get(field: str) -> str:
            if field in col_map and col_map[field] < len(row) and row[col_map[field]] is not None:
                val = row[col_map[field]]
                if isinstance(val, _dt):
                    return val.strftime("%Y-%m-%d %H:%M:%S")
                return str(val).strip()
            return ""

        # 计算装机历时(h)
        install_duration = ""
        deadline_str = _get("完成时限")
        to_install_str = _get("到装维时间")

        if deadline_str and to_install_str:
            # 尝试多种日期格式
            for fmt in ["%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S",
                        "%Y-%m-%d %H:%M", "%Y/%m/%d %H:%M",
                        "%Y-%m-%d", "%Y/%m/%d"]:
                try:
                    dl = deadline_str.strip()
                    ti = to_install_str.strip()
                    deadline_dt = _dt.strptime(dl, fmt)
                    to_install_dt = _dt.strptime(ti, fmt)
                    diff = deadline_dt - to_install_dt
                    hours = diff.total_seconds() / 3600.0
                    install_duration = f"{hours:.2f}"
                    break
                except (ValueError, TypeError):
                    continue

        # 计算积压时长提醒标签
        duration_warning = ""
        try:
            duration_h = float(install_duration) if install_duration else 0
            if duration_h > 48:
                duration_warning = "超48h"
            elif duration_h > 24:
                duration_warning = "超24h"
            elif duration_h > 8:
                duration_warning = "超8h"
        except (ValueError, TypeError):
            pass

        record = {
            "所属区县": _get("所属区县"),
            "覆盖场景": _get("覆盖场景"),
            "宽带账号": _get("宽带账号"),
            "服务": _get("服务"),
            "施工地址": _get("施工地址"),
            "施工人姓名": _get("施工人姓名"),
            "工单状态": _get("工单状态"),
            "受理时间": _get("受理时间"),
            "到装维时间": _get("到装维时间"),
            "完成时限": _get("完成时限"),
            "积压时长h": _get("积压时长h"),
            "装机历时(h)": install_duration,
            "时长提醒": duration_warning,
            "用户品牌": _get("用户品牌"),
            "数据来源": source_label,
        }

        backlog_records.append(record)

    logger.info(f"日报{source_label}横山: {filename} -> {len(backlog_records)} 条")
    return backlog_records


@record_notification("日报")
async def reparse_daily_report(db: AsyncSession, directory: Optional[str] = None) -> dict:
    """
    重新解析日报文件，提取横山两类/五类概况 + 宽带积压清单，更新数据库。
    删除旧数据并写入新数据。
    """
    if directory is None:
        directory = settings.watch_dir

    # 在线程池中解析
    result = await asyncio.to_thread(_parse_daily_report_files, directory)

    # 导入模型
    from app.core.models import DailyReportSummary, DailyReportBacklog
    from sqlalchemy import delete as _delete

    # 删除旧数据
    await db.execute(_delete(DailyReportSummary))
    await db.execute(_delete(DailyReportBacklog))

    # ── 写入汇总表 ──
    summary_count = 0
    if result["summary"]:
        s = result["summary"]
        two = s.get("two_cat") or {}
        five = s.get("five_cat") or {}

        drs = DailyReportSummary(
            report_date=result["report_date"] or "",
            latest_filename=result.get("filename") or "",
            two_cat_backlog_total=two.get("积压总量", ""),
            two_cat_broadband_rate=two.get("家宽转化率", ""),
            two_cat_fttr_rate=two.get("FTTR转化率", ""),
            two_cat_total_rate=two.get("总装机转化率", ""),
            five_cat_backlog_total=five.get("积压总量", ""),
            five_cat_broadband_rate=five.get("家宽转化率", ""),
            five_cat_smart_network=five.get("智能组网", ""),
            five_cat_safe_village=five.get("平安乡村", ""),
            five_cat_fttr_rate=five.get("FTTR转化率", ""),
            five_cat_total_rate=five.get("总装机转化率", ""),
        )
        db.add(drs)
        summary_count = 1

    # ── 获取或创建"日报"报表类型 ──
    stmt = select(ReportType).where(ReportType.name == "日报")
    r = await db.execute(stmt)
    report_type = r.scalar_one_or_none()
    if not report_type:
        report_type = ReportType(name="日报", category="装维生产")
        db.add(report_type)
        await db.flush()

    # 删除旧的日报 report_files/report_records
    old_files_stmt = select(ReportFile.id).where(ReportFile.report_type_id == report_type.id)
    r2 = await db.execute(old_files_stmt)
    old_file_ids = [row[0] for row in r2.all()]
    if old_file_ids:
        await db.execute(_delete(ReportRecord).where(ReportRecord.report_file_id.in_(old_file_ids)))
        await db.execute(_delete(ReportFile).where(ReportFile.report_type_id == report_type.id))

    # 创建 ReportFile
    report_file = ReportFile(
        report_type_id=report_type.id,
        filename=result["filename"] or "",
        file_path=os.path.join(directory, result["filename"] or ""),
        parse_status="parsed",
        record_count=len(result["backlog"]),
    )
    db.add(report_file)
    await db.flush()

    # ── 写入积压清单 ──
    backlog_count = 0
    for rec in result["backlog"]:
        drb = DailyReportBacklog(
            report_file_id=report_file.id,
            district=rec.get("所属区县", ""),
            coverage_scenario=rec.get("覆盖场景", ""),
            account=rec.get("宽带账号", ""),
            service=rec.get("服务", ""),
            address=rec.get("施工地址", ""),
            worker_name=rec.get("施工人姓名", ""),
            order_status=rec.get("工单状态", ""),
            accept_time=rec.get("受理时间", ""),
            to_install_time=rec.get("到装维时间", ""),
            deadline=rec.get("完成时限", ""),
            backlog_hours=rec.get("积压时长h", ""),
            install_duration_hours=rec.get("装机历时(h)", ""),
            user_brand=rec.get("用户品牌", ""),
            data_source=rec.get("数据来源", ""),
        )
        db.add(drb)
        backlog_count += 1

    # 更新 column_hint
    report_type.column_hint = BACKLOG_FIELDS.copy()
    report_type.updated_at = datetime.now(timezone.utc)

    await db.commit()

    # ── 贾维斯：数据更新后自动触发 AI 分析 ──
    try:
        from app.ai.ai_scheduler import trigger_ai_analysis
        asyncio.create_task(trigger_ai_analysis("daily-report", db))
    except Exception:
        pass

    return {
        "summary_parsed": summary_count,
        "backlog_count": backlog_count,
        "filename": result["filename"],
        "report_date": result["report_date"],
    }


async def get_daily_report_summary(db: AsyncSession) -> dict:
    """获取日报横山卡片汇总指标（两类+五类装机成功率）"""
    from app.core.models import DailyReportSummary

    stmt = select(DailyReportSummary).order_by(DailyReportSummary.id.desc()).limit(1)
    r = await db.execute(stmt)
    row = r.scalar_one_or_none()

    if not row:
        return {
            "report_date": "",
            "latest_filename": "",
            "two_cat": {"积压总量": "", "家宽转化率": "", "FTTR转化率": "", "总装机转化率": ""},
            "five_cat": {"积压总量": "", "家宽转化率": "", "智能组网": "", "平安乡村": "", "FTTR转化率": "", "总装机转化率": ""},
        }

    return {
        "report_date": row.report_date,
        "latest_filename": getattr(row, "latest_filename", "") or "",
        "two_cat": {
            "积压总量": row.two_cat_backlog_total,
            "家宽转化率": row.two_cat_broadband_rate,
            "FTTR转化率": row.two_cat_fttr_rate,
            "总装机转化率": row.two_cat_total_rate,
        },
        "five_cat": {
            "积压总量": row.five_cat_backlog_total,
            "家宽转化率": row.five_cat_broadband_rate,
            "智能组网": row.five_cat_smart_network,
            "平安乡村": row.five_cat_safe_village,
            "FTTR转化率": row.five_cat_fttr_rate,
            "总装机转化率": row.five_cat_total_rate,
        },
    }


async def get_daily_report_backlog(
    db: AsyncSession,
    page: int = 1,
    page_size: int = 50,
    sort_by: Optional[str] = None,
    order: str = "asc",
) -> dict:
    """分页获取日报横山装机积压清单
    - sort_by: install_duration_hours 时按装机历时数字排序；其他按 id 排序
    - order: asc=升序，desc=降序
    """
    from app.core.models import DailyReportBacklog
    from sqlalchemy import func as _func, cast, Numeric, nullslast

    count_stmt = select(_func.count(DailyReportBacklog.id))
    r = await db.execute(count_stmt)
    total = r.scalar() or 0

    offset = (page - 1) * page_size
    data_stmt = select(DailyReportBacklog)

    # 动态排序
    if sort_by == "install_duration_hours":
        # 装机历时是字符串存数字，需 CAST 为 NUMERIC，空值排最后
        col_expr = cast(_func.nullif(DailyReportBacklog.install_duration_hours, ""), Numeric)
        if order == "desc":
            data_stmt = data_stmt.order_by(nullslast(col_expr.desc()))
        else:
            data_stmt = data_stmt.order_by(nullslast(col_expr.asc()))
    elif sort_by and hasattr(DailyReportBacklog, sort_by):
        col = getattr(DailyReportBacklog, sort_by)
        if order == "desc":
            data_stmt = data_stmt.order_by(col.desc())
        else:
            data_stmt = data_stmt.order_by(col.asc())
    else:
        data_stmt = data_stmt.order_by(DailyReportBacklog.id.asc())
    r2 = await db.execute(data_stmt)
    rows = r2.scalars().all()

    records = []
    for row in rows:
        # 计算时长提醒标签
        duration_warning = ""
        try:
            duration_h = float(row.install_duration_hours) if row.install_duration_hours else 0
            if duration_h > 48:
                duration_warning = "超48h"
            elif duration_h > 24:
                duration_warning = "超24h"
            elif duration_h > 8:
                duration_warning = "超8h"
        except (ValueError, TypeError):
            pass

        records.append({
            "id": row.id,
            "所属区县": row.district,
            "覆盖场景": row.coverage_scenario,
            "宽带账号": row.account,
            "服务": row.service,
            "施工地址": row.address,
            "施工人姓名": row.worker_name,
            "工单状态": row.order_status,
            "受理时间": row.accept_time,
            "到装维时间": row.to_install_time,
            "完成时限": row.deadline,
            "积压时长h": row.backlog_hours,
            "装机历时(h)": row.install_duration_hours,
            "时长提醒": duration_warning,
            "用户品牌": row.user_brand,
            "数据来源": row.data_source,
        })

    return {
        "records": records,
        "total": total,
        "page": page,
        "page_size": page_size,
    }


async def get_report_records(
    report_type_id: int,
    db: AsyncSession,
    page: int = 1,
    page_size: int = 50,
) -> Dict:
    """分页获取某报表类型的记录"""
    # 先查 report_type 是否存在
    stmt = select(ReportType).where(ReportType.id == report_type_id)
    r = await db.execute(stmt)
    rt = r.scalar_one_or_none()
    if not rt:
        return {"records": [], "total": 0, "page": page, "page_size": page_size}

    # 查记录（通过 report_files 关联）
    from sqlalchemy import join
    j = join(ReportRecord, ReportFile, ReportRecord.report_file_id == ReportFile.id)
    count_stmt = select(func.count(ReportRecord.id)).select_from(j).where(
        ReportFile.report_type_id == report_type_id
    )
    r2 = await db.execute(count_stmt)
    total = r2.scalar() or 0

    offset = (page - 1) * page_size
    data_stmt = (
        select(ReportRecord.data, ReportFile.filename, ReportRecord.created_at)
        .select_from(j)
        .where(ReportFile.report_type_id == report_type_id)
        .order_by(ReportRecord.id.desc())
        .offset(offset)
        .limit(page_size)
    )
    r3 = await db.execute(data_stmt)
    rows = r3.all()

    records = []
    for row in rows:
        data, filename, created_at = row
        record = dict(data)
        record["_source_file"] = filename
        record["_created_at"] = created_at.isoformat() if created_at else ""
        records.append(record)

    return {
        "records": records,
        "total": total,
        "page": page,
        "page_size": page_size,
    }


# ── 全市装维工作量统计横山数据专用处理 ──

# "汇总"sheet 目标字段关键字
_CITY_WORKLOAD_SUMMARY_KEYWORDS = {
    "人员数量": "total_staff",
    "有工作量人数": "working_staff",
    "请假人数": "leave_staff",
    "无工作量占比": "no_work_ratio",
}

# "到个人"sheet 目标字段
_CITY_WORKLOAD_WORKER_FIELDS = [
    "姓名", "区域", "装移拆", "投诉", "LAN口", "巡检",
    "一户一案", "质差弱光", "小计",
]


def _parse_city_workload_files(directory: str) -> dict:
    """
    解析最新的"全市装维工作量统计"文件：
    1. 从"汇总"sheet 提取横山汇总指标
    2. 从"到个人"sheet 提取横山装维人员工作量明细
    返回 {"summary": dict, "workers": [dict], "filename": str, "report_date": str}
    """
    from datetime import datetime as _dt

    # 找到最新的全市装维工作量统计文件
    matching: list[str] = []
    for root, _dirs, files in os.walk(directory):
        for f in files:
            if f.startswith("~$") or f.startswith("."):
                continue
            if "全市装维工作量统计" in f and f.lower().endswith((".xlsx", ".xls")):
                matching.append(os.path.join(root, f))

    if not matching:
        return {"summary": None, "workers": [], "filename": None, "report_date": None}

    # 按修改时间排序，取最新
    matching.sort(key=lambda f: os.path.getmtime(f), reverse=True)
    latest_file = matching[0]
    filename = os.path.basename(latest_file)

    # 尝试从文件名提取日期
    report_date = None
    date_match = re.search(r'(\d{4})[-年](\d{1,2})[-月](\d{1,2})', filename)
    if date_match:
        report_date = f"{date_match.group(1)}-{date_match.group(2).zfill(2)}-{date_match.group(3).zfill(2)}"
    else:
        report_date = _dt.now().strftime("%Y-%m-%d")

    try:
        wb = openpyxl.load_workbook(latest_file, read_only=True, data_only=True)

        # ── 1. 解析"汇总"sheet ──
        summary = None
        if "汇总" in wb.sheetnames:
            ws = wb["汇总"]
            rows_data = list(ws.iter_rows(values_only=True))

            # 表头跨两行（Row1=分组标题, Row2=子标题），需要合并扫描
            # 构建 keyword -> col_index 的映射（扫描所有header行）
            keyword_col_map: dict[str, int] = {}
            hengshan_row_idx = None
            hengshan_row = None

            for idx, row in enumerate(rows_data[:5]):  # 只扫描前5行作为header区域
                row_text = " ".join(str(c) for c in row if c)
                for col_i, cell in enumerate(row):
                    if cell is None:
                        continue
                    cell_str = str(cell).strip()
                    for kw in _CITY_WORKLOAD_SUMMARY_KEYWORDS:
                        if kw in cell_str and kw not in keyword_col_map:
                            keyword_col_map[kw] = col_i

                # 同时检测横山数据行
                if "横山" in row_text or "横山县" in row_text:
                    hengshan_row_idx = idx
                    hengshan_row = row
                if "合计" in row_text:
                    break  # 合计行是数据行，不是header

            # 继续扫描剩余数据行找横山
            if hengshan_row is None:
                for idx in range(5, len(rows_data)):
                    row = rows_data[idx]
                    row_text = " ".join(str(c) for c in row if c)
                    if "横山" in row_text or "横山县" in row_text:
                        hengshan_row_idx = idx
                        hengshan_row = row
                        break

            # 提取数据
            if keyword_col_map and hengshan_row is not None:
                summary = {
                    "district": "横山",
                    "total_staff": "",
                    "working_staff": "",
                    "leave_staff": "",
                    "no_work_ratio": "",
                }
                for kw, key in _CITY_WORKLOAD_SUMMARY_KEYWORDS.items():
                    col_idx = keyword_col_map.get(kw)
                    if col_idx is not None and col_idx < len(hengshan_row):
                        val = hengshan_row[col_idx]
                        if val is not None:
                            val_str = str(val).strip()
                            # 百分比格式化
                            if "占比" in kw or "率" in kw:
                                try:
                                    pct = float(val_str)
                                    if pct < 1:
                                        val_str = f"{pct * 100:.1f}%"
                                    else:
                                        val_str = f"{pct:.1f}%"
                                except (ValueError, TypeError):
                                    pass
                            summary[key] = val_str

            if summary:
                logger.info(f"全市装维工作量统计横山汇总: {filename} -> {summary}")
            else:
                logger.warning(f"全市装维工作量统计: {filename} 汇总sheet未找到横山数据")
        else:
            logger.warning(f"全市装维工作量统计: {filename} 缺少'汇总'sheet")

        # ── 2. 解析"到个人"sheet ──
        workers: list[dict] = []
        if "到个人" in wb.sheetnames:
            ws2 = wb["到个人"]
            rows_data2 = list(ws2.iter_rows(values_only=True))

            # 表头结构: Row0=标题, Row1=分组标题(县区/姓名/累计积压量/当日工作量统计), Row2=子标题(详细列名)
            # 跳过Row0(标题行)，组合Row1+Row2来理解列含义
            if len(rows_data2) >= 3:
                row1 = rows_data2[1]  # 分组标题: 县区, 姓名, 账号, 岗位, 网格, 累计积压量, ..., 当日工作量统计, ...
                row2 = rows_data2[2]  # 子标题: 装移拆, 投诉, LAN口（到个人）, ...

                # 建立列映射
                name_col = None
                area_col = None
                grid_col = None
                # 工作类型 -> {"backlog": col_idx, "today": col_idx}
                wt_col_map: dict[str, dict[str, int]] = {}
                # 记录"累计积压量"组和"当日工作量统计"组的起始列
                backlog_group_start = None
                today_group_start = None

                # 第一遍：从Row1找到区县、姓名、以及"累计积压量"和"当日工作量统计"分组起始列
                for idx, cell in enumerate(row1):
                    if cell is None:
                        continue
                    cell_str = str(cell).strip()
                    if cell_str == "县区":
                        area_col = idx
                    elif cell_str == "姓名":
                        name_col = idx
                    elif cell_str == "网格":
                        grid_col = idx
                    elif "累计积压量" in cell_str or "积压量" in cell_str:
                        backlog_group_start = idx
                    elif "当日工作量" in cell_str or "当日工作" in cell_str:
                        today_group_start = idx

                # 如果没找到姓名或区域，尝试从Row2找
                if name_col is None or area_col is None:
                    for idx, cell in enumerate(row2):
                        if cell is None:
                            continue
                        cell_str = str(cell).strip()
                        if name_col is None and cell_str == "姓名":
                            name_col = idx
                        if area_col is None and cell_str == "县区":
                            area_col = idx

                # 第二遍：从Row2找到各工作类型的积压/当日列
                # 注意：排除"小计"和"请假"，它们不是真实工作类型
                # 积压工作类型映射（Row2中的名称 -> 简化名称）
                wt_name_map = {
                    "装移拆": "装移拆", "投诉": "投诉", "LAN口": "LAN口",
                    "LAN口（到个人）": "LAN口",
                    "巡检": "巡检", "一户一案": "一户一案", "质差弱光": "质差弱光",
                }
                # 当日工作类型映射
                today_name_map = {
                    "装移拆": "装移拆", "投诉归档": "投诉", "投诉": "投诉",
                    "LAN口": "LAN口",
                    "巡检（当日）": "巡检", "巡检": "巡检",
                    "一户一案": "一户一案", "质差弱光": "质差弱光",
                }
                excluded_work_types = {"小计", "请假"}

                for idx, cell in enumerate(row2):
                    if cell is None:
                        continue
                    cell_str = str(cell).strip()
                    # 跳过排除的工作类型列
                    if cell_str in excluded_work_types:
                        continue

                    # 判断属于积压组还是当日组
                    if backlog_group_start is not None and idx >= backlog_group_start:
                        if today_group_start is not None and idx >= today_group_start:
                            # 在当日组
                            wt_simple = today_name_map.get(cell_str, cell_str)
                            for orig, simple in today_name_map.items():
                                if orig in cell_str:
                                    wt_simple = simple
                                    break
                            if wt_simple not in excluded_work_types and wt_simple not in wt_col_map:
                                wt_col_map[wt_simple] = {"backlog": None, "today": None}
                            if wt_simple not in excluded_work_types:
                                wt_col_map[wt_simple]["today"] = idx
                        else:
                            # 在积压组（当日组还没开始）
                            for orig, simple in wt_name_map.items():
                                if orig in cell_str:
                                    if simple not in excluded_work_types and simple not in wt_col_map:
                                        wt_col_map[simple] = {"backlog": None, "today": None}
                                    if simple not in excluded_work_types:
                                        wt_col_map[simple]["backlog"] = idx
                                    break

                # 流式遍历数据行（从Row3开始）
                for row in rows_data2[3:]:
                    if not row:
                        continue

                    # 提取区域，过滤横山
                    area = ""
                    if area_col is not None and area_col < len(row) and row[area_col]:
                        area = str(row[area_col]).strip()

                    # 只保留横山人员
                    if "横山" not in area:
                        continue

                    # 提取姓名
                    name_idx = name_col if name_col is not None else 1
                    if name_idx >= len(row) or not row[name_idx]:
                        continue
                    worker_name = str(row[name_idx]).strip()
                    if not worker_name or worker_name in ("nan", "#N/A"):
                        continue

                    # 提取网格
                    worker_grid = ""
                    if grid_col is not None and grid_col < len(row) and row[grid_col]:
                        worker_grid = str(row[grid_col]).strip()

                    # 提取各工作类型的积压和当日数据
                    workload: dict[str, dict[str, int]] = {}
                    total_backlog = 0
                    total_today = 0

                    for wt, cols in wt_col_map.items():
                        backlog_idx = cols.get("backlog")
                        today_idx = cols.get("today")

                        backlog_val = 0
                        today_val = 0

                        if backlog_idx is not None and backlog_idx < len(row) and row[backlog_idx] is not None:
                            try:
                                v = str(row[backlog_idx]).strip()
                                if v and v not in ("#N/A", "nan"):
                                    backlog_val = int(float(v))
                            except (ValueError, TypeError):
                                pass

                        if today_idx is not None and today_idx < len(row) and row[today_idx] is not None:
                            try:
                                v = str(row[today_idx]).strip()
                                if v and v not in ("#N/A", "nan"):
                                    today_val = int(float(v))
                            except (ValueError, TypeError):
                                pass

                        workload[wt] = {"backlog": backlog_val, "today": today_val}
                        total_backlog += backlog_val
                        total_today += today_val

                    workers.append({
                        "worker_name": worker_name,
                        "area": area,
                        "grid": worker_grid,
                        "workload": workload,
                        "total_backlog": total_backlog,
                        "total_today": total_today,
                    })

                logger.info(f"全市装维工作量统计横山到个人: {filename} -> {len(workers)} 人")
            else:
                logger.warning(f"全市装维工作量统计: {filename} '到个人'sheet数据行不足")
        else:
            logger.warning(f"全市装维工作量统计: {filename} 缺少'到个人'sheet")

        wb.close()
        return {
            "summary": summary,
            "workers": workers,
            "filename": filename,
            "report_date": report_date,
        }

    except Exception as e:
        logger.error(f"解析全市装维工作量统计失败 {filename}: {e}", exc_info=True)
        return {"summary": None, "workers": [], "filename": filename, "report_date": report_date}


@record_notification("全市装维工作量统计")
async def reparse_city_workload(db: AsyncSession, directory: Optional[str] = None) -> dict:
    """
    重新解析全市装维工作量统计文件，提取横山汇总指标和人员明细，更新数据库。
    删除旧数据并写入新数据。
    """
    if directory is None:
        directory = settings.watch_dir

    # 在线程池中解析
    result = await asyncio.to_thread(_parse_city_workload_files, directory)

    # 导入模型
    from app.core.models import CityWorkloadSummary, CityWorkloadWorker
    from sqlalchemy import delete as _delete

    # 删除旧数据
    await db.execute(_delete(CityWorkloadSummary))
    await db.execute(_delete(CityWorkloadWorker))

    # ── 写入汇总表 ──
    summary_count = 0
    if result["summary"]:
        s = result["summary"]
        cws = CityWorkloadSummary(
            report_date=result["report_date"] or "",
            district=s.get("district", "横山"),
            latest_filename=result.get("filename") or "",
            total_staff=s.get("total_staff", ""),
            working_staff=s.get("working_staff", ""),
            leave_staff=s.get("leave_staff", ""),
            no_work_ratio=s.get("no_work_ratio", ""),
        )
        db.add(cws)
        summary_count = 1

    # ── 获取或创建报表类型 ──
    stmt = select(ReportType).where(ReportType.name == "全市装维工作量统计")
    r = await db.execute(stmt)
    report_type = r.scalar_one_or_none()
    if not report_type:
        report_type = ReportType(name="全市装维工作量统计", category="装维生产")
        db.add(report_type)
        await db.flush()

    # 删除旧的 report_files/report_records
    old_files_stmt = select(ReportFile.id).where(ReportFile.report_type_id == report_type.id)
    r2 = await db.execute(old_files_stmt)
    old_file_ids = [row[0] for row in r2.all()]
    if old_file_ids:
        await db.execute(_delete(ReportRecord).where(ReportRecord.report_file_id.in_(old_file_ids)))
        await db.execute(_delete(ReportFile).where(ReportFile.report_type_id == report_type.id))

    # 创建 ReportFile
    report_file = ReportFile(
        report_type_id=report_type.id,
        filename=result["filename"] or "",
        file_path=os.path.join(directory, result["filename"] or ""),
        parse_status="parsed",
        record_count=len(result["workers"]),
    )
    db.add(report_file)
    await db.flush()

    # ── 写入人员明细 ──
    worker_count = 0
    for rec in result["workers"]:
        cw = CityWorkloadWorker(
            report_file_id=report_file.id,
            worker_name=rec["worker_name"],
            area=rec["area"],
            grid=rec.get("grid", ""),
            workload=rec["workload"],
            total_backlog=rec["total_backlog"],
            total_today=rec["total_today"],
        )
        db.add(cw)
        worker_count += 1

    # 更新 column_hint
    report_type.column_hint = ["姓名", "区域", "装移拆", "投诉", "LAN口", "巡检", "一户一案", "质差弱光", "小计"]
    report_type.updated_at = datetime.now(timezone.utc)

    await db.commit()

    # ── 贾维斯：数据更新后自动触发 AI 分析 ──
    try:
        from app.ai.ai_scheduler import trigger_ai_analysis
        asyncio.create_task(trigger_ai_analysis("city-workload", db))
    except Exception:
        pass

    return {
        "summary_parsed": summary_count,
        "worker_count": worker_count,
        "filename": result["filename"],
        "report_date": result["report_date"],
    }


async def get_city_workload_summary(db: AsyncSession) -> dict:
    """获取全市装维工作量统计横山卡片汇总指标"""
    from app.core.models import CityWorkloadSummary

    stmt = select(CityWorkloadSummary).order_by(CityWorkloadSummary.id.desc()).limit(1)
    r = await db.execute(stmt)
    row = r.scalar_one_or_none()

    if not row:
        return {
            "district": "横山",
            "total_staff": "",
            "working_staff": "",
            "leave_staff": "",
            "no_work_ratio": "",
            "report_date": "",
            "latest_filename": "",
        }

    return {
        "district": row.district,
        "total_staff": row.total_staff,
        "working_staff": row.working_staff,
        "leave_staff": row.leave_staff,
        "no_work_ratio": row.no_work_ratio,
        "report_date": row.report_date,
        "latest_filename": getattr(row, "latest_filename", "") or "",
    }


async def get_city_workload_workers(db: AsyncSession) -> dict:
    """获取全市装维工作量统计横山装维人员工作量明细列表"""
    from app.core.models import CityWorkloadWorker
    from sqlalchemy import func as _func

    count_stmt = select(_func.count(CityWorkloadWorker.id))
    r = await db.execute(count_stmt)
    total = r.scalar() or 0

    data_stmt = (
        select(CityWorkloadWorker)
        .order_by(CityWorkloadWorker.total_backlog.desc())
    )
    r2 = await db.execute(data_stmt)
    rows = r2.scalars().all()

    workers = []
    for row in rows:
        workers.append({
            "id": row.id,
            "worker_name": row.worker_name,
            "area": row.area,
            "grid": row.grid or "",
            "workload": row.workload or {},
            "total_backlog": row.total_backlog,
            "total_today": row.total_today,
        })

    return {
        "workers": workers,
        "total": total,
    }


# ── 五类工单退撤单情况专用处理 ──

def _parse_five_category_withdrawal_files(directory: str) -> dict:
    """
    解析最新的"五类工单退撤单情况"文件：
    1. 从"通报1"sheet 提取横山日粒度和月粒度退撤指标
    2. 从"装机退撤"sheet 提取横山退撤单明细（筛选条件）
    返回 {
        "summary": dict | None,
        "details": [dict],
        "filename": str,
        "report_date": str,
    }
    """
    from datetime import datetime as _dt

    # 找到最新的五类工单退撤单情况文件
    matching: list[str] = []
    for root, _dirs, files in os.walk(directory):
        for f in files:
            if f.startswith("~$") or f.startswith("."):
                continue
            if "五类工单退撤单情况" in f and f.lower().endswith((".xlsx", ".xls")):
                matching.append(os.path.join(root, f))

    if not matching:
        return {"summary": None, "details": [], "filename": None, "report_date": None}

    # 按修改时间排序，取最新
    matching.sort(key=lambda f: os.path.getmtime(f), reverse=True)
    latest_file = matching[0]
    filename = os.path.basename(latest_file)

    try:
        wb = openpyxl.load_workbook(latest_file, read_only=True, data_only=True)

        summary = None
        report_date = None

        # ── 1. 解析"通报1"sheet（汇总指标）──
        if "通报1" in wb.sheetnames:
            ws = wb["通报1"]
            rows_data = list(ws.iter_rows(values_only=True))

            if rows_data:
                # 从标题行提取日期
                if rows_data[0] and rows_data[0][0]:
                    title = str(rows_data[0][0])
                    date_match = re.search(r'(\d{1,2})月(\d{1,2})日', title)
                    if date_match:
                        now = _dt.now()
                        report_date = f"{now.year}-{date_match.group(1).zfill(2)}-{date_match.group(2).zfill(2)}"

                # 日粒度：横山县在 rows_data[10] (索引10)，列3=退撤总量，列6=退撤单重装量总量
                # 月粒度：横山县在 rows_data[28] (索引28)，列3=退撤总量，列6=退撤单重装量总量
                # （注意：实际行索引可能因文件而异，这里用动态查找）

                # 动态查找横山县的行
                def find_hengshan_row(start_idx, end_idx):
                    for i in range(start_idx, min(end_idx, len(rows_data))):
                        row = rows_data[i]
                        if row and row[0] and "横山" in str(row[0]):
                            return row, i
                    return None, -1

                # 日粒度：从第5行开始找（跳过表头）
                day_row, _ = find_hengshan_row(5, 20)
                # 月粒度：从第23行开始找（月粒度数据区域）
                month_row, _ = find_hengshan_row(23, 40)

                day_withdrawal = ""
                day_reinstall = ""
                month_withdrawal = ""
                month_reinstall = ""

                if day_row and len(day_row) > 6:
                    day_withdrawal = str(day_row[3]).strip() if day_row[3] is not None else ""
                    day_reinstall = str(day_row[6]).strip() if day_row[6] is not None else ""

                if month_row and len(month_row) > 6:
                    month_withdrawal = str(month_row[3]).strip() if month_row[3] is not None else ""
                    month_reinstall = str(month_row[6]).strip() if month_row[6] is not None else ""

                summary = {
                    "district": "横山",
                    "day_withdrawal_total": day_withdrawal,
                    "day_reinstall_total": day_reinstall,
                    "month_withdrawal_total": month_withdrawal,
                    "month_reinstall_total": month_reinstall,
                }

                logger.info(f"五类工单退撤单情况汇总: 日退撤={day_withdrawal}, 日重装={day_reinstall}, 月退撤={month_withdrawal}, 月重装={month_reinstall}")

        # ── 2. 解析"装机退撤"sheet（明细数据）──
        details = []
        if "装机退撤" in wb.sheetnames:
            ws2 = wb["装机退撤"]

            # 列索引映射（基于实际Excel结构）
            COL_IDX = {
                'district': 32,        # 所属区县
                'scene': 97,            # 场景
                'tichong1': 2,          # 剔重1
                'huilao': 4,            # 是否回捞
                'account': 14,          # 宽带账号
                'global_access': 102,   # 全球通标识
                'service_type': 18,     # 服务类型
                'construction_address': 19,  # 施工地址
                'accept_department': 40,     # 受理部门
                'accept_time': 35,      # 受理时间
                'to_install_time': 37,  # 到装维时间
                'deadline': 52,         # 完成时限
                'natural_duration': 44, # 处理时长（自然时）
                'return_time': 53,      # 回单时间
                'archive_time': 54,     # 归档时间
                'return_note': 116,     # 回单备注信息
                'specific_reason': 8,   # 具体原因
            }

            # 遍历数据行（跳过表头行0）
            for i, row in enumerate(ws2.iter_rows(values_only=True)):
                if i == 0:
                    continue  # 跳过表头

                if not row or len(row) <= max(COL_IDX.values()):
                    continue

                # 获取筛选字段
                district_val = str(row[COL_IDX['district']]).strip() if row[COL_IDX['district']] else ''
                scene_val = str(row[COL_IDX['scene']]).strip() if row[COL_IDX['scene']] else ''
                tichong1_val = str(row[COL_IDX['tichong1']]).strip() if row[COL_IDX['tichong1']] else ''
                huilao_val = str(row[COL_IDX['huilao']]).strip() if row[COL_IDX['huilao']] else ''
                accept_time_val = str(row[COL_IDX['accept_time']]).strip() if row[COL_IDX['accept_time']] else ''

                # 筛选条件：所属区县=横山县，剔重1=正常
                if district_val != '横山县':
                    continue
                if tichong1_val != '正常':
                    continue
                # 场景筛选：如果场景字段有值，必须包含"家庭"
                if scene_val and '家庭' not in scene_val:
                    continue

                # 提取各字段
                def _get_val(key):
                    idx = COL_IDX.get(key)
                    if idx is not None and idx < len(row) and row[idx] is not None:
                        return str(row[idx]).strip()
                    return ''

                accept_time_str = _get_val('accept_time')
                deadline_str = _get_val('deadline')
                return_time_str = _get_val('return_time')

                # 计算疑似超时退单
                suspected_timeout = '未知'
                if return_time_str and deadline_str:
                    try:
                        return_dt = None
                        deadline_dt = None
                        for fmt in ['%Y-%m-%d %H:%M:%S', '%Y/%m/%d %H:%M:%S', '%Y-%m-%d', '%Y/%m/%d']:
                            try:
                                if not return_dt:
                                    return_dt = _dt.strptime(return_time_str[:19], fmt)
                            except (ValueError, TypeError):
                                pass
                            try:
                                if not deadline_dt:
                                    deadline_dt = _dt.strptime(deadline_str[:19], fmt)
                            except (ValueError, TypeError):
                                pass
                        if return_dt and deadline_dt:
                            diff = deadline_dt - return_dt
                            hours = diff.total_seconds() / 3600.0
                            suspected_timeout = '是' if hours < 12 else '否'
                    except Exception:
                        suspected_timeout = '未知'
                elif not return_time_str:
                    suspected_timeout = '未知'

                record = {
                    'district': district_val,
                    'account': _get_val('account'),
                    'global_access': _get_val('global_access'),
                    'service_type': _get_val('service_type'),
                    'construction_address': _get_val('construction_address'),
                    'accept_department': _get_val('accept_department'),
                    'accept_time': accept_time_str,
                    'to_install_time': _get_val('to_install_time'),
                    'deadline': deadline_str,
                    'natural_duration': _get_val('natural_duration'),
                    'return_time': return_time_str,
                    'archive_time': _get_val('archive_time'),
                    'suspected_timeout': suspected_timeout,
                    'return_note': _get_val('return_note'),
                    'specific_reason': _get_val('specific_reason'),
                }
                details.append(record)

            logger.info(f"五类工单退撤单情况明细: {filename} -> {len(details)} 条")

        wb.close()
        return {
            "summary": summary,
            "details": details,
            "filename": filename,
            "report_date": report_date,
        }

    except Exception as e:
        logger.error(f"解析五类工单退撤单情况失败 {filename}: {e}", exc_info=True)
        return {"summary": None, "details": [], "filename": filename, "report_date": None}



@record_notification("五类工单退撤单情况")
async def reparse_five_category_withdrawal(db: AsyncSession, directory: Optional[str] = None) -> dict:
    """
    重新解析五类工单退撤单情况文件，
    提取横山汇总指标和退撤单明细，更新数据库。
    """
    if directory is None:
        directory = settings.watch_dir

    from app.core.models import FiveCategoryWithdrawalSummary, FiveCategoryWithdrawalDetail
    from sqlalchemy import delete as _delete

    # 在线程池中解析
    result = await asyncio.to_thread(_parse_five_category_withdrawal_files, directory)

    # 删除旧数据
    await db.execute(_delete(FiveCategoryWithdrawalSummary))
    await db.execute(_delete(FiveCategoryWithdrawalDetail))

    summary_count = 0
    if result["summary"]:
        s = result["summary"]
        fcws = FiveCategoryWithdrawalSummary(
            report_date=result["report_date"] or "",
            district=s["district"],
            latest_filename=result.get("filename") or "",
            day_withdrawal_total=s["day_withdrawal_total"],
            day_reinstall_total=s["day_reinstall_total"],
            month_withdrawal_total=s["month_withdrawal_total"],
            month_reinstall_total=s["month_reinstall_total"],
        )
        db.add(fcws)
        summary_count = 1

    # 获取或创建"五类工单退撤单情况"报表类型
    stmt = select(ReportType).where(ReportType.name == "五类工单退撤单情况")
    r = await db.execute(stmt)
    report_type = r.scalar_one_or_none()
    if not report_type:
        report_type = ReportType(name="五类工单退撤单情况", category="装维生产")
        db.add(report_type)
        await db.flush()

    # 删除旧的 report_files/report_records
    old_files_stmt = select(ReportFile.id).where(ReportFile.report_type_id == report_type.id)
    r2 = await db.execute(old_files_stmt)
    old_file_ids = [row[0] for row in r2.all()]
    if old_file_ids:
        await db.execute(_delete(ReportRecord).where(ReportRecord.report_file_id.in_(old_file_ids)))
        await db.execute(_delete(ReportFile).where(ReportFile.report_type_id == report_type.id))

    # 创建 ReportFile
    report_file = ReportFile(
        report_type_id=report_type.id,
        filename=result["filename"] or "",
        file_path=os.path.join(directory, result["filename"] or ""),
        parse_status="parsed",
        record_count=len(result["details"]),
    )
    db.add(report_file)
    await db.flush()

    # 写入明细数据
    detail_count = 0
    for rec in result["details"]:
        fcd = FiveCategoryWithdrawalDetail(
            report_file_id=report_file.id,
            district=rec["district"],
            account=rec["account"],
            global_access=rec["global_access"],
            service_type=rec["service_type"],
            construction_address=rec["construction_address"],
            accept_department=rec["accept_department"],
            accept_time=rec["accept_time"],
            to_install_time=rec["to_install_time"],
            deadline=rec["deadline"],
            natural_duration=rec["natural_duration"],
            return_time=rec["return_time"],
            archive_time=rec["archive_time"],
            suspected_timeout=rec["suspected_timeout"],
            return_note=rec["return_note"],
            specific_reason=rec["specific_reason"],
        )
        db.add(fcd)
        detail_count += 1

    await db.commit()

    return {
        "summary_parsed": summary_count,
        "detail_count": detail_count,
        "filename": result["filename"],
        "report_date": result["report_date"],
    }


async def get_five_category_withdrawal_summary(db: AsyncSession) -> dict:
    """获取五类工单退撤单情况横山卡片指标"""
    from app.core.models import FiveCategoryWithdrawalSummary

    stmt = select(FiveCategoryWithdrawalSummary).order_by(FiveCategoryWithdrawalSummary.id.desc()).limit(1)
    r = await db.execute(stmt)
    row = r.scalar_one_or_none()

    if not row:
        return {
            "district": "横山",
            "day_withdrawal_total": "",
            "day_reinstall_total": "",
            "month_withdrawal_total": "",
            "month_reinstall_total": "",
            "report_date": "",
            "latest_filename": "",
        }

    return {
        "district": row.district,
        "day_withdrawal_total": row.day_withdrawal_total,
        "day_reinstall_total": row.day_reinstall_total,
        "month_withdrawal_total": row.month_withdrawal_total,
        "month_reinstall_total": row.month_reinstall_total,
        "report_date": row.report_date,
        "latest_filename": getattr(row, "latest_filename", "") or "",
    }


async def get_five_category_withdrawal_details(
    db: AsyncSession,
    page: int = 1,
    page_size: int = 50,
) -> dict:
    """分页获取五类工单退撤单情况横山退撤单明细"""
    from app.core.models import FiveCategoryWithdrawalDetail
    from sqlalchemy import func as _func

    count_stmt = select(_func.count(FiveCategoryWithdrawalDetail.id))
    r = await db.execute(count_stmt)
    total = r.scalar() or 0

    offset = (page - 1) * page_size
    data_stmt = (
        select(FiveCategoryWithdrawalDetail)
        .order_by(FiveCategoryWithdrawalDetail.id.desc())
        .offset(offset)
        .limit(page_size)
    )
    r2 = await db.execute(data_stmt)
    rows = r2.scalars().all()

    records = []
    for row in rows:
        records.append({
            "id": row.id,
            "district": row.district,
            "account": row.account,
            "global_access": row.global_access,
            "service_type": row.service_type,
            "construction_address": row.construction_address,
            "accept_department": row.accept_department,
            "accept_time": row.accept_time,
            "to_install_time": row.to_install_time,
            "deadline": row.deadline,
            "natural_duration": row.natural_duration,
            "return_time": row.return_time,
            "archive_time": row.archive_time,
            "suspected_timeout": row.suspected_timeout,
            "return_note": row.return_note,
            "specific_reason": row.specific_reason,
        })

    return {
        "records": records,
        "total": total,
        "page": page,
        "page_size": page_size,
    }


# ── 宽带在途投诉清单横山数据专用处理 ──
# 宽带在途投诉清单保留字段（来自"到县区"sheet）
COMPLAINT_BACKLOG_FIELDS = [
    "10086积压",      # backlog_10086
    "全球通积压",      # backlog_global
    "2200000积压",    # backlog_2200000
    "86线下积压",     # backlog_86_offline
    "合计",            # total_backlog
    "前一日积压量",   # previous_day_backlog
    "环比",            # ratio
]


def _parse_complaint_backlog_files(directory: str) -> dict:
    """
    解析最新的"宽带在途投诉清单"文件：
    从"到县区"sheet 提取横山在途投诉汇总数据。
    返回 {"summary": dict, "filename": str, "report_date": str}
    """
    from datetime import datetime as _dt

    # 找到最新的宽带在途投诉清单文件
    matching: list[str] = []
    for root, _dirs, files in os.walk(directory):
        for f in files:
            if f.startswith("~$") or f.startswith("."):
                continue
            if "宽带在途投诉清单" in f and f.lower().endswith((".xlsx", ".xls")):
                matching.append(os.path.join(root, f))

    if not matching:
        return {"summary": None, "filename": None, "report_date": None}

    # 按修改时间排序，取最新
    matching.sort(key=lambda f: os.path.getmtime(f), reverse=True)
    latest_file = matching[0]
    filename = os.path.basename(latest_file)

    # 尝试从文件名提取日期
    report_date = None
    date_match = re.search(r'(\d{4})[-年](\d{1,2})[-月](\d{1,2})', filename)
    if date_match:
        report_date = f"{date_match.group(1)}-{date_match.group(2).zfill(2)}-{date_match.group(3).zfill(2)}"
    else:
        report_date = _dt.now().strftime("%Y-%m-%d")

    try:
        wb = openpyxl.load_workbook(latest_file, read_only=True, data_only=True)

        # ── 解析"到县区"sheet ──
        summary = None
        if "到县区" in wb.sheetnames:
            ws = wb["到县区"]
            rows_data = list(ws.iter_rows(values_only=True))

            if len(rows_data) >= 2:
                # 查找表头行（包含"10086积压"或"区县"等关键字段）
                header_row_idx = None
                header_row = None
                for idx, row in enumerate(rows_data):
                    row_text = " ".join(str(c) for c in row if c)
                    if any(kw in row_text for kw in ["10086积压", "区县", "合计"]):
                        header_row_idx = idx
                        header_row = row
                        break

                if header_row_idx is not None:
                    # 建立列映射（合并表头可能跨两行，同时扫描子表头行和上一行）
                    col_map: dict[str, int] = {}
                    target_fields = ["区县", "县区", "10086积压", "全球通积压", "2200000积压", "86线下积压", "合计", "前一日积压量", "环比"]
                    
                    # 扫描子表头行（如行2）
                    for target in target_fields:
                        for idx, cell in enumerate(header_row):
                            if cell and target in str(cell).strip():
                                col_map[target] = idx
                                break
                    
                    # 扫描上一行（合并表头行，如行1）补全缺失的列
                    if header_row_idx > 0:
                        parent_header = rows_data[header_row_idx - 1]
                        for target in target_fields:
                            if target not in col_map:
                                for idx, cell in enumerate(parent_header):
                                    if cell and target in str(cell).strip():
                                        col_map[target] = idx
                                        break

                    # 查找横山行
                    for row in rows_data[header_row_idx + 1:]:
                        if not row or len(row) <= 1:
                            continue
                        
                        # 尝试多种方式匹配区县名
                        district_val = ""
                        for key in ["区县", "县区"]:
                            if key in col_map and col_map[key] < len(row) and row[col_map[key]]:
                                district_val = str(row[col_map[key]]).strip()
                                if district_val:
                                    break
                        
                        if district_val == "横山" or district_val == "横山县":
                            # 提取各字段值
                            def _safe_str(key: str) -> str:
                                idx = col_map.get(key)
                                if idx is not None and idx < len(row) and row[idx] is not None:
                                    val = row[idx]
                                    # 数值类型直接转字符串
                                    if isinstance(val, (int, float)):
                                        return str(int(val)) if isinstance(val, int) or val == int(val) else str(val)
                                    return str(val).strip()
                                return ""

                            summary = {
                                "district": "横山",
                                "backlog_10086": _safe_str("10086积压"),
                                "backlog_global": _safe_str("全球通积压"),
                                "backlog_2200000": _safe_str("2200000积压"),
                                "backlog_86_offline": _safe_str("86线下积压"),
                                "total_backlog": _safe_str("合计"),
                                "previous_day_backlog": _safe_str("前一日积压量"),
                                "ratio": _safe_str("环比"),
                            }
                            break

            wb.close()
            
            if summary:
                logger.info(f"宽带在途投诉清单横山: {filename} -> {summary}")
            else:
                logger.warning(f"宽带在途投诉清单: {filename} 未找到横山数据")
                
            return {
                "summary": summary,
                "filename": filename,
                "report_date": report_date,
            }
        else:
            logger.warning(f"宽带在途投诉清单: {filename} 缺少'到县区'sheet")
            wb.close()
            return {"summary": None, "filename": filename, "report_date": report_date}

    except Exception as e:
        logger.error(f"解析宽带在途投诉清单失败 {filename}: {e}", exc_info=True)
        return {"summary": None, "filename": filename, "report_date": report_date}


@record_notification("宽带在途投诉清单")
async def reparse_complaint_backlog(db: AsyncSession, directory: Optional[str] = None) -> dict:
    """
    重新解析宽带在途投诉清单文件，提取横山在途投诉汇总数据，更新数据库。
    删除旧数据并写入新数据。
    """
    if directory is None:
        directory = settings.watch_dir

    # 在线程池中解析
    result = await asyncio.to_thread(_parse_complaint_backlog_files, directory)

    # 导入模型
    from app.core.models import ComplaintBacklogSummary
    from sqlalchemy import delete as _delete

    # 删除旧汇总数据
    await db.execute(_delete(ComplaintBacklogSummary))

    summary_count = 0
    if result["summary"]:
        s = result["summary"]
        cbs = ComplaintBacklogSummary(
            report_date=result["report_date"] or "",
            district=s["district"],
            latest_filename=result.get("filename") or "",
            backlog_10086=s["backlog_10086"],
            backlog_global=s["backlog_global"],
            backlog_2200000=s["backlog_2200000"],
            backlog_86_offline=s["backlog_86_offline"],
            total_backlog=s["total_backlog"],
            previous_day_backlog=s["previous_day_backlog"],
            ratio=s["ratio"],
        )
        db.add(cbs)
        summary_count = 1

    # 获取或创建"宽带在途投诉清单"报表类型
    stmt = select(ReportType).where(ReportType.name == "宽带在途投诉清单")
    r = await db.execute(stmt)
    report_type = r.scalar_one_or_none()
    if not report_type:
        report_type = ReportType(name="宽带在途投诉清单", category="投诉")
        db.add(report_type)
        await db.flush()

    # 删除旧的 report_files/report_records
    old_files_stmt = select(ReportFile.id).where(ReportFile.report_type_id == report_type.id)
    r2 = await db.execute(old_files_stmt)
    old_file_ids = [row[0] for row in r2.all()]
    if old_file_ids:
        await db.execute(_delete(ReportRecord).where(ReportRecord.report_file_id.in_(old_file_ids)))
        await db.execute(_delete(ReportFile).where(ReportFile.report_type_id == report_type.id))

    # 创建 ReportFile
    if result["filename"]:
        report_file = ReportFile(
            report_type_id=report_type.id,
            filename=result["filename"] or "",
            file_path=os.path.join(directory, result["filename"] or ""),
            parse_status="parsed",
            record_count=1 if result["summary"] else 0,
        )
        db.add(report_file)

    # 更新 column_hint
    report_type.column_hint = COMPLAINT_BACKLOG_FIELDS.copy()
    report_type.updated_at = datetime.now(timezone.utc)

    await db.commit()

    return {
        "summary_parsed": summary_count,
        "filename": result["filename"],
        "report_date": result["report_date"],
    }


async def get_complaint_backlog_summary(db: AsyncSession) -> dict:
    """获取宽带在途投诉清单横山卡片指标"""
    from app.core.models import ComplaintBacklogSummary

    stmt = select(ComplaintBacklogSummary).order_by(ComplaintBacklogSummary.id.desc()).limit(1)
    r = await db.execute(stmt)
    row = r.scalar_one_or_none()

    if not row:
        return {
            "district": "横山",
            "backlog_10086": "",
            "backlog_global": "",
            "backlog_2200000": "",
            "backlog_86_offline": "",
            "total_backlog": "",
            "previous_day_backlog": "",
            "ratio": "",
            "report_date": "",
            "latest_filename": "",
        }

    return {
        "district": row.district,
        "backlog_10086": row.backlog_10086,
        "backlog_global": row.backlog_global,
        "backlog_2200000": row.backlog_2200000,
        "backlog_86_offline": row.backlog_86_offline,
        "total_backlog": row.total_backlog,
        "previous_day_backlog": row.previous_day_backlog,
        "ratio": row.ratio,
        "report_date": row.report_date,
        "latest_filename": getattr(row, "latest_filename", "") or "",
    }


# ── 10086投诉积压(督办)专用处理 ──

COMPLAINT_10086_SUMMARY_FIELDS = [
    "合计未超时积压", "今日需处理量", "家宽业务", "合计超时积压", "合计积压",
]

COMPLAINT_10086_DETAIL_FIELDS = [
    "所属区县", "超时时限", "宽带帐号", "全球通属性", "客户联系方式",
    "客户催单次数", "小区名称", "处理人姓名", "是否上门服务",
    "投诉分类5级", "回复内容",
]


def _parse_complaint_10086_files(directory: str) -> dict:
    """
    解析最新的"投诉积压通报新"文件：
    1. 从"表"sheet 提取横山汇总指标
    2. 从"10086积压清单"sheet 提取横山明细数据
    """
    from datetime import datetime as _dt, timedelta

    matching: list[str] = []
    for root, _dirs, files in os.walk(directory):
        for f in files:
            if f.startswith("~$") or f.startswith("."):
                continue
            if "投诉积压通报新" in f and f.lower().endswith((".xlsx", ".xls")):
                matching.append(os.path.join(root, f))

    if not matching:
        return {"summary": None, "details": [], "filename": None, "report_date": None}

    matching.sort(key=lambda f: os.path.getmtime(f), reverse=True)
    latest_file = matching[0]
    filename = os.path.basename(latest_file)

    # 尝试从文件名提取日期
    report_date = None
    date_match = re.search(r'(\d{1,2})\.(\d{1,2})', filename)
    if date_match:
        now_year = _dt.now().year
        report_date = f"{now_year}-{date_match.group(1).zfill(2)}-{date_match.group(2).zfill(2)}"
    else:
        report_date = _dt.now().strftime("%Y-%m-%d")

    try:
        wb = openpyxl.load_workbook(latest_file, read_only=True, data_only=True)

        # ── 实时计算变量（不依赖缓存的NOW()）──
        _now = _dt.now()
        _today_day = _now.day
        _tomorrow_day = (_now + timedelta(days=1)).day
        _rt_overdue = 0
        _rt_not_overdue = 0
        _rt_total = 0
        _rt_warn_2h = 0
        _rt_overdue_2_4h = 0
        _rt_today_need = 0
        _rt_broadband = 0
        _rt_parsed = False

        # ── 解析"表"sheet ──
        summary = None
        if "表" in wb.sheetnames:
            ws = wb["表"]
            rows_data = list(ws.iter_rows(values_only=True))

            # 找到包含"合计未超时积压"的子表头行
            header_row_idx = None
            header_row = None
            for idx, row in enumerate(rows_data):
                row_text = " ".join(str(c) for c in row if c)
                if "合计未超时积压" in row_text:
                    header_row_idx = idx
                    header_row = row
                    break

            if header_row_idx is not None:
                col_map: dict[str, int] = {}
                target_fields = ["区县", "县区", "合计未超时积压", "今日需处理量", "家宽业务", "合计超时积压", "合计积压"]

                for target in target_fields:
                    for idx2, cell in enumerate(header_row):
                        if cell and target in str(cell).strip():
                            col_map[target] = idx2
                            break

                # 扫描上一行补全（合并表头）
                if header_row_idx > 0:
                    parent_header = rows_data[header_row_idx - 1]
                    for target in target_fields:
                        if target not in col_map:
                            for idx2, cell in enumerate(parent_header):
                                if cell and target in str(cell).strip():
                                    col_map[target] = idx2
                                    break

                # 再上上行
                if header_row_idx > 1:
                    grand_header = rows_data[header_row_idx - 2]
                    for target in target_fields:
                        if target not in col_map:
                            for idx2, cell in enumerate(grand_header):
                                if cell and target in str(cell).strip():
                                    col_map[target] = idx2
                                    break

                # 特殊处理："预警2小时超时"和"2-4小时超时"在"含夜间"和"剔除夜间"两个区域都有
                # 需要匹配"10086积压（剔除夜间）"区域的列
                # 先定位"10086积压（剔除夜间）"父表头的列范围
                exclude_night_start = None
                exclude_night_end = None
                # 在所有父表头行中查找包含"10086积压"且"剔除夜间"的单元格
                for row_offset in [1, 2]:
                    if header_row_idx >= row_offset:
                        upper_row = rows_data[header_row_idx - row_offset]
                        for idx2, cell in enumerate(upper_row):
                            if cell:
                                cell_str = str(cell).strip()
                                if "10086" in cell_str and "剔除夜间" in cell_str:
                                    exclude_night_start = idx2
                                    break
                        if exclude_night_start is not None:
                            break

                # 找到"剔除夜间"区域的结束列（在同一行的下一个非空单元格 - 1）
                if exclude_night_start is not None:
                    # 使用 grand_header (Row 1) 来确定区域结束
                    end_header = rows_data[header_row_idx - 1] if header_row_idx > 0 else header_row
                    for idx2 in range(exclude_night_start + 1, len(end_header)):
                        if end_header[idx2] is not None and str(end_header[idx2]).strip():
                            exclude_night_end = idx2
                            break
                    if exclude_night_end is None:
                        exclude_night_end = len(header_row)

                    # 在"剔除夜间"区域内匹配"预警2小时超时"和"2-4小时超时"
                    for target in ["预警2小时超时", "2-4小时超时"]:
                        for idx2 in range(exclude_night_start, exclude_night_end):
                            if idx2 < len(header_row) and header_row[idx2] and target in str(header_row[idx2]).strip():
                                col_map[target] = idx2
                                break

                # 查找横山行
                for row in rows_data[header_row_idx + 1:]:
                    if not row or len(row) <= 1:
                        continue
                    district_val = ""
                    for key in ["区县", "县区"]:
                        if key in col_map and col_map[key] < len(row) and row[col_map[key]]:
                            district_val = str(row[col_map[key]]).strip()
                            if district_val:
                                break

                    if district_val in ("横山", "横山县"):
                        def _safe_str(key: str) -> str:
                            idx2 = col_map.get(key)
                            if idx2 is not None and idx2 < len(row) and row[idx2] is not None:
                                val = row[idx2]
                                if isinstance(val, (int, float)):
                                    return str(int(val)) if isinstance(val, int) or val == int(val) else str(val)
                                return str(val).strip()
                            return ""

                        summary = {
                            "district": "横山",
                            "total_not_overdue": _safe_str("合计未超时积压"),
                            "today_need_process": _safe_str("今日需处理量"),
                            "broadband_business": _safe_str("家宽业务"),
                            "total_overdue": _safe_str("合计超时积压"),
                            "total_backlog": _safe_str("合计积压"),
                            "warn_2h_overdue": _safe_str("预警2小时超时"),
                            "overdue_2_4h": _safe_str("2-4小时超时"),
                        }
                        break

        # ── 解析"10086积压清单"sheet ──
        details: list[dict] = []
        if "10086积压清单" in wb.sheetnames:
            ws2 = wb["10086积压清单"]
            rows2 = list(ws2.iter_rows(values_only=True))

            if len(rows2) >= 2:
                header2 = rows2[0]
                # 建立列映射
                col_map2: dict[str, int] = {}
                detail_fields = ["所属区县", "客服受理时间", "客服派单到装维时间",
                                 "超时时限", "8小时处理时限",
                                 "宽带帐号", "全球通属性",
                                 "客户联系方式", "客户催单次数", "小区名称",
                                 "处理人姓名", "是否上门服务",
                                 "投诉分类1级", "投诉分类2级",
                                 "投诉分类5级", "回复内容"]

                for target in detail_fields:
                    for idx2, cell in enumerate(header2):
                        if cell and target in str(cell).strip():
                            col_map2[target] = idx2
                            break

                # 提取横山数据
                _rt_detail_rows = []
                for row in rows2[1:]:
                    if not row or len(row) <= 2:
                        continue
                    district_val = ""
                    if "所属区县" in col_map2:
                        idx2 = col_map2["所属区县"]
                        if idx2 < len(row) and row[idx2]:
                            district_val = str(row[idx2]).strip()

                    if district_val in ("横山", "横山县"):
                        def _safe_str2(key: str) -> str:
                            idx3 = col_map2.get(key)
                            if idx3 is not None and idx3 < len(row) and row[idx3] is not None:
                                val = row[idx3]
                                if isinstance(val, (int, float)):
                                    return str(int(val)) if isinstance(val, int) or val == int(val) else str(val)
                                return str(val).strip()
                            return ""

                        details.append({
                            "district": "横山",
                            "timeout_deadline": _safe_str2("超时时限"),
                            "broadband_account": _safe_str2("宽带帐号"),
                            "global_access": _safe_str2("全球通属性"),
                            "customer_contact": _safe_str2("客户联系方式"),
                            "customer_urge_count": _safe_str2("客户催单次数"),
                            "community_name": _safe_str2("小区名称"),
                            "handler_name": _safe_str2("处理人姓名"),
                            "is_door_service": _safe_str2("是否上门服务"),
                            "complaint_category5": _safe_str2("投诉分类5级"),
                            "reply_content": _safe_str2("回复内容"),
                        })

                        # ── 收集原始datetime值用于实时计算 ──
                        _rt_row_data = {"timeout_val": None, "eight_val": None,
                                     "dispatch_val": None, "cat1": "", "cat2": ""}
                        _tc = col_map2.get("超时时限")
                        if _tc is not None and _tc < len(row) and row[_tc]:
                            _v = row[_tc]
                            if isinstance(_v, _dt):
                                _rt_row_data["timeout_val"] = _v
                            elif isinstance(_v, str):
                                try:
                                    _rt_row_data["timeout_val"] = _dt.strptime(_v, "%Y-%m-%d %H:%M:%S")
                                except Exception:
                                    pass
                        _ec = col_map2.get("8小时处理时限")
                        if _ec is not None and _ec < len(row) and row[_ec]:
                            _v = row[_ec]
                            if isinstance(_v, _dt):
                                _rt_row_data["eight_val"] = _v
                            elif isinstance(_v, str):
                                try:
                                    _rt_row_data["eight_val"] = _dt.strptime(_v, "%Y-%m-%d %H:%M:%S")
                                except Exception:
                                    pass
                        _dc = col_map2.get("客服派单到装维时间")
                        if _dc is not None and _dc < len(row) and row[_dc]:
                            _v = row[_dc]
                            if isinstance(_v, _dt):
                                _rt_row_data["dispatch_val"] = _v
                            elif isinstance(_v, str):
                                try:
                                    _rt_row_data["dispatch_val"] = _dt.strptime(_v, "%Y-%m-%d %H:%M:%S")
                                except Exception:
                                    pass
                        _c1 = col_map2.get("投诉分类1级")
                        if _c1 is not None and _c1 < len(row) and row[_c1]:
                            _rt_row_data["cat1"] = str(row[_c1]).strip()
                        _c2 = col_map2.get("投诉分类2级")
                        if _c2 is not None and _c2 < len(row) and row[_c2]:
                            _rt_row_data["cat2"] = str(row[_c2]).strip()
                        _rt_detail_rows.append(_rt_row_data)

        wb.close()

        # ── 用实时计算结果覆盖汇总值（避免缓存公式值过时）──
        if _rt_detail_rows and summary:
            _rt_total = len(_rt_detail_rows)
            _rt_overdue = 0
            _rt_not_overdue = 0
            _rt_warn_2h = 0
            _rt_overdue_2_4h = 0
            _rt_today_need = 0
            _rt_broadband = 0

            for _rd in _rt_detail_rows:
                _timeout_val = _rd["timeout_val"]
                _eight_val = _rd["eight_val"]
                _cat1 = _rd["cat1"]
                _cat2 = _rd["cat2"]

                if _timeout_val:
                    _is_not_overdue = _timeout_val > _now
                    if _is_not_overdue:
                        _rt_not_overdue += 1
                    else:
                        _rt_overdue += 1

                    _timeout_date = _timeout_val.date() if hasattr(_timeout_val, "date") else None
                    if _timeout_date and _is_not_overdue:
                        if _timeout_date.day in (_today_day, _tomorrow_day):
                            _rt_today_need += 1

                _process_candidates = [v for v in (_timeout_val, _eight_val) if v is not None]
                _process_deadline = min(_process_candidates) if _process_candidates else None

                if _process_deadline:
                    _pre_h = (_process_deadline - _now).total_seconds() / 3600
                    if 0 <= _pre_h < 2:
                        _rt_warn_2h += 1
                    if 2 <= _pre_h < 4:
                        _rt_overdue_2_4h += 1

                if "家宽" in f"{_cat1}{_cat2}" or "家庭" in f"{_cat1}{_cat2}":
                    _rt_broadband += 1

            summary["total_not_overdue"] = str(_rt_not_overdue)
            summary["total_overdue"] = str(_rt_overdue)
            summary["total_backlog"] = str(_rt_total)
            summary["warn_2h_overdue"] = str(_rt_warn_2h)
            summary["overdue_2_4h"] = str(_rt_overdue_2_4h)
            summary["today_need_process"] = str(_rt_today_need)
            summary["broadband_business"] = str(_rt_broadband)
            logger.info(f"10086投诉积压 实时计算覆盖: 未超时={_rt_not_overdue}, 超时={_rt_overdue}, 预警2h={_rt_warn_2h}, 2-4h={_rt_overdue_2_4h}")

        if summary:
            logger.info(f"10086投诉积压横山汇总: {filename} -> {summary}")
        else:
            logger.warning(f"10086投诉积压: {filename} 未找到横山汇总数据")
        logger.info(f"10086投诉积压横山明细: {len(details)} 条")

        return {
            "summary": summary,
            "details": details,
            "filename": filename,
            "report_date": report_date,
        }

    except Exception as e:
        logger.error(f"解析10086投诉积压失败 {filename}: {e}", exc_info=True)
        return {"summary": None, "details": [], "filename": filename, "report_date": report_date}


@record_notification("10086投诉积压")
async def reparse_complaint_10086(db: AsyncSession, directory: Optional[str] = None) -> dict:
    """
    重新解析10086投诉积压(督办)文件，提取横山汇总+明细数据，更新数据库。
    """
    if directory is None:
        directory = settings.watch_dir

    result = await asyncio.to_thread(_parse_complaint_10086_files, directory)

    from app.core.models import Complaint10086Summary, Complaint10086Detail
    from sqlalchemy import delete as _delete

    # 删除旧汇总+明细数据
    await db.execute(_delete(Complaint10086Summary))
    await db.execute(_delete(Complaint10086Detail))

    summary_count = 0
    if result["summary"]:
        s = result["summary"]
        cbs = Complaint10086Summary(
            report_date=result["report_date"] or "",
            district=s["district"],
            latest_filename=result.get("filename") or "",
            total_not_overdue=s["total_not_overdue"],
            today_need_process=s["today_need_process"],
            broadband_business=s["broadband_business"],
            total_overdue=s["total_overdue"],
            total_backlog=s["total_backlog"],
            warn_2h_overdue=s.get("warn_2h_overdue", ""),
            overdue_2_4h=s.get("overdue_2_4h", ""),
        )
        db.add(cbs)
        summary_count = 1

    # 获取或创建报表类型
    stmt = select(ReportType).where(ReportType.name == "10086投诉积压(督办)")
    r = await db.execute(stmt)
    report_type = r.scalar_one_or_none()
    if not report_type:
        report_type = ReportType(name="10086投诉积压(督办)", category="投诉")
        db.add(report_type)
        await db.flush()

    # 删除旧的 report_files/report_records
    old_files_stmt = select(ReportFile.id).where(ReportFile.report_type_id == report_type.id)
    r2 = await db.execute(old_files_stmt)
    old_file_ids = [row[0] for row in r2.all()]
    if old_file_ids:
        await db.execute(_delete(ReportRecord).where(ReportRecord.report_file_id.in_(old_file_ids)))
        await db.execute(_delete(ReportFile).where(ReportFile.report_type_id == report_type.id))

    # 创建 ReportFile
    detail_count = 0
    if result["filename"]:
        report_file = ReportFile(
            report_type_id=report_type.id,
            filename=result["filename"] or "",
            file_path=os.path.join(directory, result["filename"] or ""),
            parse_status="parsed",
            record_count=len(result["details"]),
        )
        db.add(report_file)
        await db.flush()
        detail_count = len(result["details"])

        # 写入明细
        for d in result["details"]:
            det = Complaint10086Detail(
                report_file_id=report_file.id,
                district=d["district"],
                timeout_deadline=d["timeout_deadline"],
                broadband_account=d["broadband_account"],
                global_access=d["global_access"],
                customer_contact=d["customer_contact"],
                customer_urge_count=d["customer_urge_count"],
                community_name=d["community_name"],
                handler_name=d["handler_name"],
                is_door_service=d["is_door_service"],
                complaint_category5=d["complaint_category5"],
                reply_content=d["reply_content"],
            )
            db.add(det)

    # 更新 column_hint
    report_type.column_hint = COMPLAINT_10086_DETAIL_FIELDS.copy()
    report_type.updated_at = datetime.now(timezone.utc)

    await db.commit()

    return {
        "summary_parsed": summary_count,
        "detail_parsed": detail_count,
        "filename": result["filename"],
        "report_date": result["report_date"],
    }


async def get_complaint_10086_summary(db: AsyncSession) -> dict:
    """获取10086投诉积压(督办)横山卡片指标"""
    from app.core.models import Complaint10086Summary

    stmt = select(Complaint10086Summary).order_by(Complaint10086Summary.id.desc()).limit(1)
    r = await db.execute(stmt)
    row = r.scalar_one_or_none()

    if not row:
        return {
            "district": "横山",
            "total_not_overdue": "",
            "today_need_process": "",
            "broadband_business": "",
            "total_overdue": "",
            "total_backlog": "",
            "warn_2h_overdue": "",
            "overdue_2_4h": "",
            "report_date": "",
            "latest_filename": "",
        }

    return {
        "district": row.district,
        "total_not_overdue": row.total_not_overdue,
        "today_need_process": row.today_need_process,
        "broadband_business": row.broadband_business,
        "total_overdue": row.total_overdue,
        "total_backlog": row.total_backlog,
        "warn_2h_overdue": row.warn_2h_overdue or "",
        "overdue_2_4h": row.overdue_2_4h or "",
        "report_date": row.report_date,
        "latest_filename": getattr(row, "latest_filename", "") or "",
    }


async def get_complaint_10086_details(
    db: AsyncSession,
    page: int = 1,
    page_size: int = 50,
    sort_by: Optional[str] = None,
    order: str = "asc",
) -> dict:
    """分页获取10086投诉积压(督办)横山10086积压清单明细
    - sort_by: timeout_deadline 时按超时时限排序；其他按 id 排序
    - order: asc=升序，desc=降序
    """
    from app.core.models import Complaint10086Detail
    from sqlalchemy import func as _func

    count_stmt = select(_func.count(Complaint10086Detail.id))
    r = await db.execute(count_stmt)
    total = r.scalar() or 0

    offset = (page - 1) * page_size
    data_stmt = select(Complaint10086Detail)

    # 动态排序
    if sort_by and hasattr(Complaint10086Detail, sort_by):
        col = getattr(Complaint10086Detail, sort_by)
        if order == "desc":
            data_stmt = data_stmt.order_by(col.desc())
        else:
            data_stmt = data_stmt.order_by(col.asc())
    else:
        data_stmt = data_stmt.order_by(Complaint10086Detail.id.asc())
    r2 = await db.execute(data_stmt)
    rows = r2.scalars().all()

    records = []
    for row in rows:
        records.append({
            "id": row.id,
            "district": row.district,
            "timeout_deadline": row.timeout_deadline,
            "broadband_account": row.broadband_account,
            "global_access": row.global_access,
            "customer_contact": row.customer_contact,
            "customer_urge_count": row.customer_urge_count,
            "community_name": row.community_name,
            "handler_name": row.handler_name,
            "is_door_service": row.is_door_service,
            "complaint_category5": row.complaint_category5,
            "reply_content": row.reply_content,
        })

    return {
        "records": records,
        "total": total,
        "page": page,
        "page_size": page_size,
    }


# ── 2200000及时率通报 ──

def _parse_complaint_2200000_files(directory: str) -> dict:
    """
    解析最新的"2200000及时率通报"文件：
    1. 从"通报"sheet 提取横山汇总指标
    2. 从"累计在途需处理（自接）"sheet 提取横山在途明细（剔重=正常）
    3. 从"往月积压"sheet 提取横山往月明细
    """
    from datetime import datetime as _dt, timedelta

    matching: list[str] = []
    for root, _dirs, files in os.walk(directory):
        for f in files:
            if f.startswith("~$") or f.startswith("."):
                continue
            if "2200000及时率" in f and f.lower().endswith((".xlsx", ".xls")):
                matching.append(os.path.join(root, f))

    if not matching:
        return {"summary": None, "details": [], "filename": None, "report_date": None}

    matching.sort(key=lambda f: os.path.getmtime(f), reverse=True)
    latest_file = matching[0]
    filename = os.path.basename(latest_file)

    # 尝试从文件名提取日期
    report_date = None
    date_match = re.search(r'(\d{1,2})\.(\d{1,2})', filename)
    if date_match:
        now_year = _dt.now().year
        report_date = f"{now_year}-{date_match.group(1).zfill(2)}-{date_match.group(2).zfill(2)}"
    else:
        report_date = _dt.now().strftime("%Y-%m-%d")

    try:
        wb = openpyxl.load_workbook(latest_file, read_only=True, data_only=True)

        # ── 解析"通报"sheet ──
        summary = None
        if "通报" in wb.sheetnames:
            ws = wb["通报"]
            rows_data = list(ws.iter_rows(values_only=True))

            # 找到包含"月派单量"的子表头行
            header_row_idx = None
            header_row = None
            for idx, row in enumerate(rows_data):
                row_text = " ".join(str(c) for c in row if c)
                if "月派单量" in row_text:
                    header_row_idx = idx
                    header_row = row
                    break

            if header_row_idx is not None:
                col_map: dict[str, int] = {}
                target_fields = ["县区", "月派单量", "超时积压", "未超时积压", "累计在途", "往月积压",
                                 "预警4小时超时", "升级投诉量"]

                for target in target_fields:
                    for idx2, cell in enumerate(header_row):
                        if cell and target in str(cell).strip():
                            col_map[target] = idx2
                            break

                # 扫描上一行补全
                if header_row_idx > 0:
                    parent_header = rows_data[header_row_idx - 1]
                    for target in target_fields:
                        if target not in col_map:
                            for idx2, cell in enumerate(parent_header):
                                if cell and target in str(cell).strip():
                                    col_map[target] = idx2
                                    break

                # 再上上行
                if header_row_idx > 1:
                    grand_header = rows_data[header_row_idx - 2]
                    for target in target_fields:
                        if target not in col_map:
                            for idx2, cell in enumerate(grand_header):
                                if cell and target in str(cell).strip():
                                    col_map[target] = idx2
                                    break

                # 查找横山行
                for row in rows_data[header_row_idx + 1:]:
                    if not row or len(row) <= 1:
                        continue
                    district_val = ""
                    district_key = "县区" if "县区" in col_map else None
                    if district_key and col_map[district_key] < len(row) and row[col_map[district_key]]:
                        district_val = str(row[col_map[district_key]]).strip()

                    if district_val in ("横山", "横山县"):
                        def _safe_str(key: str) -> str:
                            idx2 = col_map.get(key)
                            if idx2 is not None and idx2 < len(row) and row[idx2] is not None:
                                val = row[idx2]
                                if isinstance(val, (int, float)):
                                    return str(int(val)) if isinstance(val, int) or val == int(val) else str(val)
                                return str(val).strip()
                            return ""

                        summary = {
                            "district": "横山",
                            "monthly_dispatch": _safe_str("月派单量"),
                            "overdue_backlog": _safe_str("超时积压"),
                            "not_overdue_backlog": _safe_str("未超时积压"),
                            "total_in_transit": _safe_str("累计在途"),
                            "previous_month_backlog": _safe_str("往月积压"),
                            "warn_4h_overdue": _safe_str("预警4小时超时"),
                            "escalate_complaint": _safe_str("升级投诉量"),
                        }
                        break

        # ── 解析"累计在途需处理（自接）"sheet ──
        details: list[dict] = []
        # 用实时时间计算超时状态（不依赖缓存公式值）
        _now = _dt.now()
        _overdue_cnt = 0
        _not_overdue_cnt = 0
        _warn_4h_cnt = 0
        _parsed_detail = False
        sheet1_name = None
        for sn in wb.sheetnames:
            if "累计在途" in sn and "自接" in sn:
                sheet1_name = sn
                break

        if sheet1_name:
            _parsed_detail = True
            ws1 = wb[sheet1_name]
            rows1 = list(ws1.iter_rows(values_only=True))

            if len(rows1) >= 2:
                header1 = rows1[0]
                # 建立列映射 - 使用动态匹配
                col_map1: dict[str, int] = {}
                detail_fields1 = ["所属区县", "超时时限", "客服受理时间", "宽带帐号", "是否重要客户",
                                  "客户联系方式", "施工地址", "处理人姓名", "剔重"]

                for target in detail_fields1:
                    for idx2, cell in enumerate(header1):
                        if cell and target in str(cell).strip():
                            col_map1[target] = idx2
                            break

                # 注意"超时时限"可能出现两次（列4和列29），需要最后一个（AD列）
                for idx2, cell in enumerate(header1):
                    if cell and "超时时限" in str(cell).strip():
                        col_map1["超时时限"] = idx2

                accept_time_col = col_map1.get("客服受理时间")

                for row in rows1[1:]:
                    if not row or len(row) <= 2:
                        continue
                    district_val = ""
                    if "所属区县" in col_map1:
                        idx2 = col_map1["所属区县"]
                        if idx2 < len(row) and row[idx2]:
                            district_val = str(row[idx2]).strip()

                    if district_val not in ("横山", "横山县"):
                        continue

                    # 检查剔重
                    tichong_val = ""
                    if "剔重" in col_map1:
                        idx2 = col_map1["剔重"]
                        if idx2 < len(row) and row[idx2]:
                            tichong_val = str(row[idx2]).strip()

                    if tichong_val != "正常":
                        continue

                    # 用当前实时时间计算超时状态（不依赖缓存的公式值）
                    accept_time = None
                    if accept_time_col is not None and accept_time_col < len(row) and row[accept_time_col]:
                        val = row[accept_time_col]
                        if isinstance(val, _dt):
                            accept_time = val
                        elif isinstance(val, str):
                            try:
                                accept_time = _dt.strptime(val, "%Y-%m-%d %H:%M:%S")
                            except:
                                pass

                    is_timeout = False
                    is_warn_4h = False
                    if accept_time:
                        # 超时时限 = 客服受理时间 + 1天
                        deadline_24h = accept_time + timedelta(days=1)
                        is_timeout = _now > deadline_24h
                        # 预警4小时 = (客服受理时间 + 2天 - NOW) * 24 < 4
                        deadline_48h = accept_time + timedelta(hours=48)
                        remaining = (deadline_48h - _now).total_seconds() / 3600
                        is_warn_4h = 0 <= remaining < 4

                    if accept_time:
                        if is_timeout:
                            _overdue_cnt += 1
                        else:
                            _not_overdue_cnt += 1
                        if is_warn_4h:
                            _warn_4h_cnt += 1

                    def _safe_str1(key: str) -> str:
                        idx3 = col_map1.get(key)
                        if idx3 is not None and idx3 < len(row) and row[idx3] is not None:
                            val = row[idx3]
                            if isinstance(val, _dt):
                                return val.strftime("%Y-%m-%d %H:%M:%S")
                            return str(val).strip()
                        return ""

                    details.append({
                        "district": "横山",
                        "timeout_deadline": _safe_str1("超时时限"),
                        "broadband_account": _safe_str1("宽带帐号"),
                        "is_important_customer": _safe_str1("是否重要客户"),
                        "customer_contact": _safe_str1("客户联系方式"),
                        "construction_address": _safe_str1("施工地址"),
                        "handler_name": _safe_str1("处理人姓名"),
                        "category": "在途",
                    })

        # ── 解析"往月积压"sheet ──
        if "往月积压" in wb.sheetnames:
            ws2 = wb["往月积压"]
            rows2 = list(ws2.iter_rows(values_only=True))

            if len(rows2) >= 2:
                header2 = rows2[0]
                col_map2: dict[str, int] = {}
                detail_fields2 = ["所属区县", "超时时限", "宽带帐号", "是否重要客户",
                                  "客户联系方式", "施工地址", "处理人姓名"]

                for target in detail_fields2:
                    for idx2, cell in enumerate(header2):
                        if cell and target in str(cell).strip():
                            col_map2[target] = idx2
                            break

                for row in rows2[1:]:
                    if not row or len(row) <= 2:
                        continue
                    district_val = ""
                    if "所属区县" in col_map2:
                        idx2 = col_map2["所属区县"]
                        if idx2 < len(row) and row[idx2]:
                            district_val = str(row[idx2]).strip()

                    if district_val not in ("横山", "横山县"):
                        continue

                    def _safe_str2(key: str) -> str:
                        idx3 = col_map2.get(key)
                        if idx3 is not None and idx3 < len(row) and row[idx3] is not None:
                            val = row[idx3]
                            if isinstance(val, _dt):
                                return val.strftime("%Y-%m-%d %H:%M:%S")
                            return str(val).strip()
                        return ""

                    details.append({
                        "district": "横山",
                        "timeout_deadline": _safe_str2("超时时限"),
                        "broadband_account": _safe_str2("宽带帐号"),
                        "is_important_customer": _safe_str2("是否重要客户"),
                        "customer_contact": _safe_str2("客户联系方式"),
                        "construction_address": _safe_str2("施工地址"),
                        "handler_name": _safe_str2("处理人姓名"),
                        "category": "往月",
                    })

        wb.close()

        # 用实时计算的明细数据覆盖汇总值（避免缓存公式值过时）
        if _parsed_detail and summary:
            summary["overdue_backlog"] = str(_overdue_cnt)
            summary["not_overdue_backlog"] = str(_not_overdue_cnt)
            summary["total_in_transit"] = str(_overdue_cnt + _not_overdue_cnt)
            summary["warn_4h_overdue"] = str(_warn_4h_cnt)
            _prev_cnt = sum(1 for d in details if d["category"] == "往月")
            summary["previous_month_backlog"] = str(_prev_cnt)

        if summary:
            logger.info(f"2200000及时率横山汇总: {filename} -> {summary}")
        else:
            logger.warning(f"2200000及时率: {filename} 未找到横山汇总数据")
        logger.info(f"2200000及时率横山明细: {len(details)} 条")

        return {
            "summary": summary,
            "details": details,
            "filename": filename,
            "report_date": report_date,
        }

    except Exception as e:
        logger.error(f"解析2200000及时率通报失败 {filename}: {e}", exc_info=True)
        return {"summary": None, "details": [], "filename": filename, "report_date": report_date}


@record_notification("2200000及时率通报")
async def reparse_complaint_2200000(db: AsyncSession, directory: Optional[str] = None) -> dict:
    """重新解析2200000及时率通报文件，提取横山汇总+明细数据，更新数据库。"""
    if directory is None:
        directory = settings.watch_dir

    result = await asyncio.to_thread(_parse_complaint_2200000_files, directory)

    from app.core.models import Complaint2200000Summary, Complaint2200000Detail
    from sqlalchemy import delete as _delete

    # 删除旧汇总+明细数据
    await db.execute(_delete(Complaint2200000Summary))
    await db.execute(_delete(Complaint2200000Detail))

    summary_count = 0
    if result["summary"]:
        s = result["summary"]
        cbs = Complaint2200000Summary(
            report_date=result["report_date"] or "",
            district=s["district"],
            latest_filename=result.get("filename") or "",
            monthly_dispatch=s["monthly_dispatch"],
            overdue_backlog=s["overdue_backlog"],
            not_overdue_backlog=s["not_overdue_backlog"],
            total_in_transit=s["total_in_transit"],
            previous_month_backlog=s["previous_month_backlog"],
            warn_4h_overdue=s["warn_4h_overdue"],
            escalate_complaint=s["escalate_complaint"],
        )
        db.add(cbs)
        summary_count = 1

    # 获取或创建报表类型
    stmt = select(ReportType).where(ReportType.name == "2200000及时率通报")
    r = await db.execute(stmt)
    report_type = r.scalar_one_or_none()
    if not report_type:
        report_type = ReportType(name="2200000及时率通报", category="投诉")
        db.add(report_type)
        await db.flush()

    # 删除旧的 report_files/report_records
    old_files_stmt = select(ReportFile.id).where(ReportFile.report_type_id == report_type.id)
    r2 = await db.execute(old_files_stmt)
    old_file_ids = [row[0] for row in r2.all()]
    if old_file_ids:
        await db.execute(_delete(ReportRecord).where(ReportRecord.report_file_id.in_(old_file_ids)))
        await db.execute(_delete(ReportFile).where(ReportFile.report_type_id == report_type.id))

    # 创建 ReportFile
    detail_count = 0
    if result["filename"]:
        report_file = ReportFile(
            report_type_id=report_type.id,
            filename=result["filename"] or "",
            file_path=os.path.join(directory, result["filename"] or ""),
            parse_status="parsed",
            record_count=len(result["details"]),
        )
        db.add(report_file)
        await db.flush()
        detail_count = len(result["details"])

        # 写入明细
        for d in result["details"]:
            detail = Complaint2200000Detail(
                report_file_id=report_file.id,
                district=d["district"],
                timeout_deadline=d["timeout_deadline"],
                broadband_account=d["broadband_account"],
                is_important_customer=d["is_important_customer"],
                customer_contact=d["customer_contact"],
                construction_address=d["construction_address"],
                handler_name=d["handler_name"],
                category=d["category"],
            )
            db.add(detail)

    await db.commit()

    # ── 贾维斯：数据更新后自动触发 AI 分析 ──
    try:
        from app.ai.ai_scheduler import trigger_ai_analysis
        asyncio.create_task(trigger_ai_analysis("complaint_2200000", db))
    except Exception:
        pass

    return {
        "summary_parsed": summary_count,
        "detail_parsed": detail_count,
        "filename": result["filename"],
        "report_date": result["report_date"],
    }


async def get_complaint_2200000_summary(db: AsyncSession) -> dict:
    """获取2200000及时率通报横山卡片指标"""
    from app.core.models import Complaint2200000Summary

    stmt = select(Complaint2200000Summary).order_by(Complaint2200000Summary.id.desc()).limit(1)
    r = await db.execute(stmt)
    row = r.scalar_one_or_none()

    if not row:
        return {
            "district": "横山",
            "monthly_dispatch": "",
            "overdue_backlog": "",
            "not_overdue_backlog": "",
            "total_in_transit": "",
            "previous_month_backlog": "",
            "warn_4h_overdue": "",
            "escalate_complaint": "",
            "report_date": "",
            "latest_filename": "",
        }

    return {
        "district": row.district,
        "monthly_dispatch": row.monthly_dispatch,
        "overdue_backlog": row.overdue_backlog,
        "not_overdue_backlog": row.not_overdue_backlog,
        "total_in_transit": row.total_in_transit,
        "previous_month_backlog": row.previous_month_backlog,
        "warn_4h_overdue": row.warn_4h_overdue,
        "escalate_complaint": row.escalate_complaint,
        "report_date": row.report_date,
        "latest_filename": getattr(row, "latest_filename", "") or "",
    }


async def get_complaint_2200000_details(
    db: AsyncSession,
    page: int = 1,
    page_size: int = 50,
    sort_by: Optional[str] = None,
    order: str = "asc",
) -> dict:
    """分页获取2200000及时率通报横山明细
    - sort_by: 排序字段名（对应 Complaint2200000Detail 的列名），None 时不排序
    - order: asc=升序，desc=降序
    """
    from app.core.models import Complaint2200000Detail
    from sqlalchemy import func as _func

    count_stmt = select(_func.count(Complaint2200000Detail.id))
    r = await db.execute(count_stmt)
    total = r.scalar() or 0

    offset = (page - 1) * page_size
    data_stmt = select(Complaint2200000Detail)

    # 动态排序
    if sort_by and hasattr(Complaint2200000Detail, sort_by):
        col = getattr(Complaint2200000Detail, sort_by)
        if order == "desc":
            data_stmt = data_stmt.order_by(col.desc())
        else:
            data_stmt = data_stmt.order_by(col.asc())
    else:
        data_stmt = data_stmt.order_by(Complaint2200000Detail.id.asc())

    data_stmt = data_stmt.offset(offset).limit(page_size)
    r2 = await db.execute(data_stmt)
    rows = r2.scalars().all()

    records = []
    for row in rows:
        records.append({
            "id": row.id,
            "district": row.district,
            "timeout_deadline": row.timeout_deadline,
            "broadband_account": row.broadband_account,
            "is_important_customer": row.is_important_customer,
            "customer_contact": row.customer_contact,
            "construction_address": row.construction_address,
            "handler_name": row.handler_name,
            "category": row.category,
        })

    return {
        "records": records,
        "total": total,
        "page": page,
        "page_size": page_size,
    }


# ── 线下派单处理情况 ──

def _parse_offline_dispatch_files(directory: str) -> dict:
    """解析线下派单处理情况报表
    - 月派单量：从"通报"sheet读取（=累计在途+归档量，不受NOW()影响）
    - 超时积压/未超时积压/累计在途/预警4h超时：从"累计在途需处理（自接）"sheet实时计算
      （避免col7=是否超时、col2=预警时限等含NOW()公式的缓存值过时）
    """
    import re as _re
    from datetime import datetime as _dt, timedelta as _td, timezone as _tz

    pattern = os.path.join(directory, "**", "*线下派单处理情况*")
    files = glob.glob(pattern, recursive=True)
    if not files:
        return {"summary": None, "report_date": "", "filename": ""}

    latest_file = max(files, key=os.path.getmtime)
    filename = os.path.basename(latest_file)
    details = []

    wb = openpyxl.load_workbook(latest_file, data_only=True)

    if "通报" not in wb.sheetnames:
        logger.warning(f"线下派单处理情况: 文件 {filename} 缺少'通报'sheet")
        return {"summary": None, "report_date": "", "filename": filename}

    ws = wb["通报"]

    # 提取通报日期
    report_date = ""
    title_cell = ws.cell(row=1, column=1).value
    if title_cell:
        m = _re.search(r'截止(\d+)月(\d+)日', str(title_cell))
        if m:
            report_date = f"2026-{m.group(1).zfill(2)}-{m.group(2).zfill(2)}"

    # 找横山区数据行
    district_filter = "横山"
    hengshan_row = None
    for row in range(5, ws.max_row + 1):
        cell_val = ws.cell(row=row, column=1).value
        if cell_val and district_filter in str(cell_val):
            hengshan_row = row
            break

    if not hengshan_row:
        logger.warning(f"线下派单处理情况: 未找到横山区数据")
        return {"summary": None, "report_date": report_date, "filename": filename}

    # 月派单量：从通报sheet读取（=累计在途+归档量，不受NOW()影响）
    def _val(r, c):
        v = ws.cell(row=r, column=c).value
        return str(v) if v is not None else ""

    monthly_dispatch = _val(hengshan_row, 5)

    # ── 实时计算：超时积压/未超时积压/累计在途/预警4h超时 ──
    # 从"累计在途需处理（自接）"sheet实时计算
    # Excel公式链：
    #   col20 = 客服受理时间（固定值）
    #   col5  = 超时时限 = 客服受理时间 + 1天
    #   col8  = 48h超时时限 = 客服受理时间 + 2天
    #   col6  = 当前时间 = NOW()（公式，缓存值过时）
    #   col7  = 是否超时(24h) = IF(col6>col5,"是","否")（公式，缓存值过时）
    #   col2  = 预警超时时限 = (col8-col6)*24（公式，缓存值过时）
    # 通报sheet：
    #   超时积压 = COUNTIFS(区县=横山, col7="是", col72="装维调度系统")
    #   未超时积压 = COUNTIFS(区县=横山, col7="否", col72="装维调度系统")
    #   预警4h超时 = COUNTIFS(区县=横山, col2>=0, col2<4)
    #
    # 实时计算方式：用 客服受理时间 + timedelta(days=1) vs datetime.now() 判断超时
    overdue_backlog = ""
    not_overdue_backlog = ""
    total_in_transit = ""
    warn_4h_overdue = ""

    if "累计在途需处理（自接）" in wb.sheetnames:
        ws_detail = wb["累计在途需处理（自接）"]
        BJ_TZ = _tz(_td(hours=8))
        now_bj = _dt.now(BJ_TZ)

        _overdue_cnt = 0
        _not_overdue_cnt = 0
        _warn_4h_cnt = 0

        for row in range(2, ws_detail.max_row + 1):
            district = ws_detail.cell(row=row, column=16).value
            if not district or district_filter not in str(district):
                continue
            source = ws_detail.cell(row=row, column=72).value
            if source != "装维调度系统":
                continue

            # 客服受理时间 (col20)
            accept_time_raw = ws_detail.cell(row=row, column=20).value
            if accept_time_raw is None:
                continue

            # 解析时间
            if isinstance(accept_time_raw, _dt):
                accept_time = accept_time_raw
            elif isinstance(accept_time_raw, str):
                try:
                    accept_time = _dt.strptime(accept_time_raw, "%Y-%m-%d %H:%M:%S")
                except ValueError:
                    try:
                        accept_time = _dt.strptime(accept_time_raw, "%Y-%m-%d %H:%M")
                    except ValueError:
                        continue
            else:
                continue

            if accept_time.tzinfo is None:
                accept_time = accept_time.replace(tzinfo=BJ_TZ)

            # 24h超时判断：超时时限 = 受理时间 + 1天
            deadline_24h = accept_time + _td(days=1)
            is_overdue_24h = now_bj > deadline_24h

            if is_overdue_24h:
                _overdue_cnt += 1
            else:
                _not_overdue_cnt += 1

            # 预警4h超时：距离48h时限不到4小时且未超48h
            deadline_48h = accept_time + _td(days=2)
            hours_to_48h = (deadline_48h - now_bj).total_seconds() / 3600
            if 0 <= hours_to_48h < 4:
                _warn_4h_cnt += 1

        overdue_backlog = str(_overdue_cnt)
        not_overdue_backlog = str(_not_overdue_cnt)
        total_in_transit = str(_overdue_cnt + _not_overdue_cnt)
        warn_4h_overdue = str(_warn_4h_cnt)
        logger.info(
            f"线下派单实时计算: 超时={_overdue_cnt}, 未超时={_not_overdue_cnt}, "
            f"累计={total_in_transit}, 预警4h={_warn_4h_cnt}"
        )
    else:
        # fallback：用通报sheet的缓存值
        logger.warning("线下派单: 缺少'累计在途需处理（自接）'sheet，使用通报sheet缓存值")
        overdue_backlog = _val(hengshan_row, 6)
        not_overdue_backlog = _val(hengshan_row, 7)
        total_in_transit = _val(hengshan_row, 8)
        warn_4h_overdue = _val(hengshan_row, 10)

    summary = {
        "district": district_filter,
        "monthly_dispatch": monthly_dispatch,
        "overdue_backlog": overdue_backlog,
        "not_overdue_backlog": not_overdue_backlog,
        "total_in_transit": total_in_transit,
        "warn_4h_overdue": warn_4h_overdue,
    }

    # ── 解析明细：累计在途需处理（自接）+ 往月积压 ──
    def _build_col_map(ws_detail) -> dict:
        """从首行构建 列名→索引(0-based) 映射"""
        m = {}
        for i, cell in enumerate(ws_detail[1]):
            if cell.value:
                m[str(cell.value).strip()] = i
        return m

    def _safe_str(row, idx):
        if idx is None or idx >= len(row) or row[idx] is None:
            return ""
        return str(row[idx]).strip()

    for _sheet_name, _category in [("累计在途需处理（自接）", "在途"), ("往月积压", "往月")]:
        if _sheet_name not in wb.sheetnames:
            logger.warning(f"线下派单: 缺少'{_sheet_name}'sheet")
            continue
        _ws = wb[_sheet_name]
        _col_map = _build_col_map(_ws)
        _rows = list(_ws.iter_rows(values_only=True))
        if len(_rows) <= 1:
            continue

        _col_district = _col_map.get("所属区县")
        _col_tichong  = _col_map.get("剔重")
        _col_timeout   = _col_map.get("超时时限")
        _col_account   = _col_map.get("宽带帐号")
        _col_vip      = _col_map.get("是否重要客户")
        _col_contact   = _col_map.get("客户联系方式")
        _col_address   = _col_map.get("施工地址")
        _col_handler   = _col_map.get("处理人姓名")

        for _row in _rows[1:]:
            _district = _safe_str(_row, _col_district)
            if not _district or "横山" not in _district:
                continue
            # 剔重="正常" 过滤（列不存在时跳过该过滤）
            if _col_tichong is not None:
                if _safe_str(_row, _col_tichong) != "正常":
                    continue
            details.append({
                "district":           _district,
                "timeout_limit":       _safe_str(_row, _col_timeout),
                "broadband_account":  _safe_str(_row, _col_account),
                "is_vip_customer":    _safe_str(_row, _col_vip),
                "customer_contact":    _safe_str(_row, _col_contact),
                "construction_address": _safe_str(_row, _col_address),
                "handler_name":        _safe_str(_row, _col_handler),
                "category":            _category,
            })

    logger.info(f"线下派单处理情况: 解析完成, report_date={report_date}, summary={summary}, details={len(details)}条")
    wb.close()
    return {"summary": summary, "details": details, "report_date": report_date, "filename": filename}


@record_notification("线下派单处理情况")
async def reparse_offline_dispatch(db: AsyncSession, directory: Optional[str] = None) -> dict:
    """重新解析线下派单处理情况文件，提取横山汇总+明细数据，更新数据库。"""
    if directory is None:
        directory = settings.watch_dir

    result = await asyncio.to_thread(_parse_offline_dispatch_files, directory)

    from app.core.models import OfflineDispatchSummary, OfflineDispatchDetail
    from sqlalchemy import delete as _delete, text as _text

    # ── 迁移：如明细表缺少新列，删除并重建 ──
    try:
        await db.execute(_text("SELECT timeout_limit FROM offline_dispatch_detail LIMIT 1"))
    except Exception:
        logger.info("线下派单: 明细表缺少新列，正在重建...")
        from app.core.database import engine as _async_engine
        async with _async_engine.begin() as conn:
            await conn.run_sync(lambda _: OfflineDispatchDetail.__table__.drop(_, checkfirst=True))
            await conn.run_sync(lambda _: OfflineDispatchDetail.__table__.create(_))
        await db.rollback()  # 重置事务状态

    # 删除旧数据
    await db.execute(_delete(OfflineDispatchSummary))
    await db.execute(_delete(OfflineDispatchDetail))

    summary_count = 0
    if result["summary"]:
        s = result["summary"]
        ods = OfflineDispatchSummary(
            report_date=result["report_date"] or "",
            district=s["district"],
            latest_filename=result.get("filename") or "",
            monthly_dispatch=s["monthly_dispatch"],
            overdue_backlog=s["overdue_backlog"],
            not_overdue_backlog=s["not_overdue_backlog"],
            total_in_transit=s["total_in_transit"],
            warn_4h_overdue=s["warn_4h_overdue"],
        )
        db.add(ods)
        summary_count = 1

    # 保存明细
    detail_count = 0
    for d in result.get("details", []):
        odd = OfflineDispatchDetail(
            report_file_id=None,
            district=d["district"],
            timeout_limit=d["timeout_limit"],
            broadband_account=d["broadband_account"],
            is_vip_customer=d["is_vip_customer"],
            customer_contact=d["customer_contact"],
            construction_address=d["construction_address"],
            handler_name=d["handler_name"],
            category=d["category"],
        )
        db.add(odd)
        detail_count += 1

    # 获取或创建"线下派单处理情况"报表类型
    stmt = select(ReportType).where(ReportType.name == "线下派单处理情况")
    r = await db.execute(stmt)
    report_type = r.scalar_one_or_none()
    if not report_type:
        report_type = ReportType(name="线下派单处理情况", category="投诉")
        db.add(report_type)
        await db.flush()

    # 删除旧的 report_files/report_records
    old_files_stmt = select(ReportFile.id).where(ReportFile.report_type_id == report_type.id)
    r2 = await db.execute(old_files_stmt)
    old_file_ids = [row[0] for row in r2.all()]
    if old_file_ids:
        await db.execute(_delete(ReportRecord).where(ReportRecord.report_file_id.in_(old_file_ids)))
        await db.execute(_delete(ReportFile).where(ReportFile.report_type_id == report_type.id))

    # 创建 ReportFile
    if result.get("filename"):
        report_file = ReportFile(
            report_type_id=report_type.id,
            filename=result["filename"] or "",
            file_path=os.path.join(directory, result["filename"] or ""),
            parse_status="parsed",
            record_count=detail_count,
        )
        db.add(report_file)

    await db.commit()

    # ── 贾维斯：数据更新后自动触发 AI 分析 ──
    try:
        from app.ai.ai_scheduler import trigger_ai_analysis
        asyncio.create_task(trigger_ai_analysis("offline-dispatch", db))
    except Exception:
        pass

    return {
        "summary_parsed": summary_count,
        "detail_parsed": detail_count,
        "filename": result["filename"],
        "report_date": result["report_date"],
    }


async def get_offline_dispatch_details(
    db: AsyncSession,
    page: int = 1,
    page_size: int = 50,
    category: Optional[str] = None,
    sort_by: Optional[str] = None,
    order: str = "asc",
) -> dict:
    """分页获取线下派单处理情况横山明细数据（支持按分类筛选和排序）
    - sort_by: timeout_limit 时按超时时限排序；其他按 id 排序
    - order: asc=升序，desc=降序
    """
    from app.core.models import OfflineDispatchDetail
    from sqlalchemy import func as _func

    stmt = select(OfflineDispatchDetail)
    if category:
        stmt = stmt.where(OfflineDispatchDetail.category == category)

    count_stmt = select(_func.count(OfflineDispatchDetail.id))
    if category:
        count_stmt = count_stmt.where(OfflineDispatchDetail.category == category)
    r = await db.execute(count_stmt)
    total = r.scalar() or 0

    offset = (page - 1) * page_size

    # 动态排序
    if sort_by and hasattr(OfflineDispatchDetail, sort_by):
        col = getattr(OfflineDispatchDetail, sort_by)
        if order == "desc":
            stmt = stmt.order_by(col.desc())
        else:
            stmt = stmt.order_by(col.asc())
    else:
        stmt = stmt.order_by(OfflineDispatchDetail.id.asc())

    data_stmt = stmt.offset(offset).limit(page_size)
    r2 = await db.execute(data_stmt)
    rows = r2.scalars().all()

    records = []
    for row in rows:
        records.append({
            "id": row.id,
            "district": row.district,
            "timeout_limit": row.timeout_limit,
            "broadband_account": row.broadband_account,
            "is_vip_customer": row.is_vip_customer,
            "customer_contact": row.customer_contact,
            "construction_address": row.construction_address,
            "handler_name": row.handler_name,
            "category": row.category,
        })

    return {
        "records": records,
        "total": total,
        "page": page,
        "page_size": page_size,
    }


async def get_offline_dispatch_summary(db: AsyncSession) -> dict:
    """获取线下派单处理情况横山卡片指标"""
    from app.core.models import OfflineDispatchSummary

    stmt = select(OfflineDispatchSummary).order_by(OfflineDispatchSummary.id.desc()).limit(1)
    r = await db.execute(stmt)
    row = r.scalar_one_or_none()

    if not row:
        return {
            "district": "横山",
            "monthly_dispatch": "",
            "overdue_backlog": "",
            "not_overdue_backlog": "",
            "total_in_transit": "",
            "warn_4h_overdue": "",
            "report_date": "",
            "latest_filename": "",
        }

    return {
        "district": row.district,
        "monthly_dispatch": row.monthly_dispatch,
        "overdue_backlog": row.overdue_backlog,
        "not_overdue_backlog": row.not_overdue_backlog,
        "total_in_transit": row.total_in_transit,
        "warn_4h_overdue": row.warn_4h_overdue,
        "report_date": row.report_date,
        "latest_filename": getattr(row, "latest_filename", "") or "",
    }


def _parse_retry_warning_files(directory: str) -> dict:
    """解析最新的"重投预警工单梳理"文件：
    1. 从"预警通报"sheet 提取横山汇总指标
    2. 从"预警1清单"sheet 提取横山重投预警明细
    3. 从"预警2催修未恢复"sheet 提取横山客户催修明细
    返回 {
        "summary": dict | None,
        "retry_details": [dict],
        "repair_details": [dict],
        "filename": str,
        "report_date": str,
    }
    """
    import re as _re

    pattern = os.path.join(directory, "**", "*重投预警工单梳理*")
    files = sorted(glob.glob(pattern, recursive=True), key=os.path.getmtime, reverse=True)
    # 过滤 Excel 临时文件（~$ 开头）
    files = [f for f in files if not os.path.basename(f).startswith("~$")]
    if not files:
        return {"summary": None, "retry_details": [], "repair_details": [], "filename": None, "report_date": None}

    latest_file = files[0]
    filename = os.path.basename(latest_file)

    try:
        wb = openpyxl.load_workbook(latest_file, read_only=True, data_only=True)

        # ── 1. 解析"预警通报"sheet（汇总指标）──
        summary = None
        report_date = ""

        if "预警通报" in wb.sheetnames:
            ws = wb["预警通报"]
            rows_data = list(ws.iter_rows(values_only=True))

            if rows_data:
                # 提取通报日期：优先从标题/文件名匹配"截至/截止N月N日"
                _date_source = filename + " " + (str(rows_data[0][0]) if rows_data[0] and rows_data[0][0] else "")
                _m = _re.search(r'截至?(\d{1,2})月(\d{1,2})日', _date_source)
                if _m:
                    report_date = f"2026-{_m.group(1).zfill(2)}-{_m.group(2).zfill(2)}"
                else:
                    _m2 = _re.search(r'截至?(\d+)日', _date_source)
                    if _m2:
                        report_date = f"2026-??-{_m2.group(1).zfill(2)}"

                # 横山区数据在 rows_data[8] (Excel row 9)
                hengshan_row = None
                for i, row in enumerate(rows_data):
                    if row and row[0] and "横山" in str(row[0]):
                        hengshan_row = row
                        break

                if hengshan_row:
                    summary = {
                        "district": "横山",
                        "retry_2_times":    str(hengshan_row[5]).strip()  if len(hengshan_row) > 5  and hengshan_row[5]  is not None else "",
                        "retry_3_times":    str(hengshan_row[6]).strip()  if len(hengshan_row) > 6  and hengshan_row[6]  is not None else "",
                        "retry_4plus_times": str(hengshan_row[7]).strip()  if len(hengshan_row) > 7  and hengshan_row[7]  is not None else "",
                        "total_in_transit":  str(hengshan_row[8]).strip()  if len(hengshan_row) > 8  and hengshan_row[8]  is not None else "",
                        "daily_closed":      str(hengshan_row[9]).strip()  if len(hengshan_row) > 9  and hengshan_row[9]  is not None else "",
                        "repair_total":      str(hengshan_row[11]).strip() if len(hengshan_row) > 11 and hengshan_row[11] is not None else "",
                        "repair_in_transit": str(hengshan_row[12]).strip() if len(hengshan_row) > 12 and hengshan_row[12] is not None else "",
                        "repair_closed":     str(hengshan_row[13]).strip() if len(hengshan_row) > 13 and hengshan_row[13] is not None else "",
                    }

        # ── 2. 解析"预警1清单"sheet（重投预警明细）──
        retry_details = []
        if "预警1清单" in wb.sheetnames:
            ws1 = wb["预警1清单"]
            rows1 = list(ws1.iter_rows(values_only=True))
            if rows1 and len(rows1) > 1:
                # 搜索表头行（可能有标题行，表头不一定在第0行）
                header_row_idx1 = None
                for ri, row in enumerate(rows1):
                    row_text = " ".join(str(c) for c in row if c)
                    if "所属区县" in row_text or "重投次数" in row_text:
                        header_row_idx1 = ri
                        break
                if header_row_idx1 is None:
                    header_row_idx1 = 0

                # 构建列名→索引映射（不覆盖已有key，保留第一个出现的列）
                header = rows1[header_row_idx1]
                col_map = {}
                for i, cell in enumerate(header):
                    if cell:
                        key = str(cell).strip()
                        if key and key not in col_map:
                            col_map[key] = i

                # 辅助函数：安全获取列索引（用 is not None 避免列0被 falsy 截断）
                def _col(*names):
                    for name in names:
                        v = col_map.get(name)
                        if v is not None:
                            return v
                    return None

                col_district   = _col("所属区县", "分局")
                col_retry     = _col("重投次数", "重投")
                col_account   = _col("宽带帐号", "宽带账号")
                col_global    = _col("是否全球通用户", "是否全球通", "全球通属性")
                col_contact    = _col("客户联系方式")
                col_address   = _col("小区名称", "施工地址")
                col_days      = _col("历时天数", "超时时限")
                col_handler   = _col("处理人姓名")
                col_complaint = _col("投诉内容")

                logger.info(f"重投预警-预警1清单: 表头行={header_row_idx1}, col_retry={col_retry}, col_global={col_global}, col_district={col_district}")

                def _safe_str(row, idx):
                    if idx is None or idx >= len(row) or row[idx] is None:
                        return ""
                    return str(row[idx]).strip()

                for row in rows1[header_row_idx1 + 1:]:
                    district = _safe_str(row, col_district)
                    if not district or "横山" not in district:
                        continue
                    retry_details.append({
                        "district":         district,
                        "retry_count":      _safe_str(row, col_retry),
                        "broadband_account": _safe_str(row, col_account),
                        "is_global_user":   _safe_str(row, col_global),
                        "customer_contact":  _safe_str(row, col_contact),
                        "construction_address": _safe_str(row, col_address),
                        "days_elapsed":     _safe_str(row, col_days),
                        "handler_name":      _safe_str(row, col_handler),
                        "complaint_content": _safe_str(row, col_complaint),
                    })

        # ── 3. 解析"预警2催修未恢复"sheet（客户催修明细）──
        repair_details = []
        if "预警2催修未恢复" in wb.sheetnames:
            ws2 = wb["预警2催修未恢复"]
            rows2 = list(ws2.iter_rows(values_only=True))
            if rows2 and len(rows2) > 1:
                # 搜索表头行（可能有标题行，表头不一定在第0行）
                header_row_idx2 = None
                for ri, row in enumerate(rows2):
                    for cell in row:
                        if cell and ("县区" in str(cell) or "催修次数" in str(cell)):
                            header_row_idx2 = ri
                            break
                    if header_row_idx2 is not None:
                        break
                if header_row_idx2 is None:
                    header_row_idx2 = 0

                header2 = rows2[header_row_idx2]
                col_map2 = {}
                for i, cell in enumerate(header2):
                    if cell:
                        col_map2[str(cell).strip()] = i

                col_district2  = col_map2.get("县区")
                col_account2  = col_map2.get("账号")
                col_call_num  = col_map2.get("来电号码")
                col_address2  = col_map2.get("地址")
                col_reg_date  = col_map2.get("登记日期")
                col_repair_count = col_map2.get("催修次数")

                logger.info(f"重投预警-预警2催修: 表头行={header_row_idx2}, col_map2={col_map2}, col_repair_count={col_repair_count}")

                def _safe_str2(row, idx):
                    if idx is None or idx >= len(row) or row[idx] is None:
                        return ""
                    return str(row[idx]).strip()

                for row in rows2[header_row_idx2 + 1:]:
                    district = _safe_str2(row, col_district2)
                    if not district or "横山" not in district:
                        continue
                    repair_details.append({
                        "district":      district,
                        "repair_count":  _safe_str2(row, col_repair_count),
                        "account":       _safe_str2(row, col_account2),
                        "call_number":   _safe_str2(row, col_call_num),
                        "address":       _safe_str2(row, col_address2),
                        "register_date": _safe_str2(row, col_reg_date),
                    })

        wb.close()
        logger.info(f"重投预警工单梳理: 解析完成, summary={summary is not None}, retry={len(retry_details)}条, repair={len(repair_details)}条")
        return {
            "summary": summary,
            "retry_details": retry_details,
            "repair_details": repair_details,
            "filename": filename,
            "report_date": report_date,
        }

    except Exception as e:
        logger.error(f"重投预警工单梳理解析失败: {e}", exc_info=True)
        return {"summary": None, "retry_details": [], "repair_details": [], "filename": filename, "report_date": report_date}


@record_notification("重投预警工单梳理")
async def reparse_retry_warning(db: AsyncSession, directory: Optional[str] = None) -> dict:
    """重新解析重投预警工单梳理文件，更新数据库。"""
    if directory is None:
        directory = settings.watch_dir

    result = await asyncio.to_thread(_parse_retry_warning_files, directory)

    from app.core.models import RetryWarningSummary, RetryWarningDetail, CustomerRepairDetail
    from sqlalchemy import delete as _delete

    # 删除旧数据
    await db.execute(_delete(RetryWarningSummary))
    await db.execute(_delete(RetryWarningDetail))
    await db.execute(_delete(CustomerRepairDetail))

    summary_count = 0
    if result["summary"]:
        s = result["summary"]
        rws = RetryWarningSummary(
            report_date=result["report_date"] or "",
            district=s["district"],
            latest_filename=result.get("filename") or "",
            retry_2_times=s["retry_2_times"],
            retry_3_times=s["retry_3_times"],
            retry_4plus_times=s["retry_4plus_times"],
            total_in_transit=s["total_in_transit"],
            daily_closed=s["daily_closed"],
            repair_total=s["repair_total"],
            repair_in_transit=s["repair_in_transit"],
            repair_closed=s["repair_closed"],
        )
        db.add(rws)
        summary_count = 1

    # 保存重投预警明细
    retry_count = 0
    for d in result.get("retry_details", []):
        rwd = RetryWarningDetail(
            district=d["district"],
            retry_count=d["retry_count"],
            broadband_account=d["broadband_account"],
            is_global_user=d["is_global_user"],
            customer_contact=d["customer_contact"],
            construction_address=d["construction_address"],
            days_elapsed=d["days_elapsed"],
            handler_name=d["handler_name"],
            complaint_content=d["complaint_content"],
        )
        db.add(rwd)
        retry_count += 1

    # 保存客户催修明细
    repair_count = 0
    for d in result.get("repair_details", []):
        crd = CustomerRepairDetail(
            district=d["district"],
            repair_count=d["repair_count"],
            account=d["account"],
            call_number=d["call_number"],
            address=d["address"],
            register_date=d["register_date"],
        )
        db.add(crd)
        repair_count += 1

    # 获取或创建"重投预警工单梳理"报表类型
    stmt = select(ReportType).where(ReportType.name == "重投预警工单梳理")
    r = await db.execute(stmt)
    report_type = r.scalar_one_or_none()
    if not report_type:
        report_type = ReportType(name="重投预警工单梳理", category="投诉")
        db.add(report_type)
        await db.flush()

    # 删除旧的 report_files/report_records
    from sqlalchemy import delete as _delete_rf
    old_files_stmt = select(ReportFile.id).where(ReportFile.report_type_id == report_type.id)
    r2 = await db.execute(old_files_stmt)
    old_file_ids = [row[0] for row in r2.all()]
    if old_file_ids:
        await db.execute(_delete_rf(ReportRecord).where(ReportRecord.report_file_id.in_(old_file_ids)))
        await db.execute(_delete_rf(ReportFile).where(ReportFile.report_type_id == report_type.id))

    # 创建 ReportFile
    if result.get("filename"):
        report_file = ReportFile(
            report_type_id=report_type.id,
            filename=result["filename"] or "",
            file_path=os.path.join(directory, result["filename"] or ""),
            parse_status="parsed",
            record_count=retry_count + repair_count,
        )
        db.add(report_file)

    await db.commit()

    return {
        "summary_parsed": summary_count,
        "retry_detail_parsed": retry_count,
        "repair_detail_parsed": repair_count,
        "filename": result["filename"],
        "report_date": result["report_date"],
    }


async def get_retry_warning_summary(db: AsyncSession) -> dict:
    """获取重投预警工单梳理横山卡片指标"""
    from app.core.models import RetryWarningSummary
    from sqlalchemy import select

    stmt = select(RetryWarningSummary).order_by(RetryWarningSummary.id.desc()).limit(1)
    r = await db.execute(stmt)
    row = r.scalar_one_or_none()

    if not row:
        return {
            "district": "横山",
            "retry_2_times": "",
            "retry_3_times": "",
            "retry_4plus_times": "",
            "total_in_transit": "",
            "daily_closed": "",
            "repair_total": "",
            "repair_in_transit": "",
            "repair_closed": "",
            "report_date": "",
            "latest_filename": "",
        }

    return {
        "district": row.district,
        "retry_2_times": row.retry_2_times,
        "retry_3_times": row.retry_3_times,
        "retry_4plus_times": row.retry_4plus_times,
        "total_in_transit": row.total_in_transit,
        "daily_closed": row.daily_closed,
        "repair_total": row.repair_total,
        "repair_in_transit": row.repair_in_transit,
        "repair_closed": row.repair_closed,
        "report_date": row.report_date,
        "latest_filename": getattr(row, "latest_filename", "") or "",
    }

# ── 企宽故障率横山数据专用处理 ──
# 企宽故障率月清单保留字段
ENTERPRISE_BROADBAND_FAULT_FIELDS = [
    "区县",
    "OLT名称",
    "OLT-IP",
    "PON口",
    "账号",
    "告警累计",
    "告警加权时长",
]


def _parse_enterprise_broadband_fault_files(directory: str) -> dict:
    """
    解析最新"企宽故障率"文件：
    1. 从"报表"sheet 提取横山汇总指标
    2. 从"企宽故障率月清单"sheet 提取横山明细
    返回 {"summary": dict, "records": [dict], "filename": str, "report_date": str}
    """
    from datetime import datetime as _dt

    # 找到最新的企宽故障率文件
    matching: list[str] = []
    for root, _dirs, files in os.walk(directory):
        for f in files:
            if f.startswith("~$") or f.startswith("."):
                continue
            if "企宽故障率" in f and f.lower().endswith((".xlsx", ".xls")):
                matching.append(os.path.join(root, f))

    if not matching:
        return {"summary": None, "records": [], "filename": None, "report_date": None}

    # 按修改时间排序，取最新
    matching.sort(key=lambda f: os.path.getmtime(f), reverse=True)
    latest_file = matching[0]
    filename = os.path.basename(latest_file)

    # 尝试从文件名提取日期
    report_date = None
    date_match = re.search(r'(\d{4})[-年](\d{1,2})[-月](\d{1,2})', filename)
    if date_match:
        report_date = f"{date_match.group(1)}-{date_match.group(2).zfill(2)}-{date_match.group(3).zfill(2)}"
    else:
        report_date = _dt.now().strftime("%Y-%m-%d")

    try:
        wb = openpyxl.load_workbook(latest_file, read_only=True, data_only=True)

        summary = None
        records: list[dict] = []

        # ── 1. 解析"报表"sheet ──
        if "报表" in wb.sheetnames:
            ws = wb["报表"]
            rows_data = list(ws.iter_rows(values_only=True))

            logger.info(f"企宽故障率: 报表sheet共有 {len(rows_data)} 行, 文件: {filename}")

            if len(rows_data) >= 2:
                # 关键字 → 数据库字段名 映射
                # 用关键字匹配，表头可能是 "6月企宽故障率测算" 等带月份前缀的形式
                _EBS_KEYWORD_MAP = {
                    "企宽故障率测算": "fault_rate",
                    "采集故障次数": "fault_count",
                    "累计告警时长": "total_alarm_duration",
                    "企宽未恢复告警工单": "unrecoverd_work_orders",
                }

                # 扫描前若干行，建立 keyword -> col_index 映射
                # 关键字可能出现在标题行和数据行中，后扫描到的（真实表头）优先覆盖
                keyword_col_map: dict[str, int] = {}
                hengshan_row_idx = None
                hengshan_row = None

                header_scan_limit = min(15, len(rows_data))
                for idx in range(header_scan_limit):
                    row = rows_data[idx]
                    if not row:
                        continue
                    # 检查每个单元格是否包含某个关键字
                    for col_i, cell in enumerate(row):
                        if cell is None:
                            continue
                        cell_str = str(cell).strip()
                        for kw in _EBS_KEYWORD_MAP:
                            if kw in cell_str:
                                # 后到覆盖：标题行可能也包含关键字，真实表头在后面
                                keyword_col_map[kw] = col_i

                    # 同时检测横山数据行（第一列或任意列包含"横山"）
                    row_text = " ".join(str(c) for c in row if c)
                    if "横山" in row_text:
                        hengshan_row_idx = idx
                        hengshan_row = row

                # 如果前几行没找到横山行，继续扫描剩余行
                if hengshan_row is None:
                    for idx in range(header_scan_limit, len(rows_data)):
                        row = rows_data[idx]
                        if not row:
                            continue
                        row_text = " ".join(str(c) for c in row if c)
                        if "横山" in row_text:
                            hengshan_row_idx = idx
                            hengshan_row = row
                            break

                logger.info(
                    f"企宽故障率: keyword_col_map={keyword_col_map}, "
                    f"横山行idx={hengshan_row_idx}"
                )

                if keyword_col_map and hengshan_row is not None:
                    summary = {
                        "district": "横山",
                        "fault_rate": "",
                        "fault_count": "",
                        "total_alarm_duration": "",
                        "unrecoverd_work_orders": "",
                    }
                    for kw, key in _EBS_KEYWORD_MAP.items():
                        col_idx = keyword_col_map.get(kw)
                        if col_idx is not None and col_idx < len(hengshan_row):
                            val = hengshan_row[col_idx]
                            if val is not None:
                                summary[key] = str(val).strip()
                    logger.info(f"企宽故障率: 提取summary={summary}")
                else:
                    reason = []
                    if not keyword_col_map:
                        reason.append("未找到指标关键字列")
                    if hengshan_row is None:
                        reason.append("未找到横山行")
                    logger.warning(f"企宽故障率: {'; '.join(reason)}, 文件: {filename}")
            else:
                logger.warning(f"企宽故障率: 报表sheet行数不足({len(rows_data)}行), 文件: {filename}")

            logger.info(f"企宽故障率报表sheet最终结果: {filename} -> summary={summary}")

        # ── 2. 解析"企宽故障率月清单"sheet ──
        if "企宽故障率月清单" in wb.sheetnames:
            ws = wb["企宽故障率月清单"]
            row_iter = ws.iter_rows(values_only=True)

            # 读取表头行
            try:
                header_row = next(row_iter)
            except StopIteration:
                header_row = ()

            if header_row:
                # 建立列名 -> 列索引映射
                col_map: dict[str, int] = {}
                target_fields = ENTERPRISE_BROADBAND_FAULT_FIELDS
                for idx, cell in enumerate(header_row):
                    if cell is None:
                        continue
                    cell_str = str(cell).strip()
                    for field in target_fields:
                        if field in cell_str:
                            col_map[field] = idx
                            break

                # 检查必要字段
                missing = [f for f in ["区县", "OLT名称", "账号"] if f not in col_map]
                if not missing:
                    max_col_idx = max(col_map.values())
                    for row in row_iter:
                        if not row or len(row) <= max_col_idx:
                            continue
                        district_val = str(row[col_map["区县"]]).strip() if "区县" in col_map and row[col_map["区县"]] is not None else ""
                        if district_val != "横山":
                            continue

                        def _get_val(field: str) -> str:
                            idx = col_map.get(field)
                            if idx is not None and idx < len(row) and row[idx] is not None:
                                return str(row[idx]).strip()
                            return ""

                        record = {
                            "区县": district_val,
                            "OLT名称": _get_val("OLT名称"),
                            "OLT-IP": _get_val("OLT-IP"),
                            "PON口": _get_val("PON口"),
                            "账号": _get_val("账号"),
                            "告警累计": _get_val("告警累计"),
                            "告警加权时长": _get_val("告警加权时长"),
                        }
                        records.append(record)

                    logger.info(f"企宽故障率月清单: {filename} -> {len(records)} 条横山记录")
                else:
                    logger.warning(f"企宽故障率月清单缺少字段: {missing}")
        else:
            logger.warning(f"企宽故障率文件 {filename} 缺少'企宽故障率月清单'sheet")

        wb.close()
        return {
            "summary": summary,
            "records": records,
            "filename": filename,
            "report_date": report_date,
        }

    except Exception as e:
        logger.error(f"解析企宽故障率文件失败 {filename}: {e}", exc_info=True)
        return {"summary": None, "records": [], "filename": filename, "report_date": report_date}


@record_notification("企宽故障率")
async def reparse_enterprise_broadband_fault(db: AsyncSession, directory: Optional[str] = None) -> dict:
    """
    重新解析企宽故障率文件，提取横山汇总指标和故障明细，更新数据库。
    删除旧数据并写入新数据。
    """
    if directory is None:
        directory = settings.watch_dir

    # 在线程池中解析
    result = await asyncio.to_thread(_parse_enterprise_broadband_fault_files, directory)

    # ── 写入汇总表 ──
    from app.core.models import EnterpriseBroadbandFaultSummary, EnterpriseBroadbandFaultRecord
    from sqlalchemy import delete as _delete

    await db.execute(_delete(EnterpriseBroadbandFaultSummary))
    await db.execute(_delete(EnterpriseBroadbandFaultRecord))

    summary_count = 0
    if result["summary"]:
        s = result["summary"]
        ebfs = EnterpriseBroadbandFaultSummary(
            report_date=result["report_date"] or "",
            district="横山",
            latest_filename=result.get("filename") or "",
            fault_rate=s.get("fault_rate", ""),
            fault_count=s.get("fault_count", ""),
            total_alarm_duration=s.get("total_alarm_duration", ""),
            unrecoverd_work_orders=s.get("unrecoverd_work_orders", ""),
        )
        db.add(ebfs)
        summary_count = 1

    # ── 写入明细表 ──
    # 先获取或创建"企宽故障率"报表类型用于关联 ReportFile
    stmt = select(ReportType).where(ReportType.name == "企宽故障率")
    r = await db.execute(stmt)
    report_type = r.scalar_one_or_none()
    if not report_type:
        report_type = ReportType(name="企宽故障率", category="质差整治")
        db.add(report_type)
        await db.flush()

    # 删除旧的 report_files/report_records
    old_files_stmt = select(ReportFile.id).where(ReportFile.report_type_id == report_type.id)
    r2 = await db.execute(old_files_stmt)
    old_file_ids = [row[0] for row in r2.all()]
    if old_file_ids:
        await db.execute(_delete(ReportRecord).where(ReportRecord.report_file_id.in_(old_file_ids)))
        await db.execute(_delete(ReportFile).where(ReportFile.id.in_(old_file_ids)))

    # 创建 ReportFile
    if result["filename"]:
        report_file = ReportFile(
            report_type_id=report_type.id,
            filename=result["filename"],
            file_path=os.path.join(directory, result["filename"]),
            parse_status="parsed",
            record_count=len(result["records"]),
        )
        db.add(report_file)
        await db.flush()

        for rec in result["records"]:
            ebfr = EnterpriseBroadbandFaultRecord(
                report_file_id=report_file.id,
                district=rec["区县"],
                olt_name=rec.get("OLT名称", ""),
                olt_ip=rec.get("OLT-IP", ""),
                pon_port=rec.get("PON口", ""),
                account=rec.get("账号", ""),
                alarm_total=rec.get("告警累计", ""),
                alarm_weighted_duration=rec.get("告警加权时长", ""),
            )
            db.add(ebfr)

    await db.commit()

    return {
        "summary_parsed": summary_count,
        "detail_count": len(result["records"]),
        "filename": result["filename"],
        "report_date": result["report_date"],
    }


async def get_enterprise_broadband_fault_summary(db: AsyncSession) -> dict:
    """获取企宽故障率横山卡片指标"""
    from app.core.models import EnterpriseBroadbandFaultSummary
    stmt = select(EnterpriseBroadbandFaultSummary).order_by(EnterpriseBroadbandFaultSummary.id.desc()).limit(1)
    r = await db.execute(stmt)
    row = r.scalar_one_or_none()
    if not row:
        return {
            "district": "横山",
            "fault_rate": "",
            "fault_count": "",
            "total_alarm_duration": "累计告警时长",
            "unrecoverd_work_orders": "企宽未恢复告警工单",
            "report_date": "",
            "latest_filename": "",
        }
    return {
        "district": row.district,
        "fault_rate": row.fault_rate,
        "fault_count": row.fault_count,
        "total_alarm_duration": row.total_alarm_duration,
        "unrecoverd_work_orders": row.unrecoverd_work_orders,
        "report_date": row.report_date,
        "latest_filename": getattr(row, "latest_filename", "") or "",
    }


async def get_enterprise_broadband_fault_details(
    db: AsyncSession,
    page: int = 1,
    page_size: int = 50,
    sort_field: Optional[str] = None,
    sort_order: Optional[str] = None,
    district_filter: Optional[str] = None,
) -> dict:
    """分页获取企宽故障率横山故障明细（支持排序和筛选）"""
    from app.core.models import EnterpriseBroadbandFaultRecord
    from sqlalchemy import func as _func, select as _select

    # 构建查询
    stmt = select(EnterpriseBroadbandFaultRecord)
    if district_filter:
        stmt = stmt.where(EnterpriseBroadbandFaultRecord.district == district_filter)

    # 排序：字段是 String 类型，需要 CAST 为 FLOAT 才能正确数值排序
    if sort_field and sort_field in ("alarm_total", "alarm_weighted_duration"):
        from sqlalchemy import cast as _cast, Float as _Float
        order_col = _cast(getattr(EnterpriseBroadbandFaultRecord, sort_field), _Float)
        if sort_order == "desc":
            stmt = stmt.order_by(order_col.desc())
        else:
            stmt = stmt.order_by(order_col.asc())
    else:
        stmt = stmt.order_by(EnterpriseBroadbandFaultRecord.id.desc())

    # 统计总数
    count_stmt = _select(func.count(EnterpriseBroadbandFaultRecord.id))
    if district_filter:
        count_stmt = count_stmt.where(EnterpriseBroadbandFaultRecord.district == district_filter)
    r = await db.execute(count_stmt)
    total = r.scalar() or 0

    # 分页
    offset = (page - 1) * page_size
    stmt = stmt.offset(offset).limit(page_size)
    r2 = await db.execute(stmt)
    rows = r2.scalars().all()

    records = []
    for row in rows:
        records.append({
            "id": row.id,
            "district": row.district,
            "olt_name": row.olt_name,
            "olt_ip": row.olt_ip,
            "pon_port": row.pon_port,
            "account": row.account,
            "alarm_total": row.alarm_total,
            "alarm_weighted_duration": row.alarm_weighted_duration,
        })

    return {
        "records": records,
        "total": total,
        "page": page,
        "page_size": page_size,
    }


# ── 质差小区弱光工单处理完成率横山数据专用处理 ──

# 当月工单清单保留字段（中文列名）
POOR_QUALITY_WORK_ORDER_DETAIL_FIELDS = [
    "区县",
    "工单号",
    "派单时间",
    "限定完成时间",
    "维护人员",
    "备注",
    "OLT IP",
    "OLT接口",
]


def _parse_poor_quality_work_order_file(filepath: str) -> dict:
    """
    解析单个"质差小区弱光工单处理完成率"文件。
    返回 {
        "summary": {"work_order_count", "completed_count", "completion_rate", "community_count"},
        "records": [dict],   # 当月工单清单 横山+未及时完工
        "filename": str,
        "report_date": str,
    }
    """
    import openpyxl
    import re

    filename = os.path.basename(filepath)
    result = {
        "summary": None,
        "records": [],
        "filename": filename,
        "report_date": None,
    }

    try:
        # 从文件名提取日期
        date_match = re.search(r"(\d{4})[-_]?(\d{1,2})[-_]?(\d{1,2})", filename)
        if date_match:
            try:
                result["report_date"] = f"{date_match.group(1)}-{date_match.group(2).zfill(2)}-{date_match.group(3).zfill(2)}"
            except Exception:
                pass

        wb = openpyxl.load_workbook(filepath, data_only=True)

        summary = {
            "work_order_count": "",
            "completed_count": "",
            "completion_rate": "",
            "community_count": "",
        }

        # ── 1. 解析"区县弱光工单处理完成率" sheet ──
        target_sheet = None
        for sname in wb.sheetnames:
            if "区县弱光工单处理完成率" in sname or "弱光工单处理完成率" in sname:
                target_sheet = sname
                break
        if not target_sheet and wb.sheetnames:
            target_sheet = wb.sheetnames[0]

        if target_sheet and target_sheet in wb.sheetnames:
            ws = wb[target_sheet]

            # 查找真正的表头行（含"区县"列的行，跳过标题行）
            header_row_idx = 1
            header = []
            for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 10), values_only=True), start=1):
                row_vals = [str(c or "").strip() for c in row]
                if any("区县" in v for v in row_vals):
                    header = row_vals
                    header_row_idx = r_idx
                    break
            if not header:
                header = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]

            col_idx = {}
            for i, h in enumerate(header):
                if "区县" in h:
                    col_idx["district"] = i
                elif "工单数" in h and "累计" not in h:
                    col_idx["work_order_count"] = i
                elif "累计回单" in h or "回单数" in h:
                    col_idx["completed_count"] = i
                elif "完成率" in h or "完成%" in h or "回单%" in h:
                    col_idx["completion_rate"] = i
                elif "涉及小区" in h:
                    col_idx["community_count"] = i

            for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
                district_val = str(row[col_idx.get("district", 0)] or "").strip() if col_idx.get("district") is not None and col_idx["district"] < len(row) else ""
                if "横山" in district_val:
                    if col_idx.get("work_order_count") is not None and col_idx["work_order_count"] < len(row):
                        summary["work_order_count"] = str(row[col_idx["work_order_count"]] or "").strip()
                    if col_idx.get("completed_count") is not None and col_idx["completed_count"] < len(row):
                        summary["completed_count"] = str(row[col_idx["completed_count"]] or "").strip()
                    if col_idx.get("completion_rate") is not None and col_idx["completion_rate"] < len(row):
                        summary["completion_rate"] = str(row[col_idx["completion_rate"]] or "").strip()
                    if col_idx.get("community_count") is not None and col_idx["community_count"] < len(row):
                        summary["community_count"] = str(row[col_idx["community_count"]] or "").strip()
                    break

        # ── 2. 备选：从"小区维度" sheet 提取涉及小区数（仅当汇总表未提供时） ──
        if not summary["community_count"]:
            community_count = ""
            for sname in wb.sheetnames:
                if "小区维度" in sname:
                    ws2 = wb[sname]
                    # 计算数据行数 = 涉及小区数
                    data_rows = 0
                    for i, row in enumerate(ws2.iter_rows(values_only=True)):
                        if i == 0:
                            continue
                        if any(c is not None and str(c).strip() for c in row):
                            data_rows += 1
                    community_count = str(data_rows)
                    break
            summary["community_count"] = community_count
        result["summary"] = summary

        # ── 3. 解析"当月工单清单" sheet，筛选横山 + 是否及时完工=否 ──
        # 注意："未及时工单清单" 也包含 "工单清单"，必须先精确匹配 "当月工单清单"
        detail_sheet = None
        for sname in wb.sheetnames:
            if "当月工单清单" in sname:
                detail_sheet = sname
                break
        if not detail_sheet:
            for sname in wb.sheetnames:
                if "工单清单" in sname:
                    detail_sheet = sname
                    break

        if detail_sheet and detail_sheet in wb.sheetnames:
            ws3 = wb[detail_sheet]
            header3 = [str(c.value or "").strip() for c in next(ws3.iter_rows(min_row=1, max_row=1))]

            idx = {}
            for i, h in enumerate(header3):
                if h in POOR_QUALITY_WORK_ORDER_DETAIL_FIELDS:
                    idx[h] = i
                elif "区县" in h:
                    idx["区县"] = i
                elif "工单号" in h:
                    idx["工单号"] = i
                elif "派单时间" in h or "派单" in h:
                    idx["派单时间"] = i
                elif "限定完成" in h or "完成时间" in h:
                    idx["限定完成时间"] = i
                elif "维护人员" in h or "维护人" in h:
                    idx["维护人员"] = i
                elif "备注" in h:
                    idx["备注"] = i
                elif "OLT IP" in h or "OLT_IP" in h or "oltip" in h.lower():
                    idx["OLT IP"] = i
                elif "OLT接口" in h or "OLT端口" in h or "olt接口" in h:
                    idx["OLT接口"] = i
                elif "是否及时" in h or "及时完工" in h:
                    idx["是否及时完工"] = i

            for row in ws3.iter_rows(min_row=2, values_only=True):
                # 筛选：区县包含"横山"
                district_val = str(row[idx.get("区县", 0)] or "").strip() if idx.get("区县") is not None and idx["区县"] < len(row) else ""
                if "横山" not in district_val:
                    continue
                # 筛选：是否及时完工 != "是"
                timely = str(row[idx.get("是否及时完工", -1)] or "").strip() if idx.get("是否及时完工") is not None and idx["是否及时完工"] < len(row) else ""
                if timely == "是":
                    continue

                rec = {}
                for field in POOR_QUALITY_WORK_ORDER_DETAIL_FIELDS:
                    col_i = idx.get(field)
                    if col_i is not None and col_i < len(row):
                        rec[field] = str(row[col_i] or "").strip()
                    else:
                        rec[field] = ""
                result["records"].append(rec)

            logger.info(f"质差小区弱光工单: {filename} -> 明细 {len(result['records'])} 条")

        wb.close()

    except Exception as e:
        logger.error(f"解析质差小区弱光工单文件失败 {filename}: {e}", exc_info=True)
        return {"summary": None, "records": [], "filename": filename, "report_date": None}

    return result


@record_notification("质差小区弱光工单处理完成率")
async def reparse_poor_quality_work_order(db: AsyncSession, directory: Optional[str] = None) -> dict:
    """
    重新解析质差小区弱光工单处理完成率文件，
    提取横山汇总指标和未完成工单明细，更新数据库。
    """
    if directory is None:
        directory = settings.watch_dir

    # 找到最新的匹配文件
    matching: list[str] = []
    for root, _dirs, files in os.walk(directory):
        for f in files:
            if f.startswith("~$") or f.startswith("."):
                continue
            if "质差小区弱光工单处理完成率" in f and f.lower().endswith((".xlsx", ".xls")):
                matching.append(os.path.join(root, f))

    if not matching:
        logger.warning("质差小区弱光工单: 未找到匹配文件")
        return {"summary_parsed": 0, "record_count": 0, "filename": None}

    matching.sort(key=os.path.getmtime)
    filepath = matching[-1]
    filename = os.path.basename(filepath)

    # 在线程池中解析
    result = await asyncio.to_thread(_parse_poor_quality_work_order_file, filepath)

    # ── 写入汇总表 ──
    from app.core.models import PoorQualityWorkOrderSummary, PoorQualityWorkOrderRecord
    from sqlalchemy import delete as _delete

    await db.execute(_delete(PoorQualityWorkOrderSummary))
    await db.flush()

    s = result.get("summary")
    if s:
        pwqs = PoorQualityWorkOrderSummary(
            report_date=result.get("report_date") or "",
            district="横山",
            latest_filename=filename,
            work_order_count=s.get("work_order_count", ""),
            completed_count=s.get("completed_count", ""),
            completion_rate=s.get("completion_rate", ""),
            community_count=s.get("community_count", ""),
        )
        db.add(pwqs)
        await db.flush()

    # ── 写入明细表 ──
    # 先获取或创建报表类型
    stmt = select(ReportType).where(ReportType.name == "质差小区弱光工单处理完成率")
    r = await db.execute(stmt)
    report_type = r.scalar_one_or_none()
    if not report_type:
        report_type = ReportType(name="质差小区弱光工单处理完成率", category="质差整治")
        db.add(report_type)
        await db.flush()

    # 删除旧的 report_files 和 records（注意外键顺序）
    old_files_stmt = select(ReportFile.id).where(ReportFile.report_type_id == report_type.id)
    r2 = await db.execute(old_files_stmt)
    old_file_ids = [row[0] for row in r2.all()]
    if old_file_ids:
        # 先删除 PoorQualityWorkOrderRecord（引用 ReportFile 的子表）
        await db.execute(_delete(PoorQualityWorkOrderRecord).where(
            PoorQualityWorkOrderRecord.report_file_id.in_(old_file_ids)))
        # 再删除通用的 ReportRecord
        await db.execute(_delete(ReportRecord).where(ReportRecord.report_file_id.in_(old_file_ids)))
        # 最后删除 ReportFile
        await db.execute(_delete(ReportFile).where(ReportFile.id.in_(old_file_ids)))
        await db.flush()

    # 创建新的 ReportFile
    report_file = ReportFile(
        report_type_id=report_type.id,
        filename=filename,
        file_path=filepath,
        parse_status="parsed",
        record_count=len(result.get("records", [])),
    )
    db.add(report_file)
    await db.flush()

    # 写入明细记录
    records = result.get("records", [])
    for rec in records:
        pwqr = PoorQualityWorkOrderRecord(
            report_file_id=report_file.id,
            district=rec.get("区县", ""),
            work_order_no=rec.get("工单号", ""),
            dispatch_time=rec.get("派单时间", ""),
            deadline=rec.get("限定完成时间", ""),
            maintenance_person=rec.get("维护人员", ""),
            notes=rec.get("备注", ""),
            olt_ip=rec.get("OLT IP", ""),
            olt_port=rec.get("OLT接口", ""),
        )
        db.add(pwqr)

    await db.commit()
    logger.info(f"质差小区弱光工单: 已写入数据库 summary={'是' if s else '否'}, records={len(records)}, 文件={filename}")

    return {
        "summary_parsed": 1 if s else 0,
        "record_count": len(records),
        "filename": filename,
        "report_date": result.get("report_date"),
    }


async def get_poor_quality_work_order_summary(db: AsyncSession) -> dict:
    """获取质差小区弱光工单横山卡片指标"""
    from app.core.models import PoorQualityWorkOrderSummary
    stmt = select(PoorQualityWorkOrderSummary).order_by(PoorQualityWorkOrderSummary.id.desc()).limit(1)
    r = await db.execute(stmt)
    row = r.scalar_one_or_none()
    if not row:
        return {
            "work_order_count": "",
            "completed_count": "",
            "completion_rate": "",
            "community_count": "",
            "report_date": None,
            "latest_filename": None,
        }
    return {
        "work_order_count": row.work_order_count,
        "completed_count": row.completed_count,
        "completion_rate": row.completion_rate,
        "community_count": row.community_count,
        "report_date": row.report_date,
        "latest_filename": row.latest_filename,
    }


async def get_poor_quality_work_order_details(
    db: AsyncSession,
    page: int = 1,
    page_size: int = 50,
) -> dict:
    """分页获取质差小区弱光工单横山未完成明细"""
    from app.core.models import PoorQualityWorkOrderRecord
    from sqlalchemy import func

    # 统计总数
    count_stmt = select(func.count(PoorQualityWorkOrderRecord.id))
    r = await db.execute(count_stmt)
    total = r.scalar() or 0

    # 分页查询
    offset = (page - 1) * page_size
    data_stmt = (
        select(PoorQualityWorkOrderRecord)
        .order_by(PoorQualityWorkOrderRecord.id.asc())
        .offset(offset)
        .limit(page_size)
    )
    r2 = await db.execute(data_stmt)
    rows = r2.scalars().all()

    records = []
    for row in rows:
        records.append({
            "id": row.id,
            "district": row.district,
            "work_order_no": row.work_order_no,
            "dispatch_time": row.dispatch_time,
            "deadline": row.deadline,
            "maintenance_person": row.maintenance_person,
            "notes": row.notes,
            "olt_ip": row.olt_ip,
            "olt_port": row.olt_port,
        })

    return {
        "records": records,
        "total": total,
        "page": page,
        "page_size": page_size,
    }


# ── 企宽弱光通报横山数据专用处理 ──
# 企宽清单保留字段
ENTERPRISE_BROADBAND_LOW_LIGHT_FIELDS = [
    "区县",
    "日期",
    "OLT名称",
    "OLT-IP",
    "PON口",
    "ONU-ID",
    "收光dbm",
    "小区",
    "账号-带宽",
]


def _parse_enterprise_broadband_low_light_files(directory: str) -> dict:
    """
    解析最新"企宽弱光通报"文件：
    1. 从"企宽通报"sheet 提取横山汇总指标 + 全部县区数据（用于排名）
    2. 从"企宽清单"sheet 提取横山未恢复明细
    返回 {"summary": dict, "records": [dict], "filename": str, "report_date": str}
    """
    from datetime import datetime as _dt

    # 找到最新的企宽弱光通报文件
    matching: list[str] = []
    for root, _dirs, files in os.walk(directory):
        for f in files:
            if f.startswith("~$") or f.startswith("."):
                continue
            if "企宽弱光通报" in f and f.lower().endswith((".xlsx", ".xls")):
                matching.append(os.path.join(root, f))

    if not matching:
        return {"summary": None, "records": [], "filename": None, "report_date": None}

    # 按修改时间排序，取最新
    matching.sort(key=lambda f: os.path.getmtime(f), reverse=True)
    latest_file = matching[0]
    filename = os.path.basename(latest_file)

    # 尝试从文件名提取日期
    report_date = None
    date_match = re.search(r'(\d{4})[-年](\d{1,2})[-月](\d{1,2})', filename)
    if date_match:
        report_date = f"{date_match.group(1)}-{date_match.group(2).zfill(2)}-{date_match.group(3).zfill(2)}"
    else:
        report_date = _dt.now().strftime("%Y-%m-%d")

    try:
        wb = openpyxl.load_workbook(latest_file, read_only=True, data_only=True)

        summary = None
        records: list[dict] = []

        # ── 1. 解析"企宽通报"sheet ──
        if "企宽通报" in wb.sheetnames:
            ws = wb["企宽通报"]
            rows_data = list(ws.iter_rows(values_only=True))

            logger.info(f"企宽弱光: 企宽通报sheet共有 {len(rows_data)} 行, 文件: {filename}")

            if len(rows_data) >= 3:
                # 扫描前10行，找到表头行（含"企宽总量"关键字）
                header_row = None
                header_scan_limit = min(10, len(rows_data))
                for idx in range(header_scan_limit):
                    row = rows_data[idx]
                    if not row:
                        continue
                    row_text = " ".join(str(c) for c in row if c)
                    if "企宽总量" in row_text:
                        header_row = row
                        break

                if header_row:
                    # 建立列索引：区县, 企宽总量, 月完成量, 月完成率
                    dist_idx = None
                    total_idx = None
                    completed_idx = None
                    rate_idx = None
                    for ci, cell in enumerate(header_row):
                        if cell is None:
                            continue
                        cs = str(cell).strip()
                        if ("区县" in cs or "县区" in cs) and dist_idx is None:
                            dist_idx = ci
                        elif "企宽总量" in cs and total_idx is None:
                            total_idx = ci
                        elif "月完成量" in cs and completed_idx is None:
                            completed_idx = ci
                        elif "月完成率" in cs and rate_idx is None:
                            rate_idx = ci

                    logger.info(
                        f"企宽弱光: col_indexes -> dist={dist_idx}, total={total_idx}, "
                        f"completed={completed_idx}, rate={rate_idx}"
                    )

                    # 如果没找到明确的列，用默认位置
                    if dist_idx is None:
                        dist_idx = 0
                    if total_idx is None:
                        total_idx = 1
                    if completed_idx is None:
                        completed_idx = 2
                    if rate_idx is None:
                        rate_idx = 3

                    # 收集所有县区数据用于排名
                    county_data: list[dict] = []
                    hengshan_summary = None

                    for row in rows_data[1:]:
                        if not row:
                            continue
                        max_idx = max(dist_idx, total_idx, completed_idx, rate_idx)
                        if len(row) <= max_idx:
                            continue

                        name_val = str(row[dist_idx]).strip() if dist_idx < len(row) and row[dist_idx] else ""
                        if not name_val or name_val == "None":
                            continue

                        # 跳过标题行
                        if "弱光通报" in name_val:
                            continue
                        # 跳过表头行（含"区县"/"县区"等列名）
                        if "区县" in name_val or "县区" in name_val or "企宽总量" in name_val or "月完成" in name_val:
                            continue

                        total_val = row[total_idx] if total_idx < len(row) else None
                        completed_val = row[completed_idx] if completed_idx < len(row) else None
                        rate_val = row[rate_idx] if rate_idx < len(row) else None

                        if total_val is None and rate_val is None:
                            continue

                        county_entry = {
                            "name": name_val,
                            "total_count": str(total_val).strip() if total_val is not None else "",
                            "monthly_completed": str(completed_val).strip() if completed_val is not None else "",
                            "monthly_completion_rate": str(rate_val).strip() if rate_val is not None else "",
                        }
                        county_data.append(county_entry)

                        if "横山" in name_val:
                            hengshan_summary = county_entry

                    # 计算县区排名（剔除"全市"，按完成率降序排列）
                    rankable = [c for c in county_data if "全市" not in c["name"]]
                    try:
                        rankable.sort(key=lambda x: float(x["monthly_completion_rate"]) if x["monthly_completion_rate"] else 0, reverse=True)
                    except Exception:
                        pass

                    hengshan_rank = None
                    total_rankable = len(rankable)
                    for ri, c in enumerate(rankable):
                        if "横山" in c["name"]:
                            hengshan_rank = ri + 1
                            break

                    if hengshan_summary:
                        summary = {
                            "district": "横山",
                            "total_count": hengshan_summary["total_count"],
                            "monthly_completed": hengshan_summary["monthly_completed"],
                            "monthly_completion_rate": hengshan_summary["monthly_completion_rate"],
                            "county_rank": f"第{hengshan_rank}名/{total_rankable}" if hengshan_rank else "",
                        }
                    else:
                        logger.warning(f"企宽弱光: 未找到横山行, 文件: {filename}")

                    logger.info(f"企宽弱光: 县区数={len(county_data)}, 可排名={total_rankable}, 横山排名={hengshan_rank}")
                else:
                    logger.warning(f"企宽弱光: 未找到含'企宽总量'的表头行, 文件: {filename}")
            else:
                logger.warning(f"企宽弱光: 企宽通报sheet行数不足({len(rows_data)}行), 文件: {filename}")

            logger.info(f"企宽弱光通报sheet最终结果: {filename} -> summary={summary}")

        # ── 2. 解析"企宽清单"sheet ──
        if "企宽清单" in wb.sheetnames:
            ws = wb["企宽清单"]
            row_iter = ws.iter_rows(values_only=True)

            # 读取表头行
            try:
                header_row = next(row_iter)
            except StopIteration:
                header_row = ()

            if header_row:
                # 建立列名 -> 列索引映射
                col_map: dict[str, int] = {}
                target_fields = ENTERPRISE_BROADBAND_LOW_LIGHT_FIELDS + ["是否恢复"]
                for idx, cell in enumerate(header_row):
                    if cell is None:
                        continue
                    cell_str = str(cell).strip()
                    for field in target_fields:
                        if field in cell_str:
                            col_map[field] = idx
                            break

                # 检查必要字段
                required = ["区县", "日期", "OLT名称", "OLT-IP", "ONU-ID"]
                missing_fields = [f for f in required if f not in col_map]
                if not missing_fields:
                    max_col_idx = max(col_map.values())
                    for row in row_iter:
                        if not row or len(row) <= max_col_idx:
                            continue
                        # 筛选区县=横山
                        district_val = str(row[col_map["区县"]]).strip() if "区县" in col_map and row[col_map["区县"]] is not None else ""
                        if district_val != "横山":
                            continue
                        # 剔除是否恢复=是
                        is_recovered = ""
                        if "是否恢复" in col_map and col_map["是否恢复"] < len(row) and row[col_map["是否恢复"]] is not None:
                            is_recovered = str(row[col_map["是否恢复"]]).strip()
                        if is_recovered == "是":
                            continue

                        def _get_val(field: str) -> str:
                            idx = col_map.get(field)
                            if idx is not None and idx < len(row) and row[idx] is not None:
                                return str(row[idx]).strip()
                            return ""

                        record = {
                            "区县": district_val,
                            "日期": _get_val("日期"),
                            "OLT名称": _get_val("OLT名称"),
                            "OLT-IP": _get_val("OLT-IP"),
                            "PON口": _get_val("PON口"),
                            "ONU-ID": _get_val("ONU-ID"),
                            "收光dbm": _get_val("收光dbm"),
                            "小区": _get_val("小区"),
                            "账号-带宽": _get_val("账号-带宽"),
                        }
                        records.append(record)

                    logger.info(f"企宽弱光清单: {filename} -> {len(records)} 条横山未恢复记录")
                else:
                    logger.warning(f"企宽弱光清单缺少字段: {missing_fields}")
        else:
            logger.warning(f"企宽弱光文件 {filename} 缺少'企宽清单'sheet")

        wb.close()
        return {
            "summary": summary,
            "records": records,
            "filename": filename,
            "report_date": report_date,
        }

    except Exception as e:
        logger.error(f"解析企宽弱光通报文件失败 {filename}: {e}", exc_info=True)
        return {"summary": None, "records": [], "filename": filename, "report_date": report_date}


@record_notification("企宽弱光通报")
async def reparse_enterprise_broadband_low_light(db: AsyncSession, directory: Optional[str] = None) -> dict:
    """
    重新解析企宽弱光通报文件，提取横山汇总指标和未恢复明细，更新数据库。
    删除旧数据并写入新数据。
    """
    if directory is None:
        directory = settings.watch_dir

    # 在线程池中解析
    result = await asyncio.to_thread(_parse_enterprise_broadband_low_light_files, directory)

    # ── 写入汇总表 ──
    from app.core.models import EnterpriseBroadbandLowLightSummary, EnterpriseBroadbandLowLightRecord
    from sqlalchemy import delete as _delete

    await db.execute(_delete(EnterpriseBroadbandLowLightSummary))
    await db.execute(_delete(EnterpriseBroadbandLowLightRecord))

    summary_count = 0
    if result["summary"]:
        s = result["summary"]
        eblls = EnterpriseBroadbandLowLightSummary(
            report_date=result["report_date"] or "",
            district="横山",
            latest_filename=result.get("filename") or "",
            total_count=s.get("total_count", ""),
            monthly_completed=s.get("monthly_completed", ""),
            monthly_completion_rate=s.get("monthly_completion_rate", ""),
            county_rank=s.get("county_rank", ""),
        )
        db.add(eblls)
        summary_count = 1

    # ── 写入明细表 ──
    # 先获取或创建"企宽弱光通报"报表类型用于关联 ReportFile
    stmt = select(ReportType).where(ReportType.name == "企宽弱光通报")
    r = await db.execute(stmt)
    report_type = r.scalar_one_or_none()
    if not report_type:
        report_type = ReportType(name="企宽弱光通报", category="质差整治")
        db.add(report_type)
        await db.flush()

    # 删除旧的 report_files/report_records
    old_files_stmt = select(ReportFile.id).where(ReportFile.report_type_id == report_type.id)
    r2 = await db.execute(old_files_stmt)
    old_file_ids = [row[0] for row in r2.all()]
    if old_file_ids:
        await db.execute(_delete(ReportRecord).where(ReportRecord.report_file_id.in_(old_file_ids)))
        await db.execute(_delete(ReportFile).where(ReportFile.id.in_(old_file_ids)))

    # 创建 ReportFile
    if result["filename"]:
        report_file = ReportFile(
            report_type_id=report_type.id,
            filename=result["filename"],
            file_path=os.path.join(directory, result["filename"]),
            parse_status="parsed",
            record_count=len(result["records"]),
        )
        db.add(report_file)
        await db.flush()

        for rec in result["records"]:
            ebllr = EnterpriseBroadbandLowLightRecord(
                report_file_id=report_file.id,
                district=rec["区县"],
                date=rec.get("日期", ""),
                olt_name=rec.get("OLT名称", ""),
                olt_ip=rec.get("OLT-IP", ""),
                pon_port=rec.get("PON口", ""),
                onu_id=rec.get("ONU-ID", ""),
                rx_power_dbm=rec.get("收光dbm", ""),
                community=rec.get("小区", ""),
                account_bandwidth=rec.get("账号-带宽", ""),
            )
            db.add(ebllr)

    await db.commit()

    return {
        "summary_parsed": summary_count,
        "detail_count": len(result["records"]),
        "filename": result["filename"],
        "report_date": result["report_date"],
    }


async def get_enterprise_broadband_low_light_summary(db: AsyncSession) -> dict:
    """获取企宽弱光通报横山卡片指标"""
    from app.core.models import EnterpriseBroadbandLowLightSummary
    stmt = select(EnterpriseBroadbandLowLightSummary).order_by(EnterpriseBroadbandLowLightSummary.id.desc()).limit(1)
    r = await db.execute(stmt)
    row = r.scalar_one_or_none()
    if not row:
        return {
            "district": "横山",
            "total_count": "",
            "monthly_completed": "",
            "monthly_completion_rate": "",
            "county_rank": "",
            "report_date": "",
            "latest_filename": "",
        }
    return {
        "district": row.district,
        "total_count": row.total_count,
        "monthly_completed": row.monthly_completed,
        "monthly_completion_rate": row.monthly_completion_rate,
        "county_rank": row.county_rank,
        "report_date": row.report_date,
        "latest_filename": getattr(row, "latest_filename", "") or "",
    }


async def get_enterprise_broadband_low_light_details(
    db: AsyncSession,
    page: int = 1,
    page_size: int = 50,
) -> dict:
    """分页获取企宽弱光通报横山未恢复明细"""
    from app.core.models import EnterpriseBroadbandLowLightRecord
    from sqlalchemy import func

    count_stmt = select(func.count()).select_from(EnterpriseBroadbandLowLightRecord)
    r1 = await db.execute(count_stmt)
    total = r1.scalar() or 0

    offset = (page - 1) * page_size
    data_stmt = (
        select(EnterpriseBroadbandLowLightRecord)
        .order_by(EnterpriseBroadbandLowLightRecord.id.asc())
        .offset(offset)
        .limit(page_size)
    )
    r2 = await db.execute(data_stmt)
    rows = r2.scalars().all()

    records = []
    for row in rows:
        records.append({
            "id": row.id,
            "district": row.district,
            "date": row.date,
            "olt_name": row.olt_name,
            "olt_ip": row.olt_ip,
            "pon_port": row.pon_port,
            "onu_id": row.onu_id,
            "rx_power_dbm": row.rx_power_dbm,
            "community": row.community,
            "account_bandwidth": row.account_bandwidth,
        })

    return {
        "records": records,
        "total": total,
        "page": page,
        "page_size": page_size,
    }


# ════════════════════════════════════════════════════════════════
# ── 家宽重投2次清单 解析 ──
# ════════════════════════════════════════════════════════════════

def _parse_broadband_redelivery2_files(latest_file: str, filename: str) -> dict:
    """解析家宽重投2次清单「12-15-18点通报」sheet，提取横山最新时点8项指标"""
    import openpyxl

    wb = openpyxl.load_workbook(latest_file, read_only=True, data_only=True)

    if "12-15-18点通报" not in wb.sheetnames:
        wb.close()
        raise ValueError("未找到 sheet '12-15-18点通报'")

    ws = wb["12-15-18点通报"]
    rows_data = list(ws.iter_rows(values_only=True))
    wb.close()

    logger.info(f"家宽重投2次: 12-15-18点通报 sheet 共 {len(rows_data)} 行")

    # ── 解析sheet结构 ──
    # 每个日通报块包含 4 个时间段: 12:00, 15:00, 16:30, 18:00
    # 每时间段 12 列: 县区|重投2次在途量|2次全球通量|重投3次|3次全球通量|重投4次及以上|4次全球通量|总在途重投量|全球通总积压|重投2次处理量|闭环情况|-

    # 找出所有日通报标题行（含"闭环情况"）
    day_blocks = []  # [title_row_idx, ...]
    for idx, row in enumerate(rows_data):
        if not row:
            continue
        row_text = " ".join(str(c) for c in row if c)
        if "闭环情况" in row_text and "截至" in row_text:
            day_blocks.append(idx)

    if not day_blocks:
        raise ValueError("未找到日通报数据块（含'闭环情况'的标题行）")

    logger.info(f"家宽重投2次: 找到 {len(day_blocks)} 个日通报块, 标题行索引: {day_blocks}")

    # 取最后一个日通报块
    last_block_title_row_idx = day_blocks[-1]
    title_row = rows_data[last_block_title_row_idx]

    # 统计标题行中的列分隔符("-")数量 → 时间段数
    dash_count = sum(1 for c in title_row if c is not None and str(c).strip() == "-")
    periods_count = max(dash_count, 1)
    if periods_count > 4:
        periods_count = 4  # 最多4个时间段
    if periods_count < 1:
        periods_count = 4  # 默认

    logger.info(f"家宽重投2次: 最新日块有 {periods_count} 个时间段")

    BLOCK_COLS = 12  # 每个时间段占12列

    # 找到该块的表头行（紧接标题行之后，含"重投2次在途量"）
    header_row_idx = None
    for offset in range(1, 20):
        check_idx = last_block_title_row_idx + offset
        if check_idx >= len(rows_data):
            break
        check_row = rows_data[check_idx]
        if not check_row:
            continue
        check_text = " ".join(str(c) for c in check_row if c)
        if "重投2次在途量" in check_text:
            header_row_idx = check_idx
            break

    if header_row_idx is None:
        raise ValueError("未找到表头行（含'重投2次在途量'）")

    # 找到该块的数据结束行（"全市"行或遇到空行/下一块标题）
    city_row_idx = None
    for offset in range(1, 25):
        check_idx = header_row_idx + offset
        if check_idx >= len(rows_data):
            break
        check_row = rows_data[check_idx]
        if not check_row:
            city_row_idx = check_idx - 1
            break
        first_val = str(check_row[0]).strip() if check_row[0] else ""
        if "全市" in first_val:
            city_row_idx = check_idx
            break

    if city_row_idx is None:
        city_row_idx = min(header_row_idx + 20, len(rows_data) - 1)

    logger.info(f"家宽重投2次: 表头行={header_row_idx}, 全市行={city_row_idx}")

    # ── 取最后一个时间段（最新时点）的数据 ──
    last_period_start_col = (periods_count - 1) * BLOCK_COLS

    # 指标列在时间段内的偏移
    METRIC_COL_OFFSETS = {
        "redelivery2_in_transit": 1,   # 重投2次在途量
        "global_tong_2": 2,            # 2次全球通量
        "redelivery3": 3,              # 重投3次
        "global_tong_3": 4,            # 3次全球通量
        "redelivery4_plus": 5,         # 重投4次及以上
        "global_tong_4": 6,            # 4次全球通量
        "total_in_transit": 7,         # 总在途重投量
        "redelivery2_processed": 9,    # 重投2次处理量
    }

    def safe_int_str(val):
        """安全转整数字符串"""
        if val is None:
            return "0"
        if isinstance(val, (int, float)):
            if val == int(val):
                return str(int(val))
            return str(val).strip()
        return str(val).strip()

    hengshan_data = None
    import re
    title_text = str(title_row[0]) if title_row[0] else ""
    date_match = re.search(r'截至(\d+)日', title_text)
    report_date_from_title = ""
    if date_match:
        day_num = date_match.group(1)
        month_match = re.search(r'(\d+)月(\d+)日', filename)
        if month_match:
            report_date_from_title = f"2026-{int(month_match.group(1)):02d}-{int(day_num):02d}"
        else:
            report_date_from_title = f"2026-06-{int(day_num):02d}"

    # 在数据行中找横山
    for offset in range(1, city_row_idx - header_row_idx + 1):
        data_idx = header_row_idx + offset
        if data_idx >= len(rows_data):
            break
        row = rows_data[data_idx]
        if not row:
            continue

        name_col = last_period_start_col
        if name_col >= len(row):
            continue
        name_val = str(row[name_col]).strip() if row[name_col] else ""

        if "横山" in name_val:
            hengshan_data = {}
            for key, offset_in_block in METRIC_COL_OFFSETS.items():
                col = last_period_start_col + offset_in_block
                if col < len(row):
                    hengshan_data[key] = safe_int_str(row[col])
                else:
                    hengshan_data[key] = "0"
            break

    if not hengshan_data:
        raise ValueError(f"未找到横山数据（日期: {report_date_from_title}）")

    return {
        "summary": hengshan_data,
        "report_date": report_date_from_title,
        "filename": filename,
        "time_period": "18:00",
    }


def _parse_broadband_redelivery2_from_dir(directory: str) -> dict:
    """从目录中查找最新家宽重投2次清单文件并解析"""
    candidates = []
    for root, _dirs, files in os.walk(directory):
        for f in files:
            if "家宽重投2次清单" in f and not f.startswith("~$"):
                fp = os.path.join(root, f)
                candidates.append((os.path.getmtime(fp), fp, f))

    if not candidates:
        return {"error": "未找到含'家宽重投2次清单'的报表文件"}

    candidates.sort(reverse=True)
    latest_file = candidates[0][1]
    filename = candidates[0][2]

    logger.info(f"家宽重投2次 reparse: 使用文件 {filename}")

    return _parse_broadband_redelivery2_files(latest_file, filename)


async def reparse_broadband_redelivery2(db: AsyncSession, directory: Optional[str] = None) -> dict:
    """重新解析家宽重投2次清单，提取横山最新8项指标"""
    from app.core.models import BroadbandRedelivery2Summary, ReportFile, ReportType

    if directory is None:
        directory = settings.watch_dir

    # 在线程池中解析文件
    result = await asyncio.to_thread(_parse_broadband_redelivery2_from_dir, directory)

    if "error" in result:
        return result
    summary_data = result["summary"]
    filename = result.get("filename", "")

    # 删除旧数据
    stmt = delete(BroadbandRedelivery2Summary)
    await db.execute(stmt)

    # 创建新记录
    new_summary = BroadbandRedelivery2Summary(
        report_date=result["report_date"],
        district="横山",
        latest_filename=filename,
        time_period=result["time_period"],
        redelivery2_in_transit=summary_data.get("redelivery2_in_transit", ""),
        global_tong_2=summary_data.get("global_tong_2", ""),
        redelivery3=summary_data.get("redelivery3", ""),
        global_tong_3=summary_data.get("global_tong_3", ""),
        redelivery4_plus=summary_data.get("redelivery4_plus", ""),
        global_tong_4=summary_data.get("global_tong_4", ""),
        total_in_transit=summary_data.get("total_in_transit", ""),
        redelivery2_processed=summary_data.get("redelivery2_processed", ""),
    )
    db.add(new_summary)

    # 更新或创建 ReportFile 记录（确保 ReportType 存在）
    rt_result = await db.execute(
        select(ReportType).where(ReportType.name == "家宽重投2次清单明细")
    )
    rt = rt_result.scalar_one_or_none()
    if not rt:
        rt = ReportType(name="家宽重投2次清单明细", category="")
        db.add(rt)
        await db.flush()
    if rt:
        r4_result = await db.execute(
            select(ReportFile).where(ReportFile.report_type_id == rt.id)
            .order_by(ReportFile.created_at.desc()).limit(1)
        )
        existing = r4_result.scalar_one_or_none()
        if existing:
            existing.filename = filename
            existing.created_at = datetime.now(timezone.utc)
        else:
            db.add(ReportFile(
                report_type_id=rt.id,
                filename=filename,
                created_at=datetime.now(timezone.utc),
            ))

    await db.commit()

    return {
        "summary_parsed": 1,
        "filename": filename,
        "report_date": result["report_date"],
        "time_period": result["time_period"],
    }


async def get_broadband_redelivery2_summary(db: AsyncSession) -> dict:
    """获取家宽重投2次横山最新8项卡片指标"""
    from app.core.models import BroadbandRedelivery2Summary

    result = await db.execute(
        select(BroadbandRedelivery2Summary)
        .order_by(BroadbandRedelivery2Summary.created_at.desc())
        .limit(1)
    )
    row = result.scalar_one_or_none()

    default = {
        "district": "横山",
        "redelivery2_in_transit": "—",
        "global_tong_2": "—",
        "redelivery3": "—",
        "global_tong_3": "—",
        "redelivery4_plus": "—",
        "global_tong_4": "—",
        "total_in_transit": "—",
        "redelivery2_processed": "—",
        "report_date": "",
        "time_period": "",
        "latest_filename": "",
    }

    if not row:
        return default

    return {
        "district": row.district or "横山",
        "redelivery2_in_transit": row.redelivery2_in_transit,
        "global_tong_2": row.global_tong_2,
        "redelivery3": row.redelivery3,
        "global_tong_3": row.global_tong_3,
        "redelivery4_plus": row.redelivery4_plus,
        "global_tong_4": row.global_tong_4,
        "total_in_transit": row.total_in_transit,
        "redelivery2_processed": row.redelivery2_processed,
        "report_date": row.report_date,
        "time_period": row.time_period,
        "latest_filename": row.latest_filename,
    }
